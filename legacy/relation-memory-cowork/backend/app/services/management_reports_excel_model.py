from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import get_column_letter, range_boundaries

from app.services.business_metric_model import (
    BusinessMetricModelBundle,
    BusinessMetricModelError,
    MetricDefinition,
    ModelMetadata,
    ObservationRecord,
    RelationDefinition,
    RelationInputBinding,
    SourceReference,
)


CYRILLIC_TO_LATIN = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}


class UnsupportedExcelFormula(BusinessMetricModelError):
    pass


@dataclass(frozen=True)
class ColumnContext:
    index: int
    letter: str
    key: str
    header_label: str


@dataclass(frozen=True)
class MetricRowContext:
    row_index: int
    metric_code: str
    label: str
    unit: str
    kind: str
    category: str


@dataclass
class SheetContext:
    workbook_name: str
    sheet_name: str
    report_period: str
    label_col: int
    unit_col: int | None
    metric_rows: dict[int, MetricRowContext] = field(default_factory=dict)
    columns: dict[int, ColumnContext] = field(default_factory=dict)

    def slice_for_column(self, col_index: int) -> dict[str, str]:
        column = self.columns[col_index]
        return {
            "workbook": self.workbook_name,
            "sheet": self.sheet_name,
            "column": column.key,
        }


@dataclass
class ParsedToken:
    expression: str
    scalar_value: Any = None
    reference: str | None = None
    range_reference: str | None = None


def _normalize_text(value: Any) -> str:
    text = str(value or "").replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text.strip(" :;")


def _transliterate(text: str) -> str:
    result: list[str] = []
    for char in text.lower():
        result.append(CYRILLIC_TO_LATIN.get(char, char))
    return "".join(result)


def _slugify(text: str, max_length: int = 80) -> str:
    base = _transliterate(text)
    base = re.sub(r"[^a-z0-9]+", "_", base).strip("_")
    if not base:
        return "value"
    return base[:max_length]


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _to_period(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    raise BusinessMetricModelError(f"Unsupported report period value: {value!r}")


def _safe_float(value: Any) -> float | None:
    if not _is_number(value):
        return None
    return float(value)


def _infer_category(label: str) -> str:
    lowered = label.lower()
    if "выруч" in lowered:
        return "revenue"
    if "маржин" in lowered or "прибыл" in lowered:
        return "profit"
    if "рентабель" in lowered:
        return "ratio"
    if "затрат" in lowered or "себестоим" in lowered or "фот" in lowered or "амортиз" in lowered:
        return "cost"
    if "км" in lowered or "пробег" in lowered:
        return "distance"
    if "наклад" in lowered or "заяв" in lowered:
        return "volume"
    if "колич" in lowered or "водител" in lowered or "тс" in lowered:
        return "capacity"
    return "other"


class ExcelFormulaTranslator:
    PRECEDENCE = {
        "=": 1,
        "<>": 1,
        "<": 1,
        "<=": 1,
        ">": 1,
        ">=": 1,
        "+": 2,
        "-": 2,
        "*": 3,
        "/": 3,
        "^": 4,
    }

    def __init__(
        self,
        *,
        workbook_formula,
        workbook_values,
        sheet_contexts: dict[str, SheetContext],
        target_slice: dict[str, str],
        current_sheet: str,
    ):
        self.workbook_formula = workbook_formula
        self.workbook_values = workbook_values
        self.sheet_contexts = sheet_contexts
        self.target_slice = target_slice
        self.current_sheet = current_sheet
        self.tokens = []
        self.position = 0
        self.inputs: dict[str, RelationInputBinding] = {}
        self.alias_by_binding: dict[tuple[str, str], str] = {}

    def translate(self, formula: str) -> tuple[str, dict[str, RelationInputBinding]]:
        self.tokens = [item for item in Tokenizer(formula).items]
        self.position = 0
        self.inputs = {}
        self.alias_by_binding = {}
        parsed = self._parse_expression()
        if self._peek() is not None:
            raise UnsupportedExcelFormula(f"Unable to parse the full Excel formula {formula!r}.")
        used_aliases = {
            alias
            for alias in re.findall(r"\b(v\d+)\b", parsed.expression)
            if alias in self.inputs
        }
        filtered_inputs = {
            alias: binding
            for alias, binding in self.inputs.items()
            if alias in used_aliases
        }
        return parsed.expression, filtered_inputs

    def _peek(self):
        if self.position >= len(self.tokens):
            return None
        return self.tokens[self.position]

    def _consume(self):
        token = self._peek()
        if token is None:
            raise UnsupportedExcelFormula("Unexpected end of formula.")
        self.position += 1
        return token

    def _parse_expression(self, min_precedence: int = 0) -> ParsedToken:
        left = self._parse_prefix()
        while True:
            token = self._peek()
            if token is None or token.type != "OPERATOR-INFIX":
                break
            precedence = self.PRECEDENCE.get(token.value)
            if precedence is None or precedence < min_precedence:
                break
            operator_token = self._consume()
            next_precedence = precedence if operator_token.value == "^" else precedence + 1
            right = self._parse_expression(next_precedence)
            operator_value = "==" if operator_token.value == "=" else "!=" if operator_token.value == "<>" else operator_token.value
            left = ParsedToken(expression=f"({left.expression} {operator_value} {right.expression})")
        return left

    def _parse_prefix(self) -> ParsedToken:
        token = self._peek()
        if token is None:
            raise UnsupportedExcelFormula("Unexpected end of formula.")
        if token.type == "OPERATOR-PREFIX":
            operator_token = self._consume()
            operand = self._parse_expression(self.PRECEDENCE["^"])
            return ParsedToken(expression=f"({operator_token.value}{operand.expression})")
        return self._parse_primary()

    def _parse_primary(self) -> ParsedToken:
        token = self._consume()
        if token.type == "OPERAND":
            if token.subtype == "NUMBER":
                return ParsedToken(expression=str(float(token.value)), scalar_value=float(token.value))
            if token.subtype == "TEXT":
                text_value = token.value[1:-1] if token.value.startswith('"') and token.value.endswith('"') else token.value
                return ParsedToken(expression=json.dumps(text_value, ensure_ascii=False), scalar_value=text_value)
            if token.subtype == "RANGE":
                return self._parse_range_operand(token.value)
        if token.type == "PAREN" and token.subtype == "OPEN":
            inner = self._parse_expression()
            closing = self._consume()
            if closing.type != "PAREN" or closing.subtype != "CLOSE":
                raise UnsupportedExcelFormula("Expected a closing parenthesis.")
            return ParsedToken(expression=f"({inner.expression})")
        if token.type == "FUNC" and token.subtype == "OPEN":
            function_name = token.value[:-1]
            args = self._parse_function_args()
            return self._render_function(function_name, args)
        raise UnsupportedExcelFormula(f"Unsupported Excel token {token.value!r}.")

    def _parse_function_args(self) -> list[ParsedToken]:
        args: list[ParsedToken] = []
        if self._peek() is not None and self._peek().type == "FUNC" and self._peek().subtype == "CLOSE":
            self._consume()
            return args
        while True:
            args.append(self._parse_expression())
            token = self._consume()
            if token.type == "FUNC" and token.subtype == "CLOSE":
                break
            if token.type != "SEP" or token.subtype != "ARG":
                raise UnsupportedExcelFormula("Expected a function argument separator.")
        return args

    def _parse_range_operand(self, reference: str) -> ParsedToken:
        sheet_name, coordinate = self._split_sheet_reference(reference)
        if ":" in coordinate:
            return ParsedToken(
                expression="0",
                range_reference=f"{sheet_name}!{coordinate}",
            )
        expression = self._resolve_cell_expression(sheet_name, coordinate)
        return ParsedToken(
            expression=expression,
            reference=f"{sheet_name}!{coordinate}",
        )

    def _render_function(self, function_name: str, args: list[ParsedToken]) -> ParsedToken:
        normalized = function_name.upper()
        if normalized == "SUM":
            terms: list[str] = []
            for item in args:
                terms.extend(self._expand_argument_terms(item))
            return ParsedToken(expression=self._sum_terms(terms))
        if normalized == "ABS":
            return ParsedToken(expression=f"abs({args[0].expression})")
        if normalized == "ROUND":
            return ParsedToken(expression=f"round({args[0].expression}, {self._coerce_numeric_literal(args[1])})")
        if normalized == "IFERROR":
            fallback = args[1].expression
            if isinstance(args[1].scalar_value, str):
                fallback = "0"
            return ParsedToken(expression=f"iferror({args[0].expression}, {fallback})")
        if normalized == "IF":
            return ParsedToken(expression=f"if_({args[0].expression}, {args[1].expression}, {args[2].expression})")
        if normalized == "VLOOKUP":
            return ParsedToken(expression=self._resolve_vlookup(args))
        if normalized == "SUMIFS":
            terms = self._resolve_sumifs(args)
            return ParsedToken(expression=self._sum_terms(terms))
        if normalized in {"CONCATENATE", "TEXT", "EDATE"}:
            raise UnsupportedExcelFormula(f"Text/date Excel function {function_name} is not quantitative.")
        raise UnsupportedExcelFormula(f"Unsupported Excel function {function_name}.")

    def _resolve_vlookup(self, args: list[ParsedToken]) -> str:
        if len(args) < 3:
            raise UnsupportedExcelFormula("VLOOKUP expects at least three arguments.")
        lookup_value = args[0].scalar_value
        table_reference = args[1].range_reference or args[1].reference
        if table_reference is None:
            raise UnsupportedExcelFormula("VLOOKUP table reference was not preserved.")
        column_index = int(self._coerce_numeric_literal(args[2]))
        exact_match = True
        if len(args) >= 4:
            exact_match = bool(self._coerce_numeric_literal(args[3])) is False
        sheet_name, coordinate = self._split_sheet_reference(table_reference)
        min_col, min_row, max_col, max_row = range_boundaries(coordinate.replace("$", ""))
        sheet = self.workbook_values[sheet_name]
        matched_row = None
        for row_index in range(min_row, max_row + 1):
            candidate = sheet.cell(row_index, min_col).value
            if self._lookup_matches(candidate, lookup_value, exact_match=exact_match):
                matched_row = row_index
                break
        if matched_row is None:
            return "0"
        target_col = min_col + column_index - 1
        return self._resolve_cell_expression(sheet_name, f"{get_column_letter(target_col)}{matched_row}")

    def _resolve_sumifs(self, args: list[ParsedToken]) -> list[str]:
        if len(args) < 3 or len(args) % 2 == 0:
            raise UnsupportedExcelFormula("SUMIFS expects one sum range and at least one criteria pair.")
        sum_reference = args[0].range_reference or args[0].reference
        if sum_reference is None:
            raise UnsupportedExcelFormula("SUMIFS sum range reference was not preserved.")
        sum_sheet_name, sum_coordinate = self._split_sheet_reference(sum_reference)
        sum_cells = self._iter_range_cells(sum_sheet_name, sum_coordinate)

        criteria_pairs: list[tuple[list[tuple[str, str]], Any]] = []
        for index in range(1, len(args), 2):
            range_reference = args[index].range_reference or args[index].reference
            if range_reference is None:
                raise UnsupportedExcelFormula("SUMIFS criteria range reference was not preserved.")
            criteria_sheet_name, criteria_coordinate = self._split_sheet_reference(range_reference)
            criteria_pairs.append(
                (
                    self._iter_range_cells(criteria_sheet_name, criteria_coordinate),
                    args[index + 1].scalar_value,
                )
            )

        terms: list[str] = []
        for sum_position, (sheet_name, coordinate) in enumerate(sum_cells):
            matched = True
            for criteria_cells, criterion in criteria_pairs:
                criteria_sheet_name, criteria_coordinate = criteria_cells[sum_position]
                criteria_value = self.workbook_values[criteria_sheet_name][criteria_coordinate].value
                if not self._criterion_matches(criteria_value, criterion):
                    matched = False
                    break
            if matched:
                terms.append(self._resolve_cell_expression(sheet_name, coordinate))
        return terms

    def _criterion_matches(self, cell_value: Any, criterion: Any) -> bool:
        if isinstance(criterion, str):
            stripped = criterion.strip()
            operator_match = re.match(r"^(<=|>=|<>|=|<|>)(.*)$", stripped)
            if operator_match:
                operator_text, right_text = operator_match.groups()
                right_value: Any = right_text.strip()
                if re.fullmatch(r"-?\d+(\.\d+)?", right_value or ""):
                    right_value = float(right_value)
                if operator_text == "=":
                    return cell_value == right_value
                if operator_text == "<>":
                    return cell_value != right_value
                left_number = _safe_float(cell_value)
                right_number = _safe_float(right_value)
                if left_number is None or right_number is None:
                    return False
                if operator_text == "<":
                    return left_number < right_number
                if operator_text == "<=":
                    return left_number <= right_number
                if operator_text == ">":
                    return left_number > right_number
                if operator_text == ">=":
                    return left_number >= right_number
            return _normalize_text(cell_value) == stripped
        if _is_number(criterion):
            left_number = _safe_float(cell_value)
            return left_number is not None and left_number == float(criterion)
        return cell_value == criterion

    def _lookup_matches(self, candidate: Any, lookup_value: Any, *, exact_match: bool) -> bool:
        if exact_match:
            return _normalize_text(candidate) == _normalize_text(lookup_value)
        if _is_number(candidate) and _is_number(lookup_value):
            return float(candidate) >= float(lookup_value)
        return _normalize_text(candidate) >= _normalize_text(lookup_value)

    def _expand_argument_terms(self, item: ParsedToken) -> list[str]:
        if item.range_reference is not None:
            sheet_name, coordinate = self._split_sheet_reference(item.range_reference)
            return self._expand_range_references(sheet_name, coordinate)
        return [item.expression]

    def _expand_range_references(self, sheet_name: str, coordinate: str) -> list[str]:
        return [
            self._resolve_cell_expression(sheet_name, cell_coordinate)
            for _, cell_coordinate in self._iter_range_cells(sheet_name, coordinate)
        ]

    def _iter_range_cells(self, sheet_name: str, coordinate: str) -> list[tuple[str, str]]:
        min_col, min_row, max_col, max_row = range_boundaries(coordinate.replace("$", ""))
        cells: list[tuple[str, str]] = []
        for row_index in range(min_row, max_row + 1):
            for col_index in range(min_col, max_col + 1):
                cells.append((sheet_name, f"{get_column_letter(col_index)}{row_index}"))
        return cells

    def _resolve_cell_expression(self, sheet_name: str, coordinate: str) -> str:
        clean_coordinate = coordinate.replace("$", "")
        sheet_context = self.sheet_contexts[sheet_name]
        row_match = re.match(r"([A-Z]+)(\d+)$", clean_coordinate)
        if row_match is None:
            raise UnsupportedExcelFormula(f"Unsupported cell coordinate {coordinate!r}.")
        col_letters, row_number_text = row_match.groups()
        row_index = int(row_number_text)
        col_index = self._column_index_from_letters(col_letters)
        metric_context = sheet_context.metric_rows.get(row_index)
        if metric_context is not None and col_index in sheet_context.columns:
            source_slice = sheet_context.slice_for_column(col_index)
            return self._bind_metric(metric_context.metric_code, source_slice)
        value = self.workbook_values[sheet_name][clean_coordinate].value
        if _is_number(value):
            return str(float(value))
        if value in (None, "", " "):
            return "0"
        raise UnsupportedExcelFormula(
            f"Cell {sheet_name}!{clean_coordinate} resolved to unsupported non-numeric value {value!r}."
        )

    def _bind_metric(self, metric_code: str, source_slice: dict[str, str]) -> str:
        binding_key = (metric_code, json.dumps(source_slice, ensure_ascii=False, sort_keys=True))
        existing_alias = self.alias_by_binding.get(binding_key)
        if existing_alias is not None:
            return existing_alias
        alias = f"v{len(self.inputs) + 1}"
        slice_overrides = {
            key: value
            for key, value in source_slice.items()
            if self.target_slice.get(key) != value
        }
        self.inputs[alias] = RelationInputBinding(
            metric_code=metric_code,
            slice_overrides=slice_overrides,
        )
        self.alias_by_binding[binding_key] = alias
        return alias

    def _split_sheet_reference(self, reference: str) -> tuple[str, str]:
        raw = reference.replace("$", "")
        if "!" not in raw:
            return self.current_sheet, raw
        sheet_name, coordinate = raw.rsplit("!", 1)
        if sheet_name.startswith("'") and sheet_name.endswith("'"):
            sheet_name = sheet_name[1:-1].replace("''", "'")
        if "]" in sheet_name:
            sheet_name = sheet_name.split("]", 1)[1]
        return sheet_name, coordinate

    def _sum_terms(self, expressions: list[str]) -> str:
        clean = [item for item in expressions if item]
        if not clean:
            return "0"
        if len(clean) == 1:
            return clean[0]
        return f"sum_({', '.join(clean)})"

    def _column_index_from_letters(self, letters: str) -> int:
        value = 0
        for char in letters:
            value = value * 26 + (ord(char.upper()) - 64)
        return value

    def _coerce_numeric_literal(self, item: ParsedToken) -> float:
        if _is_number(item.scalar_value):
            return float(item.scalar_value)
        if isinstance(item.scalar_value, str) and re.fullmatch(r"-?\d+(\.\d+)?", item.scalar_value.strip()):
            return float(item.scalar_value.strip())
        try:
            return float(item.expression)
        except ValueError as exc:
            raise UnsupportedExcelFormula(
                f"Expected a numeric literal in Excel formula but received {item.expression!r}."
            ) from exc


class ManagementReportsExcelModelBuilder:
    def __init__(self, excel_dir: Path):
        self.excel_dir = Path(excel_dir)

    def build_bundle(
        self,
        *,
        model_key: str = "management_reports_excel_full_model",
        model_name: str = "Management Reports Excel Full Model",
        version: str = "1",
    ) -> BusinessMetricModelBundle:
        workbook_paths = sorted(self.excel_dir.glob("*.xlsx"))
        if not workbook_paths:
            raise BusinessMetricModelError(f"No Excel workbooks were found under {self.excel_dir}.")

        metrics: list[MetricDefinition] = []
        relations: list[RelationDefinition] = []
        observations: list[ObservationRecord] = []

        for workbook_path in workbook_paths:
            workbook_formula = load_workbook(workbook_path, data_only=False, read_only=False)
            workbook_values = load_workbook(workbook_path, data_only=True, read_only=False)
            sheet_contexts: dict[str, SheetContext] = {}
            for sheet_name in workbook_formula.sheetnames:
                sheet_context = self._analyze_sheet(
                    workbook_name=workbook_path.name,
                    formula_sheet=workbook_formula[sheet_name],
                    value_sheet=workbook_values[sheet_name],
                )
                if sheet_context is None:
                    continue
                sheet_contexts[sheet_name] = sheet_context

            if not sheet_contexts:
                workbook_formula.close()
                workbook_values.close()
                continue

            metrics.extend(self._build_metric_definitions(sheet_contexts))
            workbook_observations, workbook_relations = self._build_records_for_workbook(
                workbook_path=workbook_path,
                workbook_formula=workbook_formula,
                workbook_values=workbook_values,
                sheet_contexts=sheet_contexts,
            )
            observations.extend(workbook_observations)
            relations.extend(workbook_relations)
            workbook_formula.close()
            workbook_values.close()

        metrics.sort(key=lambda item: item.code)
        relations.sort(key=lambda item: item.id)
        observations.sort(key=lambda item: (item.metric_code, item.period, item.scenario, json.dumps(item.slice, ensure_ascii=False, sort_keys=True)))

        return BusinessMetricModelBundle(
            model=ModelMetadata(
                key=model_key,
                name=model_name,
                version=version,
                description="Excel-derived management report model with row metrics, spreadsheet relations, and cell observations.",
            ),
            metrics=metrics,
            relations=relations,
            observations=observations,
        )

    def _build_metric_definitions(self, sheet_contexts: dict[str, SheetContext]) -> list[MetricDefinition]:
        metrics: list[MetricDefinition] = []
        for sheet_context in sheet_contexts.values():
            for metric_context in sheet_context.metric_rows.values():
                metrics.append(
                    MetricDefinition(
                        code=metric_context.metric_code,
                        name=metric_context.label,
                        description=(
                            f"Excel row metric parsed from {sheet_context.workbook_name} / "
                            f"{sheet_context.sheet_name} / row {metric_context.row_index}."
                        ),
                        unit=metric_context.unit,
                        kind=metric_context.kind,
                        category=metric_context.category,
                        scope_tags=[sheet_context.workbook_name, sheet_context.sheet_name],
                        aliases=[
                            metric_context.label,
                            f"{sheet_context.sheet_name}: {metric_context.label}",
                        ],
                        source_refs=[
                            SourceReference(
                                kind="workbook_row",
                                workbook=sheet_context.workbook_name,
                                sheet=sheet_context.sheet_name,
                                row_label=metric_context.label,
                                note=f"row={metric_context.row_index}",
                            )
                        ],
                    )
                )
        return metrics

    def _build_records_for_workbook(
        self,
        *,
        workbook_path: Path,
        workbook_formula,
        workbook_values,
        sheet_contexts: dict[str, SheetContext],
    ) -> tuple[list[ObservationRecord], list[RelationDefinition]]:
        observations: list[ObservationRecord] = []
        relations: list[RelationDefinition] = []
        for sheet_name, sheet_context in sheet_contexts.items():
            formula_sheet = workbook_formula[sheet_name]
            value_sheet = workbook_values[sheet_name]
            for row_index, metric_context in sheet_context.metric_rows.items():
                for col_index, column_context in sheet_context.columns.items():
                    slice_filters = sheet_context.slice_for_column(col_index)
                    value = value_sheet.cell(row_index, col_index).value
                    formula = formula_sheet.cell(row_index, col_index).value
                    if _is_number(value):
                        observations.append(
                            ObservationRecord(
                                metric_code=metric_context.metric_code,
                                period=sheet_context.report_period,
                                scenario="fact",
                                slice=slice_filters,
                                value=float(value),
                                source_refs=[
                                    SourceReference(
                                        kind="workbook_cell",
                                        workbook=workbook_path.name,
                                        sheet=sheet_name,
                                        row_label=metric_context.label,
                                        note=f"cell={column_context.letter}{row_index}",
                                    )
                                ],
                            )
                        )

                    if not isinstance(formula, str) or not formula.startswith("="):
                        continue
                    translator = ExcelFormulaTranslator(
                        workbook_formula=workbook_formula,
                        workbook_values=workbook_values,
                        sheet_contexts=sheet_contexts,
                        target_slice=slice_filters,
                        current_sheet=sheet_name,
                    )
                    try:
                        expression, inputs = translator.translate(formula)
                    except UnsupportedExcelFormula:
                        continue
                    if not inputs:
                        continue
                    relations.append(
                        RelationDefinition(
                            id=f"{metric_context.metric_code}__{column_context.key}",
                            name=f"{metric_context.label} [{sheet_name} {column_context.letter}]",
                            description=(
                                f"Spreadsheet formula parsed from {workbook_path.name} / "
                                f"{sheet_name} / {column_context.letter}{row_index}."
                            ),
                            target_metric_code=metric_context.metric_code,
                            relation_kind="formula",
                            output_mode="level",
                            expression=expression,
                            inputs=inputs,
                            applies_to=slice_filters,
                            confidence=1.0,
                            source_refs=[
                                SourceReference(
                                    kind="workbook_formula",
                                    workbook=workbook_path.name,
                                    sheet=sheet_name,
                                    row_label=metric_context.label,
                                    note=f"cell={column_context.letter}{row_index}; formula={formula}",
                                )
                            ],
                            tags=[sheet_name, column_context.key],
                        )
                    )
        return observations, relations

    def _analyze_sheet(self, *, workbook_name: str, formula_sheet, value_sheet) -> SheetContext | None:
        header = self._detect_header(formula_sheet, value_sheet)
        if header is None:
            return None
        header_row, label_col = header
        unit_col = self._detect_unit_column(value_sheet, header_row, label_col)
        report_period = self._detect_report_period(value_sheet, header_row, label_col)
        metric_rows = self._detect_metric_rows(
            workbook_name=workbook_name,
            sheet_name=value_sheet.title,
            formula_sheet=formula_sheet,
            value_sheet=value_sheet,
            label_col=label_col,
            unit_col=unit_col,
        )
        numeric_columns = self._detect_data_columns(
            formula_sheet=formula_sheet,
            value_sheet=value_sheet,
            metric_rows=metric_rows,
            label_col=label_col,
            header_row=header_row,
        )
        return SheetContext(
            workbook_name=workbook_name,
            sheet_name=value_sheet.title,
            report_period=report_period,
            label_col=label_col,
            unit_col=unit_col,
            metric_rows=metric_rows,
            columns=numeric_columns,
        )

    def _detect_header(self, formula_sheet, value_sheet) -> tuple[int, int] | None:
        candidates: list[tuple[int, int]] = []
        for row_index in range(1, min(value_sheet.max_row, 30) + 1):
            for col_index in range(1, min(value_sheet.max_column, 12) + 1):
                cell_value = _normalize_text(value_sheet.cell(row_index, col_index).value)
                if "показатель" in cell_value.lower():
                    candidates.append((row_index, col_index))
        if not candidates:
            return None
        return sorted(candidates)[0]

    def _detect_unit_column(self, value_sheet, header_row: int, label_col: int) -> int | None:
        for col_index in range(label_col + 1, min(value_sheet.max_column, label_col + 3) + 1):
            header_value = _normalize_text(value_sheet.cell(header_row, col_index).value).lower()
            if "ед.измер" in header_value:
                return col_index
        return None

    def _detect_report_period(self, value_sheet, header_row: int, label_col: int) -> str:
        for row_index in range(1, min(value_sheet.max_row, max(header_row, 10)) + 1):
            for col_index in range(label_col + 1, value_sheet.max_column + 1):
                value = value_sheet.cell(row_index, col_index).value
                if isinstance(value, (datetime, date)):
                    return _to_period(value)
        raise BusinessMetricModelError(f"Unable to detect a report period in sheet {value_sheet.title!r}.")

    def _detect_metric_rows(
        self,
        *,
        workbook_name: str,
        sheet_name: str,
        formula_sheet,
        value_sheet,
        label_col: int,
        unit_col: int | None,
    ) -> dict[int, MetricRowContext]:
        metric_rows: dict[int, MetricRowContext] = {}
        for row_index in range(1, value_sheet.max_row + 1):
            label = _normalize_text(value_sheet.cell(row_index, label_col).value)
            if not label or label.lower() == "показатель" or label.lower().startswith("в том числе"):
                continue
            if not self._row_has_metric_data(formula_sheet, value_sheet, row_index, label_col):
                continue
            raw_values = [
                formula_sheet.cell(row_index, col_index).value
                for col_index in range(label_col + 1, formula_sheet.max_column + 1)
            ]
            has_constant_numeric = any(_is_number(item) for item in raw_values)
            has_formula = any(isinstance(item, str) and item.startswith("=") for item in raw_values)
            metric_code = (
                f"{_slugify(workbook_name, 30)}__"
                f"{_slugify(sheet_name, 30)}__"
                f"r{row_index:03d}__{_slugify(label, 50)}"
            )
            unit = _normalize_text(value_sheet.cell(row_index, unit_col).value) if unit_col else ""
            metric_rows[row_index] = MetricRowContext(
                row_index=row_index,
                metric_code=metric_code,
                label=label,
                unit=unit,
                kind="observed" if has_constant_numeric else "derived" if has_formula else "observed",
                category=_infer_category(label),
            )
        return metric_rows

    def _row_has_metric_data(self, formula_sheet, value_sheet, row_index: int, label_col: int) -> bool:
        for col_index in range(label_col + 1, formula_sheet.max_column + 1):
            formula_value = formula_sheet.cell(row_index, col_index).value
            observed_value = value_sheet.cell(row_index, col_index).value
            if _is_number(observed_value) or (isinstance(formula_value, str) and formula_value.startswith("=")):
                return True
        return False

    def _detect_data_columns(
        self,
        *,
        formula_sheet,
        value_sheet,
        metric_rows: dict[int, MetricRowContext],
        label_col: int,
        header_row: int,
    ) -> dict[int, ColumnContext]:
        columns: dict[int, ColumnContext] = {}
        for col_index in range(label_col + 1, value_sheet.max_column + 1):
            has_data = False
            for row_index in metric_rows:
                observed_value = value_sheet.cell(row_index, col_index).value
                formula_value = formula_sheet.cell(row_index, col_index).value
                if _is_number(observed_value) or (isinstance(formula_value, str) and formula_value.startswith("=")):
                    has_data = True
                    break
            if not has_data:
                continue
            letter = get_column_letter(col_index)
            header_parts = [
                _normalize_text(value_sheet.cell(row_index, col_index).value)
                for row_index in range(1, header_row + 1)
                if _normalize_text(value_sheet.cell(row_index, col_index).value)
            ]
            header_label = " | ".join(header_parts)
            columns[col_index] = ColumnContext(
                index=col_index,
                letter=letter,
                key=f"col_{letter.lower()}",
                header_label=header_label,
            )
        return columns
