from __future__ import annotations

import copy
import logging
import re
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Callable

import yaml
from openpyxl import Workbook, load_workbook

from app.config import settings
from app.schemas.relation_memory import (
    CandidateRelation,
    ConfirmationResponse,
    ConfirmedRelation,
    ConversationMessage,
    DocumentParseStatus,
    DocumentSummary,
    GraphQuestionResponse,
    MetricCandidate,
    MetricCandidateConfirmationResponse,
    RelationMemoryChatResponse,
    RelationMemorySessionResponse,
)
from app.services.relation_memory_agentic_enrichment import AgenticWorkbookEnrichmentResult, AgenticWorkbookRelationEnricher
from app.services.relation_memory_candidates import (
    candidate_id,
    candidate_relations_from_dependencies,
    candidate_relations_from_hypotheses,
    relation_key,
    relation_key_set,
)
from app.services.relation_memory_dynamic import (
    DomainPack,
    NormalizationAdapter,
    OllamaMetricSemanticResolver,
    OllamaRelationSemanticJudge,
    WorkbookProfiler,
)
from app.services.relation_memory_agent import (
    LlmJsonError,
    RelationMemoryAgent,
    RelationMemoryAnswerNarrator,
    RelationMemoryLlmClient,
    UserAnswerMemoryResult,
    normalize_candidate_payload,
)
from app.services.relation_memory_ingestion import (
    DocumentBundle,
    UploadedFilePayload,
    ingest_file_payloads,
    make_tempdir,
    normalize_metric_code,
    write_xlsx_payloads_to_tempdir,
)
from app.services.relation_memory_metric_map import (
    build_metric_map,
    contains_cyrillic,
    display_metric_label,
    ensure_external_context_metric_label,
    ensure_metric_entry,
    ensure_question_metric,
    extend_metric_values,
    metric_default_label,
    metric_label_for_chat,
    metric_response_items,
)
from app.services.relation_memory_neo4j import RelationMemoryNeo4jClient
from app.services.relation_memory_poc import (
    RelationMemoryPocBuilder,
    RelationMemorySnapshot,
    _canonical_header,
    _canonical_metric_code,
    _sheet_rows,
)
from app.services.relation_memory_question_models import (
    GraphQuestionAnswer,
    RUSSIAN_METRIC_ALIASES,
    RUSSIAN_METRIC_DISPLAY_LABELS,
)
from app.services.relation_memory_questions import RelationMemoryQuestionService
from app.services.relation_memory_user_language import RelationMemoryConversationContext, normalize_user_text
from app.services.relation_memory_workbook_shape import (
    DIMENSION_COLUMNS,
    ROW_REPORT_RELATION_SOURCE,
    column_has_numeric_values,
    normalize_row_metric_xlsx,
)


ALLOWED_CONFIRMATION_DECISIONS = {"approve", "reject"}
REPORT_RELATION_CANDIDATE_LIMIT = 20
DYNAMIC_RELATION_CANDIDATE_LIMIT = 25
DYNAMIC_RELATION_MIN_PENDING_SCORE = 0.55
AGENTIC_ENRICHMENT_MIN_RELATION_CANDIDATES = 2
LOGGER = logging.getLogger(__name__)


@dataclass
class ParsedWorkbookArtifact:
    document_id: str
    filename: str
    file_type: str
    parser: str
    status: str
    contracts: list[dict[str, Any]] = field(default_factory=list)
    metric_dictionary_rows: list[dict[str, Any]] = field(default_factory=list)
    dependency_rule_rows: list[dict[str, Any]] = field(default_factory=list)
    metric_candidates: list[MetricCandidate] = field(default_factory=list)
    relation_candidates: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    dataset_count: int = 0
    metric_count: int = 0
    dependency_count: int = 0


@dataclass
class RelationMemorySessionState:
    id: str
    bundle: DocumentBundle
    snapshot: RelationMemorySnapshot
    assistant_message: str
    metric_map: dict[str, dict[str, Any]]
    document_parse_statuses: list[DocumentParseStatus] = field(default_factory=list)
    parsed_document_ids: set[str] = field(default_factory=set)
    metric_candidates: dict[str, MetricCandidate] = field(default_factory=dict)
    rejected_metric_candidate_ids: set[str] = field(default_factory=set)
    pending_confirmations: dict[str, CandidateRelation] = field(default_factory=dict)
    rejected_candidate_ids: set[str] = field(default_factory=set)
    approved_relations: list[ConfirmedRelation] = field(default_factory=list)
    confirmed_relation_keys: set[tuple[str, str, str]] = field(default_factory=set)
    conversation_context: RelationMemoryConversationContext = field(default_factory=RelationMemoryConversationContext)
    open_agent_question: dict[str, Any] | None = None
    answered_agent_question_ids: set[str] = field(default_factory=set)
    memory_facts: list[dict[str, Any]] = field(default_factory=list)
    chat_history: list[ConversationMessage] = field(default_factory=list)
    last_debug_trace: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# Refactor target: after launch/test hardening, split this orchestration
# class into WorkbookArtifactParser, SessionStateBuilder, ChatRouter, and
# ConfirmationService behind characterization tests.
class RelationMemorySessionService:
    def __init__(
        self,
        *,
        graph_client: Any | None = None,
        agent: RelationMemoryAgent | None = None,
        agentic_enricher: AgenticWorkbookRelationEnricher | None = None,
    ):
        self.graph_client = graph_client
        self.agent = agent or RelationMemoryAgent()
        self.agentic_enricher = agentic_enricher or AgenticWorkbookRelationEnricher()
        self.question_service = RelationMemoryQuestionService(
            answer_narrator=RelationMemoryAnswerNarrator(
                enabled=settings.RELATION_MEMORY_LLM_NARRATOR_ENABLED,
            )
        )
        self.sessions: dict[str, RelationMemorySessionState] = {}

    def create_session(self, files: list[UploadedFilePayload]) -> RelationMemorySessionResponse:
        session_id = str(uuid.uuid4())
        bundle = ingest_file_payloads(files)
        state = self._build_session_state(session_id=session_id, bundle=bundle)
        self.sessions[session_id] = state
        return self._session_response(state)

    def add_files_to_session(self, session_id: str, files: list[UploadedFilePayload]) -> RelationMemorySessionResponse:
        previous_state = self._get_session(session_id)
        incoming_bundle = ingest_file_payloads(files)
        combined_bundle = self._merge_document_bundles(previous_state.bundle, incoming_bundle)
        rebuilt_state = self._build_session_state(
            session_id=session_id,
            bundle=combined_bundle,
            previous_state=previous_state,
        )
        self.sessions[session_id] = rebuilt_state
        return self._session_response(rebuilt_state)

    def get_session(self, session_id: str) -> RelationMemorySessionResponse:
        return self._session_response(self._get_session(session_id))

    def get_last_debug_trace(self, session_id: str) -> list[dict[str, Any]]:
        return list(self._get_session(session_id).last_debug_trace)

    def rebuild_session(self, session_id: str) -> RelationMemorySessionResponse:
        previous_state = self._get_session(session_id)
        rebuilt_state = self._build_session_state(
            session_id=session_id,
            bundle=previous_state.bundle,
            previous_state=previous_state,
        )
        self.sessions[session_id] = rebuilt_state
        return self._session_response(rebuilt_state)

    def handle_message(
        self,
        session_id: str,
        message: str,
        *,
        trace_sink: Callable[[dict[str, Any]], None] | None = None,
    ) -> tuple[str, list[CandidateRelation], list[str]]:
        state = self._get_session(session_id)
        state.last_debug_trace = []
        open_question_result = self._try_handle_open_agent_question_answer(state, message)
        if open_question_result is not None:
            assistant_message, created, warnings = open_question_result
            self._record_session_trace(
                state,
                trace_sink,
                stage="open_agent_question",
                title="OpenAgentQuestion",
                summary="Сообщение распознано как ответ на открытый вопрос агента; сохраняю только как candidate/memory fact.",
                details={
                    "created_candidates": [self._candidate_trace_payload(candidate) for candidate in created],
                    "warnings": warnings,
                    "assistant_message": assistant_message,
                },
            )
            return open_question_result
        if self._is_capability_help_question(message):
            self._record_session_trace(
                state,
                trace_sink,
                stage="chat_router",
                title="ChatRouter",
                summary="Распознал meta-вопрос о возможностях, graph-QA не запускался.",
                details={"route": "capability_help", "message": message},
            )
            return self._capability_help_answer(), [], []
        if self._is_relation_source_question(message):
            self._record_session_trace(
                state,
                trace_sink,
                stage="chat_router",
                title="ChatRouter",
                summary="Распознал meta-вопрос о происхождении связей, graph-QA не запускался.",
                details={"route": "relation_source_help", "message": message},
            )
            return self._fallback_open_question_answer(state, message), [], []
        graph_answer = self.question_service.answer(
            question=message,
            snapshot=self._question_snapshot(state),
            metric_map=self._question_metric_map(state),
            pending_confirmations=list(state.pending_confirmations.values()),
            conversation_context=state.conversation_context,
            trace_sink=trace_sink,
        )
        state.last_debug_trace = list(graph_answer.debug_trace)
        state.conversation_context = graph_answer.updated_context or state.conversation_context
        self._persist_confirmed_metric_aliases(
            state=state,
            aliases=graph_answer.confirmed_aliases,
            session_id=session_id,
        )
        if graph_answer.handled:
            created = self._create_external_context_candidates(state, graph_answer)
            return graph_answer.answer, created, list(graph_answer.warnings)
        if self.question_service.should_answer_as_question(
            message=message,
            conversation_context=state.conversation_context,
        ) or self._looks_like_question(message):
            self._record_session_trace(
                state,
                trace_sink,
                stage="open_session_question",
                title="FallbackSessionAnswer",
                summary="Relation Memory graph-QA не смог обработать вопрос; отвечаю по состоянию текущей сессии.",
                details={"message": message, "graph_intent": graph_answer.intent, "warnings": graph_answer.warnings},
            )
            return self._answer_open_session_question(state, message), [], []
        self._record_session_trace(
            state,
            trace_sink,
            stage="statement_router",
            title="StatementRouter",
            summary="Сообщение не похоже на вопрос; запускаю LLM-нормализацию новых фактов/связей.",
            details={"message": message},
        )
        result = self.agent.normalize_user_message(
            message=message,
            known_metrics=list(state.metric_map.values()),
            pending_confirmations=list(state.pending_confirmations.values()),
        )
        warnings = list(result.warnings)
        created: list[CandidateRelation] = []
        for raw_candidate in result.candidate_relations:
            candidate = self._candidate_from_raw(raw_candidate, state, source="llm_user_message")
            if candidate:
                stored = self._add_candidate_if_new(state, candidate)
                if stored is not None:
                    created.append(stored)
        assistant_message = result.assistant_message or (
            "Я выделил новые кандидатные связи и жду явного подтверждения по каждой."
            if created
            else "Не нашел новых связей в сообщении. Опишите процесс подробнее или подтвердите/отклоните конкретную связь."
        )
        self._record_session_trace(
            state,
            trace_sink,
            stage="statement_extraction",
            title="RelationNormalizer",
            summary=f"LLM-нормализация вернула {len(result.candidate_relations)} candidate relation(s); создано {len(created)}.",
            details={
                "raw_candidate_count": len(result.candidate_relations),
                "created_candidates": [self._candidate_trace_payload(candidate) for candidate in created],
                "warnings": warnings,
                "assistant_message": assistant_message,
            },
        )
        return assistant_message, created, warnings

    def _record_session_trace(
        self,
        state: RelationMemorySessionState,
        trace_sink: Callable[[dict[str, Any]], None] | None,
        *,
        stage: str,
        title: str,
        summary: str,
        details: dict[str, Any] | None = None,
    ) -> None:
        event = {
            "stage": stage,
            "title": title,
            "summary": summary,
            "details": details or {},
        }
        state.last_debug_trace.append(event)
        if trace_sink is not None:
            trace_sink(event)

    def _candidate_trace_payload(self, candidate: CandidateRelation) -> dict[str, Any]:
        return {
            "id": candidate.id,
            "source_metric_code": candidate.source_metric_code,
            "target_metric_code": candidate.target_metric_code,
            "edge_type": candidate.edge_type,
            "confidence": candidate.confidence,
            "score": candidate.score,
            "evidence_type": candidate.evidence_type,
            "source": candidate.source,
            "note": candidate.note,
            "evidence": candidate.evidence,
        }

    def _looks_like_question(self, message: str) -> bool:
        normalized = str(message or "").strip().lower()
        if not normalized:
            return False
        if "?" in normalized:
            return True
        return normalized.startswith(
            (
                "что ",
                "кто ",
                "где ",
                "когда ",
                "куда ",
                "как ",
                "какие ",
                "какая ",
                "какой ",
                "почему ",
                "зачем ",
                "сколько ",
                "можно ",
                "покажи ",
                "расскажи ",
                "объясни ",
                "объясните ",
                "сравни ",
                "сравните ",
                "найди ",
                "найдите ",
                "перечисли ",
                "перечислите ",
                "опиши ",
                "опишите ",
                "сделай обзор",
                "дай обзор",
                "дай ",
                "what ",
                "which ",
                "where ",
                "when ",
                "how ",
                "why ",
                "show ",
                "list ",
                "explain ",
                "compare ",
                "find ",
                "describe ",
                "summarize ",
                "tell ",
            )
        )

    def _try_handle_open_agent_question_answer(
        self,
        state: RelationMemorySessionState,
        message: str,
    ) -> tuple[str, list[CandidateRelation], list[str]] | None:
        open_question = state.open_agent_question
        if not open_question or not self._looks_like_answer_to_open_question(message):
            return None

        created: list[CandidateRelation] = []
        warnings: list[str] = []
        facts: list[dict[str, Any]] = []

        deterministic_candidate = (
            self._candidate_from_open_question_confirmation(state, open_question, message)
            or self._candidate_from_open_question_freeform_answer(state, open_question, message)
        )
        if deterministic_candidate is not None:
            stored = self._add_candidate_if_new(state, deterministic_candidate)
            if stored is not None:
                created.append(stored)

        if deterministic_candidate is not None:
            result = UserAnswerMemoryResult(
                assistant_message="",
                candidate_relations=[],
                memory_facts=[],
                warnings=[],
            )
        else:
            result = self.agent.interpret_user_answer(
                message=message,
                open_question=open_question,
                known_metrics=list(state.metric_map.values()),
                pending_confirmations=list(state.pending_confirmations.values()),
            )
        warnings.extend(result.warnings)
        for fact_payload in result.memory_facts:
            facts.append(self._memory_fact_from_user_answer(state, open_question, message, fact_payload))
        if not facts:
            facts.append(self._memory_fact_from_user_answer(state, open_question, message, {}))
        for fact in facts:
            for candidate in created:
                for metric_code in (candidate.source_metric_code, candidate.target_metric_code):
                    if metric_code and metric_code not in fact["metric_codes"]:
                        fact["metric_codes"].append(metric_code)

        for raw_candidate in result.candidate_relations:
            raw_candidate = dict(raw_candidate)
            raw_candidate["edge_type"] = self._coerce_edge_type(raw_candidate.get("edge_type"))
            raw_candidate.setdefault("source", "user_answer")
            ensure_external_context_metric_label(
                state.metric_map,
                str(raw_candidate.get("source_metric_code") or ""),
                str(raw_candidate.get("source_label") or ""),
            )
            ensure_external_context_metric_label(
                state.metric_map,
                str(raw_candidate.get("target_metric_code") or open_question.get("target_metric_code") or ""),
                str(raw_candidate.get("target_label") or ""),
            )
            if not raw_candidate.get("target_metric_code") and open_question.get("target_metric_code"):
                raw_candidate["target_metric_code"] = open_question.get("target_metric_code")
            candidate = self._candidate_from_raw(raw_candidate, state, source="user_answer")
            if candidate:
                stored = self._add_candidate_if_new(state, candidate)
                if stored is not None:
                    created.append(stored)

        state.memory_facts.extend(facts)
        question_id = str(open_question.get("id") or "")
        if question_id:
            state.answered_agent_question_ids.add(question_id)
        state.open_agent_question = self._next_agent_question(state)

        assistant_message = self._agent_answer_acknowledgement(
            state=state,
            llm_message=result.assistant_message,
            created=created,
            facts=facts,
            next_question=state.open_agent_question,
        )
        state.assistant_message = assistant_message
        return assistant_message, created, self._unique_warnings(warnings)

    def _looks_like_answer_to_open_question(self, message: str) -> bool:
        text = str(message or "").strip()
        if not text:
            return False
        if self._looks_like_question(text):
            return False
        normalized = text.lower()
        answer_prefixes = (
            "да",
            "нет",
            "скорее",
            "верно",
            "неверно",
            "подтверж",
            "не подтверж",
            "это ",
            "из-за",
            "из за",
            "связано",
            "влияет",
            "не влияет",
            "может",
            "обычно",
        )
        return normalized.startswith(answer_prefixes) or len(normalized.split()) >= 3

    def _candidate_from_open_question_confirmation(
        self,
        state: RelationMemorySessionState,
        open_question: dict[str, Any],
        message: str,
    ) -> CandidateRelation | None:
        if not self._is_affirmative_answer(message):
            return None
        hypothesis_ids = [str(item) for item in open_question.get("hypothesis_ids") or [] if str(item)]
        if not hypothesis_ids:
            return None
        hypotheses = {
            str(item.get("id") or ""): item
            for item in state.snapshot.hypotheses
            if item.get("id")
        }
        hypothesis = hypotheses.get(hypothesis_ids[0])
        if not hypothesis:
            return None
        source = str(hypothesis.get("source_metric_code") or "")
        target = str(hypothesis.get("target_metric_code") or open_question.get("target_metric_code") or "")
        if not source or not target:
            return None
        raw_candidate = {
            "source_metric_code": source,
            "target_metric_code": target,
            "edge_type": "driver",
            "note": (
                f"Пользователь подтвердил уточняющий вопрос: "
                f"{open_question.get('prompt') or ''} Ответ: {message}"
            ),
            "evidence": str(hypothesis.get("explanation") or message),
            "confidence": max(0.7, float(hypothesis.get("confidence") or 0.0)),
            "score": max(0.7, float(hypothesis.get("confidence") or 0.0)),
            "needs_approval_reason": "Подтверждено пользователем в ответе на вопрос агента; требуется финальное сохранение.",
            "source": "user_answer",
        }
        return self._candidate_from_raw(raw_candidate, state, source="user_answer")

    def _candidate_from_open_question_freeform_answer(
        self,
        state: RelationMemorySessionState,
        open_question: dict[str, Any],
        message: str,
    ) -> CandidateRelation | None:
        if str(open_question.get("question_type") or "") != "boundary_probe":
            return None
        target = str(open_question.get("target_metric_code") or "")
        if not target:
            return None
        source = self._source_metric_from_answer_phrase(state, message, target_metric_code=target)
        if not source:
            return None
        source_code, source_label = source
        if source_code == target:
            return None
        ensure_external_context_metric_label(state.metric_map, source_code, source_label)
        target_label = metric_label_for_chat(state.metric_map, target)
        raw_candidate = {
            "source_metric_code": source_code,
            "target_metric_code": target,
            "edge_type": "driver",
            "note": (
                f"Пользователь ответил на вопрос «{open_question.get('prompt') or ''}»: {message}"
            ),
            "evidence": str(message or "").strip(),
            "confidence": 0.72,
            "score": 0.72,
            "needs_approval_reason": "Извлечено из ответа пользователя на уточняющий вопрос агента.",
            "source": "user_answer",
        }
        ensure_external_context_metric_label(state.metric_map, target, target_label)
        return self._candidate_from_raw(raw_candidate, state, source="user_answer")

    def _source_metric_from_answer_phrase(
        self,
        state: RelationMemorySessionState,
        message: str,
        *,
        target_metric_code: str,
    ) -> tuple[str, str] | None:
        text = str(message or "").strip()
        normalized = text.lower()
        if "тариф" in normalized and ("платн" in normalized or "дорог" in normalized or "оператор" in normalized):
            return "toll_tariff_policy", "Тарифная политика операторов платных дорог"
        if "маршрут" in normalized and "платн" in normalized:
            return "paid_road_route_share", "Доля маршрутов по платным дорогам"
        if ("категор" in normalized or "класс" in normalized or "масса" in normalized) and (
            "транспорт" in normalized or "авто" in normalized
        ):
            return "vehicle_category_mix", "Структура транспорта по категориям"
        for code, metric in state.metric_map.items():
            if code == target_metric_code:
                continue
            candidates = [
                code,
                str(metric.get("label") or ""),
                *[str(alias) for alias in metric.get("aliases") or []],
                *[str(alias) for alias in metric.get("approved_aliases") or []],
                *RUSSIAN_METRIC_ALIASES.get(code, []),
            ]
            for candidate in candidates:
                candidate_text = str(candidate or "").strip().lower()
                if len(candidate_text) >= 4 and candidate_text in normalized:
                    return code, display_metric_label(code, metric)
        phrase = self._extract_driver_phrase_from_answer(text)
        if not phrase:
            return None
        code = normalize_metric_code(phrase)
        if not code or code == "metric":
            return None
        return code, phrase[:96]

    def _extract_driver_phrase_from_answer(self, text: str) -> str:
        cleaned = re.sub(r"\s+", " ", str(text or "")).strip(" .")
        if not cleaned:
            return ""
        patterns = [
            r"(?:это\s+)?(?:связано|связана|связан|зависит)\s+(?:с|от)\s+(.+)$",
            r"(?:из[\s-]*за|из-за)\s+(.+)$",
            r"(?:влияет|повлиял[ао]?|драйвер)\s+(.+)$",
        ]
        for pattern in patterns:
            match = re.search(pattern, cleaned, flags=re.IGNORECASE)
            if match:
                return match.group(1).strip(" .")
        if len(cleaned.split()) <= 8:
            return cleaned
        return ""

    def _is_affirmative_answer(self, message: str) -> bool:
        normalized = str(message or "").strip().lower()
        return bool(re.match(r"^(да|верно|подтверждаю|согласен|согласна|именно|скорее да)\b", normalized))

    def _coerce_edge_type(self, value: Any) -> str:
        edge_type = str(value or "driver").strip()
        if edge_type not in {"driver", "inverse_driver", "component", "lag"}:
            return "driver"
        return edge_type

    def _memory_fact_from_user_answer(
        self,
        state: RelationMemorySessionState,
        open_question: dict[str, Any],
        message: str,
        payload: dict[str, Any],
    ) -> dict[str, Any]:
        metric_codes = [
            str(item)
            for item in payload.get("metric_codes", []) or []
            if str(item)
        ]
        target = str(open_question.get("target_metric_code") or "")
        if target and target not in metric_codes:
            metric_codes.append(target)
        content = str(payload.get("content") or message).strip()
        fact_id = f"user_answer:{open_question.get('id') or 'question'}:{uuid.uuid4().hex[:8]}"
        return {
            "id": normalize_metric_code(fact_id),
            "kind": str(payload.get("kind") or "business_fact"),
            "content": content,
            "keywords": str(payload.get("keywords") or ""),
            "metric_codes": metric_codes,
            "confidence": self._bounded_score(payload.get("confidence"), default=0.7),
            "source_question_id": str(open_question.get("id") or ""),
            "source_question": str(open_question.get("prompt") or ""),
            "created_at": datetime.now(UTC).isoformat(),
            "source": "user_answer",
        }

    def _bounded_score(self, value: Any, *, default: float) -> float:
        try:
            score = float(value if value is not None else default)
        except (TypeError, ValueError):
            score = default
        return round(max(0.0, min(1.0, score)), 3)

    def _next_agent_question(self, state: RelationMemorySessionState) -> dict[str, Any] | None:
        for question in state.snapshot.inquiry_questions:
            question_id = str(question.get("id") or "")
            if question_id and question_id not in state.answered_agent_question_ids:
                return copy.deepcopy(question)
        return None

    def _localize_inquiry_questions(self, state: RelationMemorySessionState) -> None:
        for question in state.snapshot.inquiry_questions:
            prompt = str(question.get("prompt") or "")
            if not prompt:
                continue
            for code, metric in state.metric_map.items():
                display_label = display_metric_label(code, metric)
                candidate_labels = [
                    str(metric.get("label") or ""),
                    str(metric.get("raw_label") or ""),
                    str(code).replace("_", " ").title(),
                    *[str(item or "") for item in metric.get("source_labels") or []],
                    *[str(item or "") for item in metric.get("approved_aliases") or []],
                ]
                for candidate_label in dict.fromkeys(label.strip() for label in candidate_labels if label.strip()):
                    prompt = self._replace_prompt_label(prompt, candidate_label, display_label)
            question["prompt"] = prompt

    def _replace_prompt_label(self, prompt: str, candidate_label: str, display_label: str) -> str:
        if not candidate_label or candidate_label == display_label or contains_cyrillic(candidate_label):
            return prompt
        prompt = prompt.replace(f"«{candidate_label}»", f"«{display_label}»")
        if " " in candidate_label or any(char.isupper() for char in candidate_label):
            prompt = re.sub(
                rf"(?<![\w]){re.escape(candidate_label)}(?![\w])",
                display_label,
                prompt,
            )
        return prompt

    def _agent_answer_acknowledgement(
        self,
        *,
        state: RelationMemorySessionState,
        llm_message: str,
        created: list[CandidateRelation],
        facts: list[dict[str, Any]],
        next_question: dict[str, Any] | None,
    ) -> str:
        parts: list[str] = []
        if llm_message and not self._looks_like_english_assistant_text(llm_message):
            parts.append(str(llm_message).strip())
        elif created:
            if len(created) == 1:
                candidate = created[0]
                source_label = metric_label_for_chat(state.metric_map, candidate.source_metric_code)
                target_label = metric_label_for_chat(state.metric_map, candidate.target_metric_code)
                evidence = str(candidate.evidence or candidate.note or "").strip()
                evidence_suffix = f" Основание: {evidence[:180]}" if evidence else ""
                parts.append(
                    "Запомнил ответ и подготовил кандидатную связь для проверки: "
                    f"«{source_label}» -> «{target_label}»."
                    f"{evidence_suffix}"
                )
            else:
                parts.append(f"Запомнил ответ и подготовил {len(created)} кандидатных связей для проверки.")
        elif facts:
            parts.append("Запомнил ответ как бизнес-контекст для этой relation-memory сессии.")
        else:
            parts.append("Запомнил ответ.")
        if next_question:
            parts.append(f"Следующий вопрос: {next_question.get('prompt')}")
        else:
            parts.append("Открытых уточняющих вопросов больше нет.")
        return " ".join(part for part in parts if part).strip()

    def _answer_open_session_question(self, state: RelationMemorySessionState, question: str) -> str:
        fallback_answer = self._fallback_open_question_answer(state, question)
        if self._is_relation_source_question(question):
            return fallback_answer
        if not settings.RELATION_MEMORY_LLM_ANSWER_ENABLED:
            return fallback_answer

        client = RelationMemoryLlmClient(timeout_seconds=int(settings.RELATION_MEMORY_LLM_ANSWER_TIMEOUT_SECONDS))
        system_prompt = (
            "Ты отвечаешь на свободные вопросы бизнес-пользователя по текущей relation-memory сессии. "
            "Используй только факты из payload: documents, metrics, dependencies, observations, hypotheses, pending_confirmations. "
            "Не выдумывай новые метрики, цифры или связи. Если вопрос не покрыт данными, скажи это коротко и предложи, "
            "какой вопрос можно задать по текущему графу. Верни strict JSON only: {\"answer\": \"...\"}."
        )
        try:
            payload = client.chat_json(
                system_prompt=system_prompt,
                user_payload={
                    "question": question,
                    "session": self._open_question_context(state),
                    "fallback_answer": fallback_answer,
                },
            )
        except LlmJsonError:
            return fallback_answer
        answer = str(payload.get("answer") or "").strip()
        if self._looks_like_english_assistant_text(answer):
            return fallback_answer
        return answer or fallback_answer

    def _open_question_context(self, state: RelationMemorySessionState) -> dict[str, Any]:
        metric_lookup = {
            code: display_metric_label(code, metric)
            for code, metric in state.metric_map.items()
        }
        dependencies = []
        for item in sorted(
            state.snapshot.dependencies,
            key=lambda dependency: -float(dependency.get("strength") or 0.0),
        )[:40]:
            source = str(item.get("source_metric_code") or "")
            target = str(item.get("target_metric_code") or "")
            if not source or not target:
                continue
            dependencies.append(
                {
                    "source_metric_code": source,
                    "source_label": metric_lookup.get(source, source),
                    "target_metric_code": target,
                    "target_label": metric_lookup.get(target, target),
                    "edge_type": item.get("edge_type") or "driver",
                    "strength": item.get("strength"),
                    "reason": item.get("reason") or "",
                    "source": item.get("source") or "",
                }
            )
        pending = []
        for item in sorted(
            state.pending_confirmations.values(),
            key=lambda candidate: -float(candidate.score or candidate.confidence or 0.0),
        )[:20]:
            pending.append(
                {
                    "source_metric_code": item.source_metric_code,
                    "source_label": metric_lookup.get(item.source_metric_code, item.source_metric_code),
                    "target_metric_code": item.target_metric_code,
                    "target_label": metric_lookup.get(item.target_metric_code, item.target_metric_code),
                    "edge_type": item.edge_type,
                    "score": item.score or item.confidence,
                    "reason": item.needs_approval_reason or item.note or item.evidence,
                    "source": item.source,
                }
            )
        return {
            "documents": [document.to_dict() for document in state.bundle.documents[:12]],
            "metrics": [
                {
                    "code": code,
                    "label": metric_lookup.get(code, code),
                    "canonical_label": metric.get("label") or code,
                    "aliases": list(metric.get("aliases") or [])[:8],
                }
                for code, metric in sorted(state.metric_map.items())[:80]
            ],
            "dependencies": dependencies,
            "pending_confirmations": pending,
            "observations": state.snapshot.observations[:12],
            "hypotheses": state.snapshot.hypotheses[:12],
            "warnings": state.warnings[:12],
        }

    def _fallback_open_question_answer(self, state: RelationMemorySessionState, question: str) -> str:
        relation_count = len(state.snapshot.dependencies)
        pending_count = len(state.pending_confirmations)
        metric_count = len(state.metric_map)
        question_text = str(question or "").strip()
        if self._is_relation_source_question(question_text):
            return (
                "Связи в этой demo-сессии берутся из загруженных таблиц зависимостей, найденных наблюдений и гипотез, "
                "подтвержденной памяти и ответов пользователя на уточняющие вопросы. LLM может нормализовать вопрос "
                "или предложить кандидатную связь, но такая связь остается на проверке, пока пользователь явно ее не подтвердит."
            )
        if any(token in question_text.lower() for token in ("связ", "relation", "edge")):
            examples = []
            metric_lookup = {
                code: display_metric_label(code, metric)
                for code, metric in state.metric_map.items()
            }
            for item in state.snapshot.dependencies[:3]:
                source = str(item.get("source_metric_code") or "")
                target = str(item.get("target_metric_code") or "")
                if source and target:
                    examples.append(f"«{metric_lookup.get(source, source)}» -> «{metric_lookup.get(target, target)}»")
            suffix = f" Например: {'; '.join(examples)}." if examples else ""
            return f"В графе есть {relation_count} связей и {pending_count} связей на проверке.{suffix}"
        return (
            f"Я вижу текущую сессию: {metric_count} метрик, {relation_count} связей в графе, "
            f"{pending_count} связей на проверке. Могу отвечать по метрикам, зависимостям, причинам изменений "
            "и кандидатам на подтверждение."
        )

    def _is_capability_help_question(self, question: str) -> bool:
        normalized = normalize_user_text(question)
        return any(
            phrase in normalized
            for phrase in (
                "что я могу тут спросить",
                "что можно спросить",
                "какие вопросы можно",
                "помощь",
                "help",
            )
        )

    def _capability_help_answer(self) -> str:
        return (
            "Можно спросить: что влияет на метрику, на что влияет метрика, есть ли связь между двумя метриками, "
            "почему метрика выросла или упала, какие есть метрики, связи, гипотезы, наблюдения и pending-связи. "
            "Если данных не хватает, я задам уточняющий вопрос или предложу кандидатную связь для проверки."
        )

    def _is_relation_source_question(self, question: str) -> bool:
        normalized = normalize_user_text(question)
        if "связ" not in normalized and "relation" not in normalized and "edge" not in normalized:
            return False
        return any(
            phrase in normalized
            for phrase in (
                "как получаются",
                "как формируются",
                "как строятся",
                "откуда берутся",
                "на основе чего",
                "за счет чего",
                "из чего",
            )
        )

    def handle_chat_message(
        self,
        session_id: str,
        message: str,
        *,
        trace_sink: Callable[[dict[str, Any]], None] | None = None,
    ) -> RelationMemoryChatResponse:
        state = self._get_session(session_id)
        state.chat_history.append(self._conversation_message("user", message))
        assistant_message, created, warnings = self.handle_message(session_id, message, trace_sink=trace_sink)
        assistant_message = self._humanize_chat_message(
            state=state,
            assistant_message=assistant_message,
            created=created,
        )
        debug_trace = list(state.last_debug_trace)
        state.chat_history.append(
            self._conversation_message(
                "assistant",
                assistant_message,
                pending_confirmation_ids=[candidate.id for candidate in created],
                warnings=warnings,
                debug_trace=debug_trace,
            )
        )
        return RelationMemoryChatResponse(
            session_id=session_id,
            assistant_message=assistant_message,
            created_pending_confirmations=created,
            pending_confirmations=list(state.pending_confirmations.values()),
            metric_candidates=list(state.metric_candidates.values()),
            chat_history=list(state.chat_history),
            session=self._session_response(state),
            warnings=warnings,
            debug_trace=debug_trace,
        )

    def answer_graph_question(self, session_id: str, question: str) -> GraphQuestionResponse:
        state = self._get_session(session_id)
        answer = self.question_service.answer(
            question=question,
            snapshot=self._question_snapshot(state),
            metric_map=self._question_metric_map(state),
            pending_confirmations=list(state.pending_confirmations.values()),
            conversation_context=state.conversation_context,
        )
        state.conversation_context = answer.updated_context or state.conversation_context
        self._persist_confirmed_metric_aliases(
            state=state,
            aliases=answer.confirmed_aliases,
            session_id=session_id,
        )
        if not answer.handled:
            answer.answer = (
                "Пока понимаю только вопросы вида: что влияет на X, куда влияет X, "
                "почему X связан с Y, какие связи требуют подтверждения."
            )
        return GraphQuestionResponse(session_id=session_id, **answer.to_dict())

    def _question_snapshot(self, state: RelationMemorySessionState) -> RelationMemorySnapshot:
        snapshot = copy.deepcopy(state.snapshot)
        existing = {
            (
                str(item.get("source_metric_code") or ""),
                str(item.get("target_metric_code") or ""),
                str(item.get("edge_type") or "driver"),
            )
            for item in snapshot.dependencies
        }
        for row in self._memory_prior_dependency_rows(state):
            key = (
                str(row.get("source_metric_code") or ""),
                str(row.get("target_metric_code") or ""),
                str(row.get("edge_type") or "driver"),
            )
            if key in existing:
                continue
            snapshot.dependencies.append(row)
            existing.add(key)
        return snapshot

    def _question_metric_map(self, state: RelationMemorySessionState) -> dict[str, dict[str, Any]]:
        metric_map = copy.deepcopy(state.metric_map)
        for code, metric in metric_map.items():
            extend_metric_values(metric, "aliases", RUSSIAN_METRIC_ALIASES.get(code, []))
        for row in self._memory_prior_dependency_rows(state):
            ensure_question_metric(metric_map, str(row.get("source_metric_code") or ""))
            ensure_question_metric(metric_map, str(row.get("target_metric_code") or ""))
        return metric_map

    def _memory_prior_dependency_rows(self, state: RelationMemorySessionState) -> list[dict[str, Any]]:
        rows = []
        for prior in self._relation_memory_priors(state):
            source = str(prior.get("source_metric_code") or "")
            target = str(prior.get("target_metric_code") or "")
            if not source or not target:
                continue
            rows.append(
                {
                    "id": f"memory:{source}->{target}",
                    "source_metric_code": source,
                    "target_metric_code": target,
                    "edge_type": str(prior.get("edge_type") or "driver"),
                    "reason": str(prior.get("note") or prior.get("reason") or "Связь из подтвержденной памяти.").strip(),
                    "strength": float(prior.get("strength") or prior.get("score") or prior.get("confidence") or 1.0),
                    "source": str(prior.get("source") or "memory_prior"),
                }
            )
        return rows

    def confirm_candidate(self, session_id: str, candidate_id: str, decision: str) -> ConfirmationResponse:
        if decision not in ALLOWED_CONFIRMATION_DECISIONS:
            raise ValueError(f"Unsupported confirmation decision: {decision}")
        state = self._get_session(session_id)
        candidate = state.pending_confirmations.get(candidate_id)
        if not candidate:
            raise KeyError(f"Unknown pending candidate: {candidate_id}")

        if decision == "reject":
            state.rejected_candidate_ids.add(candidate_id)
            del state.pending_confirmations[candidate_id]
            return ConfirmationResponse(
                session_id=session_id,
                candidate_id=candidate_id,
                decision="reject",
                saved=False,
                relation=None,
                message="Связь отклонена и не сохранена в долговременную память.",
            )

        self._validate_candidate_metrics(candidate, state)
        source_metric = state.metric_map.get(candidate.source_metric_code, {})
        target_metric = state.metric_map.get(candidate.target_metric_code, {})
        source_label = display_metric_label(candidate.source_metric_code, source_metric)
        target_label = display_metric_label(candidate.target_metric_code, target_metric)
        saved = self._graph_client().save_confirmed_relation(
            source_metric_code=candidate.source_metric_code,
            target_metric_code=candidate.target_metric_code,
            edge_type=candidate.edge_type,
            lag_period=candidate.lag_period,
            note=candidate.note or candidate.evidence,
            source_session_id=session_id,
            source_document_id=candidate.source_document_id,
            source_label=source_label,
            target_label=target_label,
            evidence_text=candidate.evidence,
            evidence_type=candidate.evidence_type,
            confidence=max(candidate.confidence, candidate.score),
        )
        del state.pending_confirmations[candidate_id]
        relation = ConfirmedRelation(**saved)
        state.approved_relations.append(relation)
        state.confirmed_relation_keys.add(relation_key(candidate))
        return ConfirmationResponse(
            session_id=session_id,
            candidate_id=candidate_id,
            decision="approve",
            saved=True,
            relation=relation,
            message="Связь подтверждена и сохранена в Neo4j memory.",
        )

    def list_memory_relations(self) -> list[ConfirmedRelation]:
        return [
            ConfirmedRelation(**self._localized_confirmed_relation_payload(item))
            for item in self._graph_client().list_confirmed_relations()
        ]

    def _localized_confirmed_relation_payload(self, item: dict[str, Any]) -> dict[str, Any]:
        payload = dict(item)
        replacements: dict[str, str] = {}
        for key in ("source_metric_code", "target_metric_code"):
            code = str(payload.get(key) or "")
            display = RUSSIAN_METRIC_DISPLAY_LABELS.get(code.lower())
            if not display:
                continue
            for variant in self._metric_name_variants(code):
                replacements[variant] = display
        note = str(payload.get("note") or "")
        for source, target in sorted(replacements.items(), key=lambda item: -len(item[0])):
            note = re.sub(
                rf"(?<![0-9A-Za-zА-Яа-яЁё_]){re.escape(source)}(?![0-9A-Za-zА-Яа-яЁё_])",
                target,
                note,
                flags=re.IGNORECASE,
            )
        payload["note"] = note
        return payload

    def _metric_name_variants(self, code: str) -> list[str]:
        spaced = str(code or "").replace("_", " ").strip()
        titled = " ".join(part[:1].upper() + part[1:].lower() for part in spaced.split())
        return list(dict.fromkeys(item for item in (code, spaced, titled) if item))

    def list_metric_candidates(self, session_id: str) -> list[MetricCandidate]:
        state = self._get_session(session_id)
        return list(state.metric_candidates.values())

    def confirm_metric_candidate(
        self,
        session_id: str,
        candidate_id: str,
        decision: str,
    ) -> MetricCandidateConfirmationResponse:
        if decision not in ALLOWED_CONFIRMATION_DECISIONS:
            raise ValueError(f"Unsupported confirmation decision: {decision}")
        state = self._get_session(session_id)
        candidate = state.metric_candidates.get(candidate_id)
        if not candidate:
            raise KeyError(f"Unknown metric candidate: {candidate_id}")

        if decision == "reject":
            state.rejected_metric_candidate_ids.add(candidate_id)
            del state.metric_candidates[candidate_id]
            return MetricCandidateConfirmationResponse(
                session_id=session_id,
                candidate_id=candidate_id,
                decision="reject",
                saved=False,
                metric_candidate=None,
                message="Маппинг метрики отклонен и не сохранен в память.",
            )

        if candidate.approved:
            return MetricCandidateConfirmationResponse(
                session_id=session_id,
                candidate_id=candidate_id,
                decision="approve",
                saved=True,
                metric_candidate=candidate,
                message="Маппинг метрики уже применен в текущей сессии.",
            )

        approved = candidate.model_copy(update={"approved": True, "status": "approved"})
        graph_client = self._graph_client()
        if hasattr(graph_client, "save_metric_mapping"):
            graph_client.save_metric_mapping(
                canonical_code=approved.canonical_code,
                raw_label=approved.raw_label,
                label=approved.label or approved.raw_label,
                aliases=approved.aliases,
                unit=approved.unit,
                department=approved.department,
                source_sheet=approved.source_sheet,
                section_path=approved.section_path,
                semantic_type=approved.semantic_type,
                aggregation=approved.aggregation,
                confidence=approved.confidence,
                evidence=approved.evidence,
                source_session_id=session_id,
            )
        state.metric_candidates[candidate_id] = approved
        metric_entry = ensure_metric_entry(
            state.metric_map,
            approved.canonical_code,
            label=approved.label or approved.raw_label,
            aliases=[approved.raw_label, approved.label, *approved.aliases],
            approved_aliases=[approved.raw_label, approved.label, *approved.aliases],
            source_labels=[approved.raw_label, approved.label],
            source="approved_metric_mapping",
        )
        metric_entry["label"] = metric_entry.get("label") or approved.label or approved.raw_label
        return MetricCandidateConfirmationResponse(
            session_id=session_id,
            candidate_id=candidate_id,
            decision="approve",
            saved=True,
            metric_candidate=approved,
            message="Маппинг метрики подтвержден и сохранен в memory layer.",
        )

    def _build_session_state(
        self,
        *,
        session_id: str,
        bundle: DocumentBundle,
        previous_state: RelationMemorySessionState | None = None,
    ) -> RelationMemorySessionState:
        memory_priors = self._relation_memory_priors(previous_state)
        metric_mapping_priors = self._metric_mapping_priors(previous_state)
        (
            snapshot,
            metric_candidates,
            dynamic_relation_candidates,
            dynamic_warnings,
            document_parse_statuses,
            parsed_document_ids,
        ) = self._build_snapshot_from_uploaded_xlsx(
            bundle,
            memory_priors,
            metric_mapping_priors,
        )
        metric_map = build_metric_map(
            snapshot_metrics=snapshot.metrics,
            detected_metrics=bundle.detected_metrics,
            parsed_document_ids=parsed_document_ids,
            metric_mapping_priors=metric_mapping_priors,
            mapping_aliases=self._mapping_aliases,
        )
        state = RelationMemorySessionState(
            id=session_id,
            bundle=bundle,
            snapshot=snapshot,
            assistant_message="",
            metric_map=metric_map,
            document_parse_statuses=document_parse_statuses,
            parsed_document_ids=parsed_document_ids,
            metric_candidates={
                candidate.id: candidate
                for candidate in metric_candidates
                if not previous_state or candidate.id not in previous_state.rejected_metric_candidate_ids
            },
            rejected_metric_candidate_ids=set(previous_state.rejected_metric_candidate_ids) if previous_state else set(),
            rejected_candidate_ids=set(previous_state.rejected_candidate_ids) if previous_state else set(),
            approved_relations=list(previous_state.approved_relations) if previous_state else [],
            confirmed_relation_keys=relation_key_set(memory_priors),
            conversation_context=(
                copy.deepcopy(previous_state.conversation_context)
                if previous_state
                else RelationMemoryConversationContext()
            ),
            open_agent_question=(
                copy.deepcopy(previous_state.open_agent_question)
                if previous_state and previous_state.open_agent_question
                else None
            ),
            answered_agent_question_ids=(
                set(previous_state.answered_agent_question_ids)
                if previous_state
                else set()
            ),
            memory_facts=copy.deepcopy(previous_state.memory_facts) if previous_state else [],
            chat_history=list(previous_state.chat_history) if previous_state else [],
            warnings=self._unique_warnings([*bundle.warnings, *dynamic_warnings]),
        )
        self._localize_inquiry_questions(state)
        if state.open_agent_question is None:
            state.open_agent_question = self._next_agent_question(state)

        for candidate in dynamic_relation_candidates:
            self._add_candidate(state, candidate)

        for candidate in candidate_relations_from_dependencies(
            snapshot.dependencies,
            row_relation_source=ROW_REPORT_RELATION_SOURCE,
            limit=REPORT_RELATION_CANDIDATE_LIMIT,
        ):
            self._add_candidate(state, candidate)

        for candidate in candidate_relations_from_hypotheses(snapshot.hypotheses):
            self._add_candidate(state, candidate)

        text_result = self.agent.extract_relations_from_text(
            chunks=bundle.text_chunks,
            known_metrics=list(metric_map.values()),
        )
        state.warnings = self._unique_warnings([*state.warnings, *text_result.warnings])
        for raw_candidate in text_result.candidate_relations:
            candidate = self._candidate_from_raw(raw_candidate, state, source="llm_text")
            if candidate:
                self._add_candidate(state, candidate)

        state.assistant_message = self._base_assistant_message(
            state=state,
            assistant_message=text_result.assistant_message,
            memory_priors=memory_priors,
        )
        if not state.chat_history:
            state.chat_history.append(
                self._conversation_message(
                    "assistant",
                    self._humanize_initial_chat_message(state.assistant_message),
                )
            )
        return state

    def _session_response(self, state: RelationMemorySessionState) -> RelationMemorySessionResponse:
        return RelationMemorySessionResponse(
            session_id=state.id,
            assistant_message=state.assistant_message,
            documents=[DocumentSummary(**document.to_dict()) for document in state.bundle.documents],
            document_parse_statuses=list(state.document_parse_statuses),
            metrics=metric_response_items(state.metric_map),
            observations=state.snapshot.observations,
            hypotheses=state.snapshot.hypotheses,
            inquiry_questions=state.snapshot.inquiry_questions,
            metric_candidates=list(state.metric_candidates.values()),
            pending_confirmations=list(state.pending_confirmations.values()),
            open_agent_question=copy.deepcopy(state.open_agent_question),
            memory_facts=copy.deepcopy(state.memory_facts),
            chat_history=list(state.chat_history),
            warnings=state.warnings,
        )

    def _merge_document_bundles(self, current: DocumentBundle, incoming: DocumentBundle) -> DocumentBundle:
        return DocumentBundle(
            documents=[*current.documents, *incoming.documents],
            text_chunks=[*current.text_chunks, *incoming.text_chunks],
            tables=[*current.tables, *incoming.tables],
            detected_metrics=[*current.detected_metrics, *incoming.detected_metrics],
            xlsx_files=[*current.xlsx_files, *incoming.xlsx_files],
            warnings=self._unique_warnings([*current.warnings, *incoming.warnings]),
        )

    def _conversation_message(
        self,
        role: str,
        content: str,
        *,
        pending_confirmation_ids: list[str] | None = None,
        warnings: list[str] | None = None,
        debug_trace: list[dict[str, Any]] | None = None,
    ) -> ConversationMessage:
        return ConversationMessage(
            role=role,
            content=str(content or "").strip(),
            created_at=datetime.now(UTC).isoformat(),
            pending_confirmation_ids=list(pending_confirmation_ids or []),
            warnings=list(warnings or []),
            debug_trace=list(debug_trace or []),
        )

    def _humanize_initial_chat_message(self, message: str) -> str:
        text = str(message or "").strip()
        if not text:
            return "Файлы загружены. Можете спросить, что влияет на показатель, или описать бизнес-связь своими словами."
        return (
            text.replace("pending_confirmations", "связях для подтверждения")
            .replace("кандидатных маппингов метрик", "вариантов распознавания метрик")
            .replace("ручного подтверждения", "проверки")
        )

    def _humanize_chat_message(
        self,
        *,
        state: RelationMemorySessionState,
        assistant_message: str,
        created: list[CandidateRelation],
    ) -> str:
        if created:
            if any(candidate.source == "user_answer" for candidate in created):
                return assistant_message
            if any(candidate.source == "llm_external_context" for candidate in created):
                return assistant_message
            if len(created) == 1:
                candidate = created[0]
                source = metric_label_for_chat(state.metric_map, candidate.source_metric_code)
                target = metric_label_for_chat(state.metric_map, candidate.target_metric_code)
                return (
                    f"Понял. Я вижу возможную связь: «{source}» влияет на «{target}». "
                    "Проверьте карточку ниже и подтвердите ее, если это верно."
                )
            return (
                f"Понял. Я нашел {len(created)} возможных связей. "
                "Покажу их ниже, чтобы вы могли подтвердить только те, которые действительно верны."
            )
        text = str(assistant_message or "").strip()
        if "Не смог надежно разобрать ответ LLM" in text:
            return "Пока не могу уверенно разобрать это сообщение. Сформулируйте связь проще: что на что влияет и почему."
        if text.startswith("Не нашел новых связей"):
            return "Пока не вижу новой связи в сообщении. Можно спросить про влияние метрики или описать связь между двумя показателями."
        if self._looks_like_english_assistant_text(text):
            return "Пока не вижу новой связи в сообщении. Можно спросить про влияние метрики или описать связь между двумя показателями."
        return text

    def _create_external_context_candidates(
        self,
        state: RelationMemorySessionState,
        graph_answer: GraphQuestionAnswer,
    ) -> list[CandidateRelation]:
        created: list[CandidateRelation] = []
        for row in graph_answer.rows:
            if str(row.get("status") or "") != "external_context_candidate":
                continue
            raw_candidate = {
                "source_metric_code": row.get("source_metric_code"),
                "target_metric_code": row.get("target_metric_code"),
                "edge_type": row.get("edge_type") or "driver",
                "note": row.get("reason") or row.get("first_reason") or "",
                "evidence": row.get("reason") or row.get("first_reason") or "External context suggestion",
                "evidence_type": "external_context",
                "confidence": row.get("confidence") or row.get("score") or 0.5,
                "score": row.get("score") or row.get("confidence") or 0.5,
                "needs_approval_reason": "Предложено внешним контекстом при отсутствии подтвержденной связи в графе.",
                "source": "llm_external_context",
            }
            ensure_external_context_metric_label(
                state.metric_map,
                str(row.get("source_metric_code") or ""),
                str(row.get("source_label") or ""),
            )
            ensure_external_context_metric_label(
                state.metric_map,
                str(row.get("target_metric_code") or ""),
                str(row.get("target_label") or ""),
            )
            candidate = self._candidate_from_raw(raw_candidate, state, source="llm_external_context")
            if candidate is None:
                continue
            stored = self._add_candidate_if_new(state, candidate)
            if stored is not None:
                created.append(stored)
        return created

    def _looks_like_english_assistant_text(self, value: str) -> bool:
        text = str(value or "")
        if not text:
            return False
        lowered = text.lower()
        if any(marker in lowered for marker in ("assuming", "correcting", "candidate relationships", "here are")):
            return True
        latin = len(re.findall(r"[A-Za-z]", text))
        cyrillic = len(re.findall(r"[А-Яа-яЁё]", text))
        return latin >= 24 and latin > cyrillic * 2

    def _base_assistant_message(
        self,
        *,
        state: RelationMemorySessionState,
        assistant_message: str,
        memory_priors: list[dict[str, Any]],
    ) -> str:
        message = assistant_message or self.agent.build_initial_message(
            observations=state.snapshot.observations,
            hypotheses=state.snapshot.hypotheses,
            inquiry_questions=state.snapshot.inquiry_questions,
            memory_priors=memory_priors,
        )
        if state.metric_candidates:
            message = (
                f"{message} Найдено {len(state.metric_candidates)} кандидатных маппингов метрик "
                "для подтверждения."
            )
        if state.pending_confirmations:
            message = (
                f"{message} Также подготовлено {len(state.pending_confirmations)} кандидатных связей "
                "для ручного подтверждения в pending_confirmations."
            )
        return message

    def _build_snapshot_from_uploaded_xlsx(
        self,
        bundle: DocumentBundle,
        memory_priors: list[dict[str, Any]],
        metric_mapping_priors: list[dict[str, Any]],
    ) -> tuple[RelationMemorySnapshot, list[MetricCandidate], list[dict[str, Any]], list[str], list[DocumentParseStatus], set[str]]:
        document_parse_statuses = self._document_parse_statuses_for_non_tabular_documents(bundle)
        if not bundle.xlsx_files:
            return RelationMemorySnapshot(), [], [], [], document_parse_statuses, set()

        with make_tempdir() as temp_dir:
            xlsx_paths = write_xlsx_payloads_to_tempdir(bundle.xlsx_files, temp_dir)
            xlsx_documents = [document for document in bundle.documents if document.file_type in {"xlsx", "xlsm"}]
            parsed_artifacts: list[ParsedWorkbookArtifact] = []
            metric_candidates: dict[str, MetricCandidate] = {}
            dynamic_relation_candidates: list[dict[str, Any]] = []
            warnings: list[str] = []

            for index, (path, document) in enumerate(zip(xlsx_paths, xlsx_documents, strict=False)):
                artifact = self._parse_workbook_artifact(
                    path=path,
                    document=document,
                    index=index,
                    temp_dir=Path(temp_dir),
                    memory_priors=memory_priors,
                    metric_mapping_priors=metric_mapping_priors,
                )
                document_parse_statuses.append(
                    DocumentParseStatus(
                        document_id=artifact.document_id,
                        filename=artifact.filename,
                        file_type=artifact.file_type,
                        parser=artifact.parser,
                        status=artifact.status,
                        dataset_count=artifact.dataset_count,
                        metric_count=artifact.metric_count,
                        dependency_count=artifact.dependency_count,
                        warnings=artifact.warnings,
                    )
                )
                warnings.extend(artifact.warnings)
                if artifact.status not in {"parsed", "parsed_with_warnings"}:
                    continue
                parsed_artifacts.append(artifact)
                for candidate in artifact.metric_candidates:
                    metric_candidates[candidate.id] = candidate
                dynamic_relation_candidates.extend(artifact.relation_candidates)

            parsed_document_ids = {artifact.document_id for artifact in parsed_artifacts}
            if not parsed_artifacts:
                return (
                    RelationMemorySnapshot(),
                    list(metric_candidates.values()),
                    dynamic_relation_candidates,
                    self._unique_warnings(warnings),
                    document_parse_statuses,
                    parsed_document_ids,
                )

            snapshot = self._build_combined_snapshot(
                parsed_artifacts=parsed_artifacts,
                temp_dir=Path(temp_dir),
                memory_priors=memory_priors,
            )
            return (
                snapshot,
                list(metric_candidates.values()),
                dynamic_relation_candidates,
                self._unique_warnings(warnings),
                document_parse_statuses,
                parsed_document_ids,
            )

    def _document_parse_statuses_for_non_tabular_documents(self, bundle: DocumentBundle) -> list[DocumentParseStatus]:
        warnings_by_filename: dict[str, list[str]] = {}
        for warning in bundle.warnings:
            for document in bundle.documents:
                if document.filename in warning:
                    warnings_by_filename.setdefault(document.filename, []).append(warning)
        statuses: list[DocumentParseStatus] = []
        for document in bundle.documents:
            if document.file_type in {"xlsx", "xlsm"}:
                continue
            parser = "text_ingestion" if document.file_type in {"txt", "md", "pdf"} else "unsupported_file_type"
            status = "parsed" if document.file_type in {"txt", "md", "pdf"} else "unsupported_file_type"
            document_warnings = warnings_by_filename.get(document.filename, [])
            if document_warnings and status == "parsed":
                status = "parsed_with_warnings"
            statuses.append(
                DocumentParseStatus(
                    document_id=document.id,
                    filename=document.filename,
                    file_type=document.file_type,
                    parser=parser,
                    status=status,
                    warnings=document_warnings,
                )
            )
        return statuses

    def _parse_workbook_artifact(
        self,
        *,
        path: Path,
        document: Any,
        index: int,
        temp_dir: Path,
        memory_priors: list[dict[str, Any]],
        metric_mapping_priors: list[dict[str, Any]],
    ) -> ParsedWorkbookArtifact:
        try:
            metadata_artifact = self._metadata_artifact_from_workbook(path=path, document=document)
            if metadata_artifact is not None:
                return metadata_artifact

            dynamic_result = self._try_build_dynamic_snapshot(
                path,
                temp_dir / f"dynamic_{index}",
                memory_priors,
                metric_mapping_priors,
            )
            if dynamic_result is not None:
                snapshot, metric_candidates, relation_candidates, warnings, artifacts = dynamic_result
                metric_dictionary_rows = self._read_metadata_rows(artifacts.metric_candidates_path)
                dependency_rule_rows = self._read_metadata_rows(artifacts.dependency_rules_path)
                parser = "dynamic_row_metric"
                enrichment_metric_count = 0
                enrichment_dependency_count = 0
                if len(relation_candidates) < AGENTIC_ENRICHMENT_MIN_RELATION_CANDIDATES:
                    enrichment = self._try_build_agentic_enrichment(path=path, document=document)
                    if enrichment is not None and self._has_agentic_enrichment_payload(enrichment):
                        parser = "dynamic_row_metric+agentic_enrichment"
                        warnings = self._unique_warnings([*warnings, *enrichment.warnings])
                        metric_dictionary_rows.extend(enrichment.metric_dictionary_rows)
                        dependency_rule_rows.extend(enrichment.dependency_rule_rows)
                        metric_candidates.extend(self._metric_candidates_from_agentic_enrichment(enrichment))
                        relation_candidates.extend(enrichment.candidate_relations)
                        enrichment_metric_count = len(enrichment.metric_dictionary_rows)
                        enrichment_dependency_count = len(enrichment.dependency_rule_rows)
                return ParsedWorkbookArtifact(
                    document_id=document.id,
                    filename=document.filename,
                    file_type=document.file_type,
                    parser=parser,
                    status="parsed_with_warnings" if warnings else "parsed",
                    contracts=[self._dynamic_numeric_facts_contract(artifacts.pivoted_facts_path, index)],
                    metric_dictionary_rows=metric_dictionary_rows,
                    dependency_rule_rows=dependency_rule_rows,
                    metric_candidates=metric_candidates,
                    relation_candidates=relation_candidates,
                    warnings=warnings,
                    dataset_count=len(snapshot.datasets),
                    metric_count=len(snapshot.metrics) + enrichment_metric_count,
                    dependency_count=len(snapshot.dependencies) + enrichment_dependency_count,
                )

            mapped_path, applied_metric_candidates, direct_warnings = self._rewrite_month_based_xlsx(
                path,
                index=index,
                temp_dir=temp_dir,
                metric_mapping_priors=metric_mapping_priors,
            )
            contract = self._contract_for_xlsx(mapped_path, index)
            if "month" not in contract.get("grain", []):
                enrichment = self._try_build_agentic_enrichment(path=path, document=document)
                if enrichment is not None and self._has_agentic_enrichment_payload(enrichment):
                    return ParsedWorkbookArtifact(
                        document_id=document.id,
                        filename=document.filename,
                        file_type=document.file_type,
                        parser="agentic_excel_enrichment",
                        status="parsed_with_warnings" if enrichment.warnings else "parsed",
                        metric_dictionary_rows=enrichment.metric_dictionary_rows,
                        dependency_rule_rows=enrichment.dependency_rule_rows,
                        metric_candidates=self._metric_candidates_from_agentic_enrichment(enrichment),
                        relation_candidates=enrichment.candidate_relations,
                        warnings=enrichment.warnings,
                        metric_count=len(enrichment.metric_dictionary_rows),
                        dependency_count=len(enrichment.dependency_rule_rows),
                    )
                warning = f"Excel file {path.name} was read, but no month column or row-metric report layout was detected."
                return ParsedWorkbookArtifact(
                    document_id=document.id,
                    filename=document.filename,
                    file_type=document.file_type,
                    parser="unsupported_layout",
                    status="unsupported_layout",
                    warnings=[warning],
                )

            snapshot = self._single_contract_snapshot(
                contract,
                temp_dir=temp_dir,
                memory_priors=memory_priors,
                index=index,
            )
            return ParsedWorkbookArtifact(
                document_id=document.id,
                filename=document.filename,
                file_type=document.file_type,
                parser="direct_month_facts",
                status="parsed_with_warnings" if direct_warnings else "parsed",
                contracts=[contract],
                metric_candidates=applied_metric_candidates,
                warnings=direct_warnings,
                dataset_count=len(snapshot.datasets),
                metric_count=len(snapshot.metrics),
                dependency_count=len(snapshot.dependencies),
            )
        except Exception as exc:
            LOGGER.exception("Failed to parse workbook %s", document.filename)
            return ParsedWorkbookArtifact(
                document_id=document.id,
                filename=document.filename,
                file_type=document.file_type,
                parser="parse_failed",
                status="parse_failed",
                warnings=[f"Failed to parse workbook {document.filename}: {exc}"],
            )

    def _metadata_artifact_from_workbook(self, *, path: Path, document: Any) -> ParsedWorkbookArtifact | None:
        headers, rows = _sheet_rows(path)
        header_set = {str(header or "").strip() for header in headers}
        if {"metric_code", "label"}.issubset(header_set):
            return ParsedWorkbookArtifact(
                document_id=document.id,
                filename=document.filename,
                file_type=document.file_type,
                parser="metadata_metric_dictionary",
                status="parsed",
                metric_dictionary_rows=rows,
                metric_count=len(rows),
            )
        if {"source_metric_code", "target_metric_code"}.issubset(header_set):
            return ParsedWorkbookArtifact(
                document_id=document.id,
                filename=document.filename,
                file_type=document.file_type,
                parser="metadata_dependency_rules",
                status="parsed",
                dependency_rule_rows=rows,
                dependency_count=len(rows),
            )
        return None

    def _dynamic_numeric_facts_contract(self, pivoted_facts_path: Path, index: int) -> dict[str, Any]:
        headers, _ = _sheet_rows(pivoted_facts_path, "pivoted_facts")
        dimensions = [header for header in headers if header in DIMENSION_COLUMNS]
        metrics = [header for header in headers if header not in dimensions]
        return {
            "key": f"upload_{index}_{normalize_metric_code(pivoted_facts_path.stem)}",
            "path": str(pivoted_facts_path),
            "sheet_name": "pivoted_facts",
            "required": True,
            "kind": "numeric_facts",
            "grain": dimensions,
            "required_columns": [*dimensions, *metrics],
            "derives": {"metrics": metrics, "dimensions": dimensions},
        }

    def _read_metadata_rows(self, path: Path) -> list[dict[str, Any]]:
        _, rows = _sheet_rows(path)
        return rows

    def _single_contract_snapshot(
        self,
        contract: dict[str, Any],
        *,
        temp_dir: Path,
        memory_priors: list[dict[str, Any]],
        index: int,
    ) -> RelationMemorySnapshot:
        manifest = {
            "version": 1,
            "poc_id": f"relation-memory-chat-upload-single-{index}",
            "status": "runtime",
            "incoming_csvs": [contract],
            "golden_queries": [],
        }
        manifest_path = temp_dir / f"single_manifest_{index}.yaml"
        manifest_path.write_text(yaml.safe_dump(manifest, allow_unicode=True, sort_keys=False), encoding="utf-8")
        return RelationMemoryPocBuilder(manifest_path, dependency_priors=memory_priors).build_snapshot()

    def _build_combined_snapshot(
        self,
        *,
        parsed_artifacts: list[ParsedWorkbookArtifact],
        temp_dir: Path,
        memory_priors: list[dict[str, Any]],
    ) -> RelationMemorySnapshot:
        incoming_csvs = [contract for artifact in parsed_artifacts for contract in artifact.contracts]
        metric_rows = [row for artifact in parsed_artifacts for row in artifact.metric_dictionary_rows]
        dependency_rows = [row for artifact in parsed_artifacts for row in artifact.dependency_rule_rows]
        if metric_rows:
            metric_dictionary_path = temp_dir / "combined_metric_dictionary.xlsx"
            self._write_rows_xlsx(
                metric_dictionary_path,
                "metric_dictionary",
                [
                    "metric_code",
                    "label",
                    "description",
                    "aliases",
                    "sensitivity_level",
                    "allow_roles",
                    "preferred_dataset",
                    "semantic_type",
                ],
                metric_rows,
            )
            incoming_csvs.append(
                {
                    "key": "metric_dictionary",
                    "path": str(metric_dictionary_path),
                    "required": False,
                    "kind": "metadata",
                    "required_columns": [
                        "metric_code",
                        "label",
                        "description",
                        "aliases",
                        "sensitivity_level",
                        "allow_roles",
                        "preferred_dataset",
                    ],
                }
            )
        if dependency_rows:
            dependency_rules_path = temp_dir / "combined_dependency_rules.xlsx"
            self._write_rows_xlsx(
                dependency_rules_path,
                "dependency_rules",
                ["source_metric_code", "target_metric_code", "edge_type", "reason", "strength"],
                dependency_rows,
            )
            incoming_csvs.append(
                {
                    "key": "dependency_rules",
                    "path": str(dependency_rules_path),
                    "required": False,
                    "kind": "metadata",
                    "required_columns": [
                        "source_metric_code",
                        "target_metric_code",
                        "edge_type",
                        "reason",
                        "strength",
                    ],
                }
            )
        manifest = {
            "version": 1,
            "poc_id": "relation-memory-chat-upload",
            "status": "runtime",
            "incoming_csvs": incoming_csvs,
            "golden_queries": [],
        }
        manifest_path = temp_dir / "manifest.yaml"
        manifest_path.write_text(yaml.safe_dump(manifest, allow_unicode=True, sort_keys=False), encoding="utf-8")
        return RelationMemoryPocBuilder(
            manifest_path,
            dependency_priors=memory_priors,
        ).build_snapshot()

    def _write_rows_xlsx(self, path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name[:31]
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        workbook.save(path)
        workbook.close()

    def _try_build_agentic_enrichment(
        self,
        *,
        path: Path,
        document: Any,
    ) -> AgenticWorkbookEnrichmentResult | None:
        try:
            return self.agentic_enricher.enrich_workbook(
                path,
                document_id=document.id,
                filename=document.filename,
            )
        except Exception as exc:
            return AgenticWorkbookEnrichmentResult(
                warnings=[f"Agentic workbook enrichment skipped for {document.filename}: {exc}"]
            )

    def _has_agentic_enrichment_payload(self, enrichment: AgenticWorkbookEnrichmentResult) -> bool:
        return bool(
            enrichment.metric_dictionary_rows
            or enrichment.dependency_rule_rows
            or enrichment.candidate_relations
        )

    def _metric_candidates_from_agentic_enrichment(
        self,
        enrichment: AgenticWorkbookEnrichmentResult,
    ) -> list[MetricCandidate]:
        return [self._metric_candidate_from_agentic_metric(item) for item in enrichment.metrics]

    def _metric_candidate_from_agentic_metric(self, item: dict[str, Any]) -> MetricCandidate:
        canonical_code = str(item.get("code") or "")
        raw_label = str(item.get("label") or canonical_code)
        aliases = [str(alias).strip() for alias in item.get("aliases", []) if str(alias).strip()]
        metric_id = normalize_metric_code(
            "__".join(
                part
                for part in [
                    canonical_code,
                    str(item.get("source_sheet") or ""),
                    raw_label,
                ]
                if part
            )
        )[:160]
        return MetricCandidate(
            id=metric_id,
            metric_id=metric_id,
            canonical_code=canonical_code,
            raw_label=raw_label,
            label=raw_label,
            aliases=aliases,
            unit="",
            department="generic",
            source_sheet=str(item.get("source_sheet") or ""),
            section_path=str(item.get("source_scan_unit_id") or ""),
            semantic_type="unknown",
            aggregation="sum",
            status="proposed",
            confidence=float(item.get("confidence") or 0.8),
            evidence="agentic workbook scan metric candidate",
            approved=False,
        )

    def _try_build_dynamic_snapshot(
        self,
        path: Path,
        temp_dir: str | Path,
        memory_priors: list[dict[str, Any]],
        metric_mapping_priors: list[dict[str, Any]],
    ) -> tuple[RelationMemorySnapshot, list[MetricCandidate], list[dict[str, Any]], list[str], Any] | None:
        domain_pack = DomainPack.load("generic")
        profiler = WorkbookProfiler(domain_pack=domain_pack)
        profile = profiler.profile_workbook(path)
        if not profiler.has_row_metric_layout(profile):
            return None
        output_dir = Path(temp_dir) / "dynamic_normalized"
        semantic_resolver = self._ollama_semantic_resolver()
        relation_judge = self._ollama_relation_judge()
        artifacts = NormalizationAdapter(
            domain_pack=domain_pack,
            memory_priors=memory_priors,
            metric_mapping_priors=metric_mapping_priors,
            metric_semantic_resolver=semantic_resolver,
            relation_semantic_judge=relation_judge,
        ).normalize_workbook(
            path,
            output_dir,
            department="generic",
            source_workbook=path.name,
        )
        snapshot = RelationMemoryPocBuilder(
            artifacts.generated_manifest_path,
            dependency_priors=memory_priors,
        ).build_snapshot()
        metric_candidates = [self._metric_candidate_from_artifact(item) for item in artifacts.metric_candidates]
        relation_candidates = self._select_dynamic_relation_candidates(artifacts.relation_candidates)
        return snapshot, metric_candidates, relation_candidates, artifacts.warnings, artifacts

    def _ollama_semantic_resolver(self) -> OllamaMetricSemanticResolver | None:
        if not settings.RELATION_MEMORY_OLLAMA_ENABLED:
            return None
        return OllamaMetricSemanticResolver(
            model=settings.RELATION_MEMORY_OLLAMA_MODEL,
            base_url=settings.RELATION_MEMORY_OLLAMA_URL,
            timeout_seconds=settings.RELATION_MEMORY_OLLAMA_TIMEOUT_SECONDS,
            batch_size=settings.RELATION_MEMORY_OLLAMA_BATCH_SIZE,
        )

    def _ollama_relation_judge(self) -> OllamaRelationSemanticJudge | None:
        if not settings.RELATION_MEMORY_OLLAMA_ENABLED or not settings.RELATION_MEMORY_OLLAMA_RELATION_GATE_ENABLED:
            return None
        return OllamaRelationSemanticJudge(
            model=settings.RELATION_MEMORY_OLLAMA_MODEL,
            base_url=settings.RELATION_MEMORY_OLLAMA_URL,
            timeout_seconds=settings.RELATION_MEMORY_OLLAMA_TIMEOUT_SECONDS,
            batch_size=max(1, min(settings.RELATION_MEMORY_OLLAMA_BATCH_SIZE, 12)),
        )

    def _select_dynamic_relation_candidates(self, relations: list[dict[str, Any]]) -> list[dict[str, Any]]:
        pending_candidates = [
            item
            for item in relations
            if item.get("needs_approval") or item.get("evidence_type") in {"statistical", "row_structure", "text"}
        ]
        pending_candidates.sort(
            key=lambda item: (
                -float(item.get("score") or item.get("confidence") or 0.0),
                str(item.get("evidence_type") or ""),
                str(item.get("source_metric_code") or ""),
                str(item.get("target_metric_code") or ""),
            )
        )
        selected = [
            item
            for item in pending_candidates
            if float(item.get("score") or item.get("confidence") or 0.0) >= DYNAMIC_RELATION_MIN_PENDING_SCORE
        ][:DYNAMIC_RELATION_CANDIDATE_LIMIT]
        if not selected and pending_candidates:
            selected = pending_candidates[: min(5, DYNAMIC_RELATION_CANDIDATE_LIMIT)]
        return [self._candidate_relation_payload_from_dynamic(item) for item in selected]

    def _metric_candidate_from_artifact(self, item: dict[str, Any]) -> MetricCandidate:
        aliases = [
            alias.strip()
            for alias in str(item.get("aliases") or "").split("|")
            if alias.strip()
        ]
        metric_id = str(item.get("metric_id") or item.get("canonical_code") or uuid.uuid4())
        status = str(item.get("status") or "proposed")
        return MetricCandidate(
            id=normalize_metric_code(metric_id),
            metric_id=metric_id,
            canonical_code=str(item.get("canonical_code") or item.get("metric_code") or ""),
            raw_label=str(item.get("raw_label") or item.get("label") or ""),
            label=str(item.get("label") or item.get("raw_label") or ""),
            aliases=aliases,
            unit=str(item.get("unit") or ""),
            department=str(item.get("department") or "generic"),
            source_sheet=str(item.get("source_sheet") or ""),
            section_path=str(item.get("section_path") or ""),
            semantic_type=str(item.get("semantic_type") or "unknown"),
            aggregation=str(item.get("aggregation") or "sum"),
            status=status,
            confidence=float(item.get("confidence") or 0.0),
            evidence=str(item.get("evidence") or ""),
            approved=status == "memory_applied",
        )

    def _rewrite_month_based_xlsx(
        self,
        path: Path,
        *,
        index: int,
        temp_dir: Path,
        metric_mapping_priors: list[dict[str, Any]],
    ) -> tuple[Path, list[MetricCandidate], list[str]]:
        mapping_lookup = self._metric_mapping_lookup(metric_mapping_priors)
        if not mapping_lookup:
            return path, [], []

        workbook = load_workbook(path, read_only=True, data_only=True)
        rewritten = Workbook()
        rewritten_sheet = rewritten.active
        warnings: list[str] = []
        applied_metric_candidates: dict[str, MetricCandidate] = {}
        try:
            source_sheet = workbook[workbook.sheetnames[0]]
            rows = list(source_sheet.iter_rows(values_only=True))
            if not rows:
                return path, [], []
            headers = list(rows[0])
            rewritten_headers = []
            for header in headers:
                header_text = str(header or "").strip()
                mapped = mapping_lookup.get(self._normalize_mapping_alias(header_text))
                if mapped and _canonical_header(header_text) not in DIMENSION_COLUMNS:
                    rewritten_headers.append(mapped["canonical_code"])
                    candidate = self._metric_candidate_from_mapping_prior(mapped)
                    applied_metric_candidates[candidate.id] = candidate
                    warnings.append(
                        f"Excel file {path.name}: applied approved metric mapping '{header_text}' -> '{mapped['canonical_code']}'."
                    )
                else:
                    rewritten_headers.append(header)
            rewritten_sheet.append(rewritten_headers)
            for row in rows[1:]:
                rewritten_sheet.append(list(row))
        finally:
            workbook.close()
        if not warnings:
            rewritten.close()
            return path, [], []
        rewritten_path = temp_dir / f"direct_month_{index}_{normalize_metric_code(path.stem)}.xlsx"
        rewritten.save(rewritten_path)
        rewritten.close()
        return rewritten_path, list(applied_metric_candidates.values()), warnings

    def _metric_mapping_lookup(self, priors: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        lookup: dict[str, dict[str, Any]] = {}
        for prior in priors:
            canonical_code = str(prior.get("canonical_code") or prior.get("metric_code") or "").strip()
            if not canonical_code:
                continue
            aliases = [
                str(prior.get("raw_label") or "").strip(),
                str(prior.get("label") or "").strip(),
                canonical_code,
                canonical_code.replace("_", " "),
                *self._mapping_aliases(prior.get("aliases")),
            ]
            for alias in aliases:
                normalized = self._normalize_mapping_alias(alias)
                if normalized:
                    lookup.setdefault(normalized, prior)
        return lookup

    def _normalize_mapping_alias(self, value: str) -> str:
        return normalize_metric_code(value).strip("_")

    def _metric_candidate_from_mapping_prior(self, prior: dict[str, Any]) -> MetricCandidate:
        aliases = [
            alias
            for alias in [
                str(prior.get("raw_label") or "").strip(),
                str(prior.get("label") or "").strip(),
                *self._mapping_aliases(prior.get("aliases")),
            ]
            if alias
        ]
        metric_id = self._metric_mapping_candidate_id(prior)
        return MetricCandidate(
            id=metric_id,
            metric_id=metric_id,
            canonical_code=str(prior.get("canonical_code") or prior.get("metric_code") or ""),
            raw_label=str(prior.get("raw_label") or prior.get("label") or ""),
            label=str(prior.get("label") or prior.get("raw_label") or ""),
            aliases=list(dict.fromkeys(aliases)),
            unit=str(prior.get("unit") or ""),
            department=str(prior.get("department") or "generic"),
            source_sheet=str(prior.get("source_sheet") or ""),
            section_path=str(prior.get("section_path") or ""),
            semantic_type=str(prior.get("semantic_type") or "unknown"),
            aggregation=str(prior.get("aggregation") or "sum"),
            status="memory_applied",
            confidence=1.0,
            evidence=str(prior.get("evidence") or "approved metric mapping memory"),
            approved=True,
        )

    def _metric_mapping_candidate_id(self, prior: dict[str, Any]) -> str:
        canonical_code = str(prior.get("canonical_code") or prior.get("metric_code") or "")
        source_sheet = str(prior.get("source_sheet") or "")
        raw_label = str(prior.get("raw_label") or prior.get("label") or "")
        return normalize_metric_code(f"{canonical_code}__{source_sheet}__{raw_label}")[:160]

    def _candidate_relation_payload_from_dynamic(self, item: dict[str, Any]) -> dict[str, Any]:
        relation_type = str(item.get("relation_type") or item.get("edge_type") or "driver")
        edge_type = str(item.get("edge_type") or "driver")
        if edge_type not in {"driver", "inverse_driver", "component", "lag"}:
            edge_type = "component"
        return {
            "source_metric_code": item.get("source_metric_code"),
            "target_metric_code": item.get("target_metric_code"),
            "edge_type": edge_type,
            "relation_type": relation_type,
            "note": item.get("evidence") or "",
            "evidence": item.get("evidence") or "",
            "evidence_type": item.get("evidence_type") or "dynamic",
            "confidence": item.get("confidence") or item.get("score") or 0.5,
            "score": item.get("score") or item.get("confidence") or 0.5,
            "needs_approval_reason": item.get("needs_approval_reason") or "",
            "source": f"dynamic_{item.get('evidence_type') or 'relation'}",
            "source_document_id": item.get("source_document_id"),
        }

    def _prepare_xlsx_for_snapshot(
        self,
        path: Path,
        index: int,
        temp_dir: Path,
        bundle: DocumentBundle,
        inferred_dependency_priors: list[dict[str, Any]],
    ) -> Path:
        contract = self._contract_for_xlsx(path, index)
        if "month" in contract.get("grain", []):
            return path

        normalized_path = normalize_row_metric_xlsx(path, index, temp_dir, inferred_dependency_priors)
        if normalized_path is None:
            bundle.warnings.append(
                f"Excel file {path.name} was read, but no month column or row-metric report layout was detected."
            )
            return path

        bundle.warnings.append(
            f"Excel file {path.name} was converted from row-metric report layout into month-based facts."
        )
        return normalized_path

    def _contract_for_xlsx(self, path: Path, index: int) -> dict[str, Any]:
        headers, rows = _sheet_rows(path)
        canonical_headers = []
        for header in headers:
            canonical = _canonical_header(header)
            if canonical == str(header).strip():
                canonical = _canonical_metric_code(header)
            canonical_headers.append(canonical)
        dimensions = [header for header in canonical_headers if header in DIMENSION_COLUMNS]
        metrics = [
            header
            for header in canonical_headers
            if header not in dimensions and column_has_numeric_values(header, headers, rows)
        ]
        if not metrics:
            metrics = [header for header in canonical_headers if header not in dimensions]
        dataset_key = f"upload_{index}_{normalize_metric_code(path.stem)}"
        return {
            "key": dataset_key,
            "path": str(path),
            "required": True,
            "kind": "numeric_facts",
            "grain": dimensions,
            "required_columns": [*dimensions, *metrics],
            "derives": {
                "metrics": metrics,
                "dimensions": dimensions,
            },
        }

    def _build_metric_map(
        self,
        snapshot: RelationMemorySnapshot,
        bundle: DocumentBundle,
        *,
        parsed_document_ids: set[str],
        metric_mapping_priors: list[dict[str, Any]],
    ) -> dict[str, dict[str, Any]]:
        return build_metric_map(
            snapshot_metrics=snapshot.metrics,
            detected_metrics=bundle.detected_metrics,
            parsed_document_ids=parsed_document_ids,
            metric_mapping_priors=metric_mapping_priors,
            mapping_aliases=self._mapping_aliases,
        )

    def _candidate_from_raw(
        self,
        raw_candidate: dict[str, Any],
        state: RelationMemorySessionState,
        *,
        source: str,
    ) -> CandidateRelation | None:
        normalized = normalize_candidate_payload(raw_candidate, default_source=source)
        if not normalized:
            state.warnings.append("Candidate relation ignored because it failed schema validation.")
            return None

        for metric_code in (normalized["source_metric_code"], normalized["target_metric_code"]):
            ensure_metric_entry(
                state.metric_map,
                metric_code,
                label=metric_default_label(metric_code),
                aliases=[metric_code, metric_code.replace("_", " ")],
                source_labels=[metric_default_label(metric_code)],
                source=source,
            )

        return CandidateRelation(id=candidate_id(normalized), **normalized)

    def _add_candidate_if_new(
        self,
        state: RelationMemorySessionState,
        candidate: CandidateRelation,
    ) -> CandidateRelation | None:
        existing_ids = set(state.pending_confirmations)
        self._add_candidate(state, candidate)
        if candidate.id in existing_ids:
            return None
        return state.pending_confirmations.get(candidate.id)

    def _add_candidate(self, state: RelationMemorySessionState, raw_candidate: dict[str, Any] | CandidateRelation) -> None:
        if isinstance(raw_candidate, CandidateRelation):
            candidate = raw_candidate
        else:
            candidate = self._candidate_from_raw(raw_candidate, state, source=str(raw_candidate.get("source") or "candidate"))
            if candidate is None:
                return
        if candidate.source_metric_code == candidate.target_metric_code:
            return
        if candidate.id in state.rejected_candidate_ids:
            return
        if relation_key(candidate) in state.confirmed_relation_keys:
            return
        self._validate_candidate_metrics(candidate, state)
        state.pending_confirmations.setdefault(candidate.id, candidate)

    def _validate_candidate_metrics(self, candidate: CandidateRelation, state: RelationMemorySessionState) -> None:
        missing = [
            metric_code
            for metric_code in (candidate.source_metric_code, candidate.target_metric_code)
            if metric_code not in state.metric_map
        ]
        if missing:
            raise ValueError(f"Candidate references unknown metrics: {', '.join(missing)}")

    def _relation_memory_priors(self, previous_state: RelationMemorySessionState | None) -> list[dict[str, Any]]:
        priors = list(self._get_memory_priors())
        if previous_state:
            priors.extend(relation.model_dump() for relation in previous_state.approved_relations)
        deduped: dict[tuple[str, str, str], dict[str, Any]] = {}
        for prior in priors:
            deduped[relation_key(prior)] = prior
        return list(deduped.values())

    def _metric_mapping_priors(self, previous_state: RelationMemorySessionState | None) -> list[dict[str, Any]]:
        rejected_ids = set(previous_state.rejected_metric_candidate_ids) if previous_state else set()
        priors = []
        for prior in self._get_metric_mapping_priors():
            if self._metric_mapping_candidate_id(prior) in rejected_ids:
                continue
            priors.append(prior)
        if previous_state:
            for candidate in previous_state.metric_candidates.values():
                if not candidate.approved:
                    continue
                prior = self._metric_mapping_prior_payload(candidate)
                if self._metric_mapping_candidate_id(prior) in rejected_ids:
                    continue
                priors.append(prior)
        deduped: dict[str, dict[str, Any]] = {}
        for prior in priors:
            deduped[self._metric_mapping_candidate_id(prior)] = prior
        return list(deduped.values())

    def _metric_mapping_prior_payload(self, candidate: MetricCandidate) -> dict[str, Any]:
        return {
            "canonical_code": candidate.canonical_code,
            "metric_code": candidate.canonical_code,
            "raw_label": candidate.raw_label,
            "label": candidate.label or candidate.raw_label,
            "aliases": list(candidate.aliases),
            "unit": candidate.unit,
            "department": candidate.department,
            "source_sheet": candidate.source_sheet,
            "section_path": candidate.section_path,
            "semantic_type": candidate.semantic_type,
            "aggregation": candidate.aggregation,
            "confidence": candidate.confidence,
            "evidence": candidate.evidence,
        }

    def _mapping_aliases(self, aliases: Any) -> list[str]:
        if isinstance(aliases, str):
            return [alias.strip() for alias in aliases.split("|") if alias.strip()]
        return [str(alias).strip() for alias in aliases or [] if str(alias).strip()]

    def _unique_warnings(self, warnings: list[str]) -> list[str]:
        return list(dict.fromkeys(warning for warning in warnings if warning))

    def _persist_confirmed_metric_aliases(
        self,
        *,
        state: RelationMemorySessionState,
        aliases: list[dict[str, str]],
        session_id: str,
    ) -> None:
        if not aliases:
            return
        for alias_payload in aliases:
            canonical_code = str(alias_payload.get("canonical_code") or "").strip()
            alias = str(alias_payload.get("alias") or "").strip()
            if not canonical_code or not alias:
                continue
            metric_entry = ensure_metric_entry(
                state.metric_map,
                canonical_code,
                label=str(alias_payload.get("label") or canonical_code),
                source="user_confirmed_alias",
            )
            known_aliases = {
                normalize_metric_code(item)
                for item in [
                    canonical_code,
                    metric_entry.get("label") or "",
                    *(metric_entry.get("aliases") or []),
                    *(metric_entry.get("approved_aliases") or []),
                ]
                if str(item or "").strip()
            }
            if normalize_metric_code(alias) in known_aliases:
                continue
            extend_metric_values(metric_entry, "aliases", [alias])
            extend_metric_values(metric_entry, "approved_aliases", [alias])
            if hasattr(self._graph_client(), "save_metric_mapping"):
                self._graph_client().save_metric_mapping(
                    canonical_code=canonical_code,
                    raw_label=alias,
                    label=str(metric_entry.get("label") or alias_payload.get("label") or alias),
                    aliases=list(
                        dict.fromkeys(
                            [
                                alias,
                                str(metric_entry.get("label") or alias_payload.get("label") or alias),
                                canonical_code,
                            ]
                        )
                    ),
                    department="conversation",
                    source_sheet="user_language",
                    section_path="clarification",
                    semantic_type=str(metric_entry.get("semantic_type") or "unknown"),
                    aggregation="sum",
                    confidence=1.0,
                    evidence="user_confirmed_alias",
                    source_session_id=session_id,
                )

    def _get_memory_priors(self) -> list[dict[str, Any]]:
        graph_client = self._graph_client()
        if hasattr(graph_client, "get_memory_priors"):
            return list(graph_client.get_memory_priors())
        return []

    def _get_metric_mapping_priors(self) -> list[dict[str, Any]]:
        graph_client = self._graph_client()
        if hasattr(graph_client, "get_metric_mapping_priors"):
            return list(graph_client.get_metric_mapping_priors())
        return []

    def _get_session(self, session_id: str) -> RelationMemorySessionState:
        try:
            return self.sessions[session_id]
        except KeyError as exc:
            raise KeyError(f"Unknown relation memory session: {session_id}") from exc

    def _graph_client(self):
        if self.graph_client is None:
            self.graph_client = RelationMemoryNeo4jClient()
        return self.graph_client
