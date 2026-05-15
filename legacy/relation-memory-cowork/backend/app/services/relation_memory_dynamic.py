from __future__ import annotations

import json
import math
import re
import urllib.error
import urllib.request
import uuid
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook

from app.services.relation_memory_dynamic_utils import (
    BASE_DIR,
    DOMAIN_PACK_DIR,
    LONG_FACT_HEADERS,
    METRIC_CANDIDATE_HEADERS,
    PIVOT_DIMENSIONS,
    RELATION_CANDIDATE_HEADERS,
    RELATION_GATE_CONTEXT_LIMIT,
    RELATION_GATE_DECISIONS,
    RELATION_GATE_EDGE_TYPES,
    ROW_METRIC_HEADERS,
    RUSSIAN_MONTHS,
    SCENARIO_PATTERNS,
    SECTION_MARKER_PATTERNS,
    SEMANTIC_TYPES,
    TECHNICAL_DEFAULT_PATTERNS,
    DynamicRelationMemoryArtifacts,
    _append_reason,
    aggregation_for_metric,
    edge_type_for_relation,
    formula_signature,
    looks_like_metric_label,
    looks_like_section_marker,
    normalize_text,
    period_from_value,
    read_yaml_file,
    scenario_from_text,
    titleize_code,
)
from app.services.relation_memory_ingestion import normalize_metric_code
from app.services.relation_memory_poc import _canonical_header, _safe_float


class DomainPack:
    def __init__(self, name: str = "generic", payload: dict[str, Any] | None = None):
        self.name = name or "generic"
        self.payload = payload or {}
        self.aliases = {
            normalize_text(alias): str(code)
            for code, aliases in (self.payload.get("aliases") or {}).items()
            for alias in ([code] + list(aliases or []))
        }
        self.technical_row_patterns = list(self.payload.get("technical_row_patterns") or [])
        self.target_metric_patterns = list(self.payload.get("target_metric_patterns") or [])
        self.semantic_type_patterns = dict(self.payload.get("semantic_type_patterns") or {})
        self.relation_templates = list(self.payload.get("relation_templates") or [])
        self.expected_dimensions = list(self.payload.get("expected_dimensions") or [])
        self.external_context_suggestions = list(self.payload.get("external_context_suggestions") or [])
        self.unit_patterns = list(self.payload.get("unit_patterns") or [])

    @classmethod
    def load(cls, name: str | None = None) -> "DomainPack":
        pack_name = name or "generic"
        generic_payload = read_yaml_file(DOMAIN_PACK_DIR / "generic.yaml")
        selected_payload = read_yaml_file(DOMAIN_PACK_DIR / f"{pack_name}.yaml") if pack_name != "generic" else {}
        payload = cls._merge_payloads(generic_payload, selected_payload)
        return cls(pack_name, payload)

    @staticmethod
    def _merge_payloads(base: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
        merged = dict(base)
        for key, value in override.items():
            if isinstance(value, dict) and isinstance(merged.get(key), dict):
                merged[key] = {**merged[key], **value}
            elif isinstance(value, list) and isinstance(merged.get(key), list):
                merged[key] = [*merged[key], *value]
            else:
                merged[key] = value
        return merged

    def is_technical_label(self, label: str) -> bool:
        normalized = normalize_text(label)
        patterns = [*TECHNICAL_DEFAULT_PATTERNS, *self.technical_row_patterns]
        return any(re.search(pattern, normalized) for pattern in patterns)

    def canonical_code_for_label(self, label: str) -> tuple[str, float, str]:
        normalized = normalize_text(label)
        if normalized in self.aliases:
            return self.aliases[normalized], 0.95, "domain_alias"
        for alias, code in self.aliases.items():
            if alias and self._is_safe_partial_alias_match(normalized, alias):
                return code, 0.82, "partial_domain_alias"
        return normalize_metric_code(label), 0.55, "generated_from_label"

    def _is_safe_partial_alias_match(self, normalized_label: str, alias: str) -> bool:
        if normalized_label == alias:
            return True
        if len(alias) < 5:
            return False
        if not normalized_label.startswith(alias):
            return False
        next_char = normalized_label[len(alias) : len(alias) + 1]
        if next_char and next_char not in {" ", ",", ".", ":", ";", ")", "(", "/"}:
            return False
        # Composite labels like "ремонты+амортизация+лизинг" should not be
        # collapsed to the first short alias unless the domain pack has an
        # exact mapping for the whole phrase.
        if "+" in normalized_label and len(alias) < 14:
            return False
        return True

    def semantic_type_for_label(self, label: str, canonical_code: str) -> str:
        normalized = normalize_text(label)
        if self.is_technical_label(label):
            return "technical_check"
        for semantic_type, patterns in self.semantic_type_patterns.items():
            if any(re.search(pattern, normalized) or re.search(pattern, canonical_code) for pattern in patterns or []):
                return semantic_type
        if any(re.search(pattern, normalized) for pattern in self.target_metric_patterns):
            return "target_metric"
        if any(token in normalized for token in ("пробег", "накладн", "количество", "кол-во", "шт")):
            return "denominator"
        if any(token in normalized for token in ("индекс", "ставка", "погода", "дизель")):
            return "external_context"
        return "unknown"


class OllamaMetricSemanticResolver:
    def __init__(
        self,
        *,
        model: str = "qwen3:14b",
        base_url: str = "http://localhost:11434",
        timeout_seconds: float = 90.0,
        batch_size: int = 24,
    ):
        self.model = model
        self.base_url = base_url.rstrip("/")
        self.timeout_seconds = timeout_seconds
        self.batch_size = max(1, int(batch_size or 1))

    def resolve_batch(
        self,
        metric_contexts: list[dict[str, Any]],
        *,
        known_metric_codes: list[str],
    ) -> tuple[dict[str, dict[str, Any]], list[str]]:
        suggestions: dict[str, dict[str, Any]] = {}
        warnings: list[str] = []
        for batch_start in range(0, len(metric_contexts), self.batch_size):
            batch = metric_contexts[batch_start : batch_start + self.batch_size]
            if not batch:
                continue
            try:
                payload = self._generate_json(self._prompt(batch, known_metric_codes))
            except Exception as exc:  # noqa: BLE001 - resolver is optional; fallback must be deterministic.
                warnings.append(f"Ollama semantic resolver failed for batch {batch_start // self.batch_size + 1}: {exc}")
                continue
            for item in self._items_from_payload(payload):
                key = str(item.get("key") or "").strip()
                if not key:
                    continue
                normalized = self._normalize_suggestion(item)
                if normalized:
                    suggestions[key] = normalized
        return suggestions, warnings

    def _generate_json(self, prompt: str) -> dict[str, Any]:
        request = urllib.request.Request(
            f"{self.base_url}/api/generate",
            data=json.dumps(
                {
                    "model": self.model,
                    "prompt": prompt,
                    "stream": False,
                    "format": "json",
                    "options": {"temperature": 0, "num_ctx": 8192},
                },
                ensure_ascii=False,
            ).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=self.timeout_seconds) as response:
                body = json.loads(response.read().decode("utf-8"))
        except urllib.error.URLError as exc:
            raise RuntimeError(f"cannot reach Ollama at {self.base_url}: {exc}") from exc
        return self._parse_json_response(str(body.get("response") or ""))

    def _parse_json_response(self, text: str) -> dict[str, Any]:
        cleaned = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL | re.IGNORECASE).strip()
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
            if not match:
                raise ValueError("Ollama returned no JSON object")
            return json.loads(match.group(0))

    def _items_from_payload(self, payload: dict[str, Any]) -> list[dict[str, Any]]:
        if isinstance(payload.get("items"), list):
            return [item for item in payload["items"] if isinstance(item, dict)]
        if isinstance(payload.get("metrics"), list):
            return [item for item in payload["metrics"] if isinstance(item, dict)]
        return [payload] if payload.get("key") else []

    def _normalize_suggestion(self, item: dict[str, Any]) -> dict[str, Any] | None:
        semantic_type = str(item.get("semantic_type") or "unknown").strip()
        if semantic_type not in SEMANTIC_TYPES:
            semantic_type = "unknown"
        raw_code = str(item.get("canonical_code") or item.get("metric_code") or "").strip()
        if not raw_code:
            return None
        code = normalize_metric_code(raw_code)
        confidence = _safe_float(item.get("confidence"))
        confidence = max(0.0, min(1.0, float(confidence if confidence is not None else 0.0)))
        aliases = item.get("aliases")
        if isinstance(aliases, list):
            alias_text = "|".join(str(alias).strip() for alias in aliases if str(alias).strip())
        else:
            alias_text = str(aliases or "").strip()
        return {
            "key": str(item.get("key") or "").strip(),
            "canonical_code": code,
            "semantic_type": semantic_type,
            "label": str(item.get("label") or "").strip(),
            "aliases": alias_text,
            "confidence": round(confidence, 3),
            "reason": str(item.get("reason") or "").strip()[:500],
            "model": self.model,
        }

    def _prompt(self, metric_contexts: list[dict[str, Any]], known_metric_codes: list[str]) -> str:
        compact_contexts = [
            {
                "key": item["key"],
                "raw_label": item["raw_label"],
                "unit": item.get("unit") or "",
                "source_sheet": item.get("source_sheet") or "",
                "section_path": item.get("section_path") or "",
                "baseline_code": item.get("baseline_code") or "",
                "baseline_semantic_type": item.get("baseline_semantic_type") or "unknown",
                "baseline_evidence": item.get("baseline_evidence") or "",
            }
            for item in metric_contexts
        ]
        return (
            "You classify metric rows from a Russian FTL cost Excel workbook. "
            "Return valid JSON only. Do not create metric relationships.\n"
            "For every input item return one object with the same key.\n"
            "Allowed semantic_type values: target_metric, cost_component, business_driver, denominator, "
            "external_context, technical_check, unknown.\n"
            "canonical_code must be stable snake_case. Prefer known_metric_codes when the label clearly matches. "
            "If uncertain, keep the baseline_code and lower confidence.\n"
            "technical checks/control rows must use semantic_type=technical_check.\n"
            "Output schema: {\"items\":[{\"key\":\"...\",\"canonical_code\":\"...\",\"label\":\"...\","
            "\"aliases\":[\"...\"],\"semantic_type\":\"...\",\"confidence\":0.0,\"reason\":\"short Russian explanation\"}]}.\n"
            f"known_metric_codes={json.dumps(known_metric_codes, ensure_ascii=False)}\n"
            f"items={json.dumps(compact_contexts, ensure_ascii=False)}"
        )


class OllamaRelationSemanticJudge:
    def __init__(
        self,
        *,
        model: str = "qwen3:14b",
        base_url: str = "http://localhost:11434",
        timeout_seconds: float = 90.0,
        batch_size: int = 12,
    ):
        self.model = model
        self.base_url = base_url.rstrip("/")
        self.timeout_seconds = timeout_seconds
        self.batch_size = max(1, int(batch_size or 1))

    def judge_batch(self, relation_contexts: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], list[str]]:
        decisions: dict[str, dict[str, Any]] = {}
        warnings: list[str] = []
        for batch_start in range(0, len(relation_contexts), self.batch_size):
            batch = relation_contexts[batch_start : batch_start + self.batch_size]
            if not batch:
                continue
            try:
                payload = self._generate_json(self._prompt(batch))
            except Exception as exc:  # noqa: BLE001 - relation gate is optional and must fail closed.
                warnings.append(f"Ollama relation gate failed for batch {batch_start // self.batch_size + 1}: {exc}")
                continue
            for item in self._items_from_payload(payload):
                key = str(item.get("key") or "").strip()
                normalized = self._normalize_decision(item)
                if key and normalized:
                    decisions[key] = normalized
        return decisions, warnings

    def _generate_json(self, prompt: str) -> dict[str, Any]:
        request = urllib.request.Request(
            f"{self.base_url}/api/generate",
            data=json.dumps(
                {
                    "model": self.model,
                    "prompt": prompt,
                    "stream": False,
                    "format": "json",
                    "options": {"temperature": 0, "num_ctx": 8192},
                },
                ensure_ascii=False,
            ).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=self.timeout_seconds) as response:
                body = json.loads(response.read().decode("utf-8"))
        except urllib.error.URLError as exc:
            raise RuntimeError(f"cannot reach Ollama at {self.base_url}: {exc}") from exc
        text = re.sub(r"<think>.*?</think>", "", str(body.get("response") or ""), flags=re.DOTALL | re.IGNORECASE).strip()
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            match = re.search(r"\{.*\}", text, flags=re.DOTALL)
            if not match:
                raise ValueError("Ollama returned no JSON object")
            return json.loads(match.group(0))

    def _items_from_payload(self, payload: dict[str, Any]) -> list[dict[str, Any]]:
        if isinstance(payload.get("relations"), list):
            return [item for item in payload["relations"] if isinstance(item, dict)]
        if isinstance(payload.get("items"), list):
            return [item for item in payload["items"] if isinstance(item, dict)]
        return [payload] if payload.get("key") else []

    def _normalize_decision(self, item: dict[str, Any]) -> dict[str, Any] | None:
        decision = str(item.get("decision") or item.get("status") or "keep_pending").strip()
        if decision not in RELATION_GATE_DECISIONS:
            decision = "keep_pending"
        edge_type = str(item.get("edge_type") or "").strip()
        if edge_type not in RELATION_GATE_EDGE_TYPES:
            edge_type = ""
        confidence = _safe_float(item.get("confidence"))
        confidence = max(0.0, min(1.0, float(confidence if confidence is not None else 0.0)))
        return {
            "decision": decision,
            "edge_type": edge_type,
            "confidence": round(confidence, 3),
            "reason": str(item.get("reason") or item.get("note") or "").strip()[:500],
            "model": self.model,
        }

    def _prompt(self, relation_contexts: list[dict[str, Any]]) -> str:
        return (
            "You are a strict semantic gate for Russian management-report Excel metric relationships. "
            "Return valid JSON only. Do not invent metrics or numbers.\n"
            "Each input relation has source/target metric codes, labels, semantic types, and extraction evidence. "
            "Approve only if the direction and edge_type are economically meaningful. "
            "Reject comparison, variance, percentage, denominator-only, check/control, and reverse rollup references. "
            "Use keep_pending when evidence is plausible but not strong enough for automatic dependency_rules.\n"
            "Allowed decision values: approve, keep_pending, reject. "
            "Allowed edge_type values: component, driver, inverse_driver.\n"
            "Output schema: {\"relations\":[{\"key\":\"...\",\"decision\":\"approve|keep_pending|reject\","
            "\"edge_type\":\"component|driver|inverse_driver\",\"confidence\":0.0,\"reason\":\"short Russian explanation\"}]}.\n"
            f"relations={json.dumps(relation_contexts, ensure_ascii=False)}"
        )


class WorkbookProfiler:
    def __init__(self, *, domain_pack: DomainPack | None = None):
        self.domain_pack = domain_pack or DomainPack.load("generic")

    def profile_workbook(self, path: str | Path) -> dict[str, Any]:
        workbook_path = Path(path)
        workbook_values = load_workbook(workbook_path, data_only=True, read_only=False)
        workbook_formulas = load_workbook(workbook_path, data_only=False, read_only=False)
        warnings: list[str] = []
        sheets: list[dict[str, Any]] = []
        try:
            for values_sheet, formula_sheet in zip(workbook_values.worksheets, workbook_formulas.worksheets, strict=True):
                sheet_profile = self._profile_sheet(values_sheet, formula_sheet)
                if sheet_profile["duplicate_periods"]:
                    warnings.append(
                        f"Sheet {values_sheet.title} has duplicate period columns: "
                        f"{', '.join(sorted(sheet_profile['duplicate_periods']))}"
                    )
                sheets.append(sheet_profile)
        finally:
            workbook_values.close()
            workbook_formulas.close()
        return {
            "workbook_path": str(workbook_path),
            "workbook_name": workbook_path.name,
            "profiled_at": datetime.now(UTC).isoformat(),
            "sheets": sheets,
            "warnings": warnings,
        }

    def has_row_metric_layout(self, profile: dict[str, Any]) -> bool:
        return any(sheet.get("header_candidates") for sheet in profile.get("sheets", []))

    def _profile_sheet(self, values_sheet: Any, formula_sheet: Any) -> dict[str, Any]:
        header_candidates = self._find_header_candidates(values_sheet)
        best_header = header_candidates[0] if header_candidates else None
        metric_col = best_header["metric_col"] if best_header else None
        header_row = best_header["row_number"] if best_header else None
        period_columns = best_header["period_columns"] if best_header else []
        duplicate_periods = sorted(
            period for period, count in Counter(item["period"] for item in period_columns).items() if count > 1
        )
        candidate_metric_rows = (
            self._candidate_metric_rows(values_sheet, header_row, metric_col, period_columns)
            if header_row and metric_col
            else []
        )
        return {
            "sheet_name": values_sheet.title,
            "state": values_sheet.sheet_state,
            "max_row": values_sheet.max_row,
            "max_column": values_sheet.max_column,
            "header_candidates": header_candidates,
            "duplicate_periods": duplicate_periods,
            "candidate_metric_rows": candidate_metric_rows,
            "formula_cells": self._formula_cells(formula_sheet, metric_col),
            "comments": self._comments(values_sheet),
        }

    def _find_header_candidates(self, sheet: Any) -> list[dict[str, Any]]:
        candidates: list[dict[str, Any]] = []
        for row_idx in range(1, min(sheet.max_row or 0, 12) + 1):
            values = list(next(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)))
            metric_col = None
            for index, value in enumerate(values, start=1):
                if normalize_text(value) in ROW_METRIC_HEADERS:
                    metric_col = index
                    break
            if not metric_col:
                continue
            period_columns = []
            for index, value in enumerate(values, start=1):
                if index <= metric_col:
                    continue
                period = period_from_value(value)
                if period:
                    scenario, scenario_evidence = self._infer_scenario_for_period_column(sheet, row_idx, index)
                    period_columns.append(
                        {
                            "col": index,
                            "period": period,
                            "header_value": str(value),
                            "scenario": scenario,
                            "scenario_evidence": scenario_evidence,
                        }
                    )
            if len(period_columns) >= 2:
                candidates.append(
                    {
                        "row_number": row_idx,
                        "metric_col": metric_col,
                        "period_columns": period_columns,
                        "confidence": min(0.99, 0.5 + len(period_columns) * 0.05),
                    }
                )
        return candidates

    def _infer_scenario_for_period_column(self, sheet: Any, row_idx: int, col_idx: int) -> tuple[str, str]:
        matches: list[tuple[int, int, str, str]] = []
        max_row = row_idx
        for row_number in range(max(1, row_idx - 2), max_row + 1):
            for col_number in range(max(1, col_idx - 3), min(sheet.max_column or col_idx, col_idx + 3) + 1):
                value = sheet.cell(row_number, col_number).value
                scenario = scenario_from_text(value)
                if row_number == row_idx and col_number == col_idx and period_from_value(value) and not scenario:
                    continue
                if not scenario:
                    continue
                if scenario in {"gb", "bm"} and not (row_number == row_idx and col_number == col_idx):
                    continue
                row_distance = abs(row_idx - row_number)
                col_distance = abs(col_idx - col_number)
                exact_column_priority = 0 if col_number == col_idx else 1
                matches.append(
                    (
                        row_distance + col_distance,
                        exact_column_priority,
                        scenario,
                        f"{sheet.cell(row_number, col_number).coordinate}={value}",
                    )
                )
        if not matches:
            return "fact", "default_fact"
        matches.sort(key=lambda item: (item[0], item[1], item[2]))
        return matches[0][2], matches[0][3]

    def _candidate_metric_rows(
        self,
        sheet: Any,
        header_row: int,
        metric_col: int,
        period_columns: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        current_section: str | None = None
        current_section_row: int | None = None
        current_section_level: int | None = None
        previous_metric_label: str | None = None
        previous_metric: dict[str, Any] | None = None
        hierarchy_stack: list[dict[str, Any]] = []
        unique_period_cols = self._first_columns_by_period(period_columns)
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            label_cell = sheet.cell(row_idx, metric_col)
            label = str(label_cell.value or "").strip()
            if not label:
                continue
            normalized_label = normalize_text(label)
            if looks_like_section_marker(label):
                if previous_metric:
                    current_section = previous_metric_label
                    current_section_row = int(previous_metric["row_number"])
                    current_section_level = int(previous_metric.get("layout_level") or 0)
                continue
            is_technical = self.domain_pack.is_technical_label(label)
            value_count = 0
            numeric_periods: list[str] = []
            for period, col in unique_period_cols.items():
                if _safe_float(sheet.cell(row_idx, col).value) is not None:
                    value_count += 1
                    numeric_periods.append(period)
            unit = self._infer_unit(sheet, row_idx, metric_col, period_columns)
            if looks_like_metric_label(label):
                layout = self._row_layout_evidence(sheet, row_idx, label_cell)
                if (
                    current_section
                    and current_section_row
                    and current_section_level is not None
                    and int(layout["layout_level"]) < current_section_level
                ):
                    current_section = None
                    current_section_row = None
                    current_section_level = None
                parent = self._parent_for_row(
                    hierarchy_stack=hierarchy_stack,
                    current_section=current_section,
                    current_section_row=current_section_row,
                    current_section_level=current_section_level,
                    layout_level=int(layout["layout_level"]),
                )
                section_path = parent["label"] if parent else ""
                rows.append(
                    {
                        "row_number": row_idx,
                        "label": label,
                        "normalized_label": normalized_label,
                        "unit": unit,
                        "section_path": section_path,
                        "parent_row_number": parent["row_number"] if parent else None,
                        "parent_label": parent["label"] if parent else "",
                        "hierarchy_evidence": parent["evidence"] if parent else "",
                        **layout,
                        "is_technical": is_technical,
                        "value_count": value_count,
                        "numeric_periods": numeric_periods,
                    }
                )
                metric_event = {
                    "row_number": row_idx,
                    "label": label,
                    "layout_level": int(layout["layout_level"]),
                    "is_bold": bool(layout["is_bold"]),
                    "value_count": value_count,
                }
                self._push_hierarchy_stack(hierarchy_stack, metric_event)
                if value_count:
                    previous_metric_label = label
                    previous_metric = metric_event
        return rows

    def _row_layout_evidence(self, sheet: Any, row_idx: int, label_cell: Any) -> dict[str, Any]:
        indent = int(getattr(label_cell.alignment, "indent", 0) or 0)
        outline_level = int(getattr(sheet.row_dimensions[row_idx], "outlineLevel", 0) or 0)
        is_bold = bool(getattr(label_cell.font, "bold", False))
        fill_type = str(getattr(label_cell.fill, "fill_type", "") or "")
        layout_level = outline_level if outline_level > 0 else indent
        return {
            "indent_level": indent,
            "outline_level": outline_level,
            "layout_level": layout_level,
            "is_bold": is_bold,
            "fill_type": fill_type,
        }

    def _parent_for_row(
        self,
        *,
        hierarchy_stack: list[dict[str, Any]],
        current_section: str | None,
        current_section_row: int | None,
        current_section_level: int | None,
        layout_level: int,
    ) -> dict[str, Any] | None:
        if current_section and current_section_row and current_section_level is not None:
            return {
                "row_number": current_section_row,
                "label": current_section,
                "evidence": "section_marker",
            }
        if layout_level > 0:
            for candidate in reversed(hierarchy_stack):
                if int(candidate.get("layout_level") or 0) < layout_level:
                    return {
                        "row_number": candidate["row_number"],
                        "label": candidate["label"],
                        "evidence": "layout_indent_or_outline",
                    }
        return None

    def _push_hierarchy_stack(self, hierarchy_stack: list[dict[str, Any]], metric_event: dict[str, Any]) -> None:
        current_level = int(metric_event.get("layout_level") or 0)
        while hierarchy_stack and int(hierarchy_stack[-1].get("layout_level") or 0) >= current_level:
            hierarchy_stack.pop()
        hierarchy_stack.append(metric_event)

    def _infer_unit(self, sheet: Any, row_idx: int, metric_col: int, period_columns: list[dict[str, Any]]) -> str:
        period_cols = {item["col"] for item in period_columns}
        for col in range(metric_col + 1, min(metric_col + 4, sheet.max_column) + 1):
            if col in period_cols:
                continue
            value = sheet.cell(row_idx, col).value
            if isinstance(value, str) and 0 < len(value.strip()) <= 24:
                return value.strip()
        return ""

    def _first_columns_by_period(self, period_columns: list[dict[str, Any]]) -> dict[str, int]:
        result: dict[str, int] = {}
        for item in sorted(period_columns, key=lambda value: value["col"]):
            result.setdefault(item["period"], item["col"])
        return result

    def _formula_cells(self, sheet: Any, metric_col: int | None) -> list[dict[str, Any]]:
        formulas: list[dict[str, Any]] = []
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                label = ""
                if metric_col:
                    label = str(sheet.cell(cell.row, metric_col).value or "").strip()
                formulas.append(
                    {
                        "row": cell.row,
                        "col": cell.column,
                        "coordinate": cell.coordinate,
                        "formula": value,
                        "signature": formula_signature(value),
                        "row_label": label,
                        "referenced_row_signs": self._formula_referenced_row_signs(value),
                    }
                )
                formulas[-1]["referenced_rows"] = sorted({item["row"] for item in formulas[-1]["referenced_row_signs"]})
                if len(formulas) >= 500:
                    return formulas
        return formulas

    def _formula_referenced_row_signs(self, formula: str) -> list[dict[str, Any]]:
        refs: dict[int, dict[str, Any]] = {}
        sum_spans: list[tuple[int, int]] = []
        for match in re.finditer(r"(?:SUM|СУММ)\(([^()]*)\)", formula, flags=re.IGNORECASE):
            sum_spans.append(match.span())
            function_sign = self._operator_sign_before(formula, match.start())
            sign = -1 if function_sign == "-" else 1
            arguments = match.group(1)
            for range_match in re.finditer(
                r"(?:'[^']+'!)?\$?[A-Z]{1,3}\$?(\d+)\s*:\s*(?:'[^']+'!)?\$?[A-Z]{1,3}\$?(\d+)",
                arguments,
                flags=re.IGNORECASE,
            ):
                start_row = int(range_match.group(1))
                end_row = int(range_match.group(2))
                for row_number in range(min(start_row, end_row), max(start_row, end_row) + 1):
                    self._add_formula_ref(refs, row_number, sign, "sum_range")
            for cell_match in re.finditer(r"(?:'[^']+'!)?\$?[A-Z]{1,3}\$?(\d+)", arguments, flags=re.IGNORECASE):
                before = arguments[cell_match.start() - 1] if cell_match.start() > 0 else ""
                after = arguments[cell_match.end()] if cell_match.end() < len(arguments) else ""
                if before == ":" or after == ":":
                    continue
                self._add_formula_ref(refs, int(cell_match.group(1)), sign, "sum_argument")

        for match in re.finditer(r"(?:'[^']+'!)?\$?[A-Z]{1,3}\$?(\d+)", formula, flags=re.IGNORECASE):
            if any(start <= match.start() < end for start, end in sum_spans):
                continue
            before = self._previous_non_space(formula, match.start())
            after = self._next_non_space(formula, match.end())
            if before == ":" or after == ":":
                continue
            if before in {"*", "/", "^"} or after in {"*", "/", "^"}:
                self._add_formula_ref(refs, int(match.group(1)), 0, "non_additive")
                continue
            if before == "-":
                self._add_formula_ref(refs, int(match.group(1)), -1, "direct_subtractive")
            elif before in {"", "=", "+", "(", ","} or after in {"", "+", "-", ")", ","}:
                self._add_formula_ref(refs, int(match.group(1)), 1, "direct_additive")
            else:
                self._add_formula_ref(refs, int(match.group(1)), 0, "non_additive")
        return sorted(refs.values(), key=lambda item: (item["row"], item["mode"]))

    def _add_formula_ref(self, refs: dict[int, dict[str, Any]], row_number: int, sign: int, mode: str) -> None:
        precedence = {"sum_range": 4, "sum_argument": 4, "direct_additive": 3, "direct_subtractive": 3, "non_additive": 1}
        current = refs.get(row_number)
        candidate = {"row": row_number, "sign": sign, "mode": mode}
        if current is None or precedence.get(mode, 0) > precedence.get(str(current.get("mode") or ""), 0):
            refs[row_number] = candidate

    def _operator_sign_before(self, formula: str, index: int) -> str:
        previous = self._previous_non_space(formula, index)
        return "-" if previous == "-" else "+"

    def _previous_non_space(self, formula: str, index: int) -> str:
        cursor = index - 1
        while cursor >= 0 and formula[cursor].isspace():
            cursor -= 1
        return formula[cursor] if cursor >= 0 else ""

    def _next_non_space(self, formula: str, index: int) -> str:
        cursor = index
        while cursor < len(formula) and formula[cursor].isspace():
            cursor += 1
        return formula[cursor] if cursor < len(formula) else ""

    def _comments(self, sheet: Any) -> list[dict[str, Any]]:
        comments: list[dict[str, Any]] = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append(
                        {
                            "row": cell.row,
                            "col": cell.column,
                            "coordinate": cell.coordinate,
                            "text": str(cell.comment.text or "")[:500],
                        }
                    )
        return comments[:200]


class MetricRegistry:
    def __init__(
        self,
        *,
        domain_pack: DomainPack | None = None,
        memory_priors: list[dict[str, Any]] | None = None,
        metric_mapping_priors: list[dict[str, Any]] | None = None,
        semantic_resolver: OllamaMetricSemanticResolver | None = None,
    ):
        self.domain_pack = domain_pack or DomainPack.load("generic")
        self.memory_aliases = self._aliases_from_memory(memory_priors or [])
        self.mapping_aliases = self._aliases_from_metric_mappings(metric_mapping_priors or [])
        self.semantic_resolver = semantic_resolver
        self.candidates: dict[str, dict[str, Any]] = {}
        self.semantic_suggestions: dict[str, dict[str, Any]] = {}
        self.warnings: list[str] = []

    def prime_semantic_suggestions(self, metric_inputs: list[dict[str, Any]]) -> None:
        if not self.semantic_resolver or not metric_inputs:
            return
        contexts = []
        for item in metric_inputs:
            raw_label = str(item.get("raw_label") or "")
            canonical_code, confidence, evidence, status = self._baseline_mapping(raw_label)
            contexts.append(
                {
                    "key": self._suggestion_key(
                        raw_label=raw_label,
                        source_sheet=str(item.get("source_sheet") or ""),
                        section_path=str(item.get("section_path") or ""),
                    ),
                    "raw_label": raw_label,
                    "unit": str(item.get("unit") or ""),
                    "source_sheet": str(item.get("source_sheet") or ""),
                    "section_path": str(item.get("section_path") or ""),
                    "baseline_code": canonical_code,
                    "baseline_confidence": confidence,
                    "baseline_evidence": evidence,
                    "baseline_status": status,
                    "baseline_semantic_type": self.domain_pack.semantic_type_for_label(raw_label, canonical_code),
                }
            )
        known_codes = sorted(
            set(self.domain_pack.aliases.values())
            .union(self.memory_aliases.values())
            .union(item["canonical_code"] for item in self.mapping_aliases.values())
        )
        suggestions, warnings = self.semantic_resolver.resolve_batch(contexts, known_metric_codes=known_codes)
        self.semantic_suggestions.update(suggestions)
        self.warnings.extend(warnings)

    def propose_metric(
        self,
        *,
        raw_label: str,
        unit: str,
        department: str,
        source_sheet: str,
        section_path: str,
        formula_signature_value: str = "",
    ) -> dict[str, Any]:
        canonical_code, confidence, evidence, status = self._baseline_mapping(raw_label)
        semantic_type = self.domain_pack.semantic_type_for_label(raw_label, canonical_code)
        aliases_extra = ""
        suggestion = self.semantic_suggestions.get(
            self._suggestion_key(raw_label=raw_label, source_sheet=source_sheet, section_path=section_path)
        )
        if suggestion:
            canonical_code, semantic_type, confidence, status, evidence, aliases_extra = self._merge_semantic_suggestion(
                raw_label=raw_label,
                baseline_code=canonical_code,
                baseline_confidence=confidence,
                baseline_evidence=evidence,
                baseline_status=status,
                baseline_semantic_type=semantic_type,
                suggestion=suggestion,
            )
        if semantic_type == "technical_check":
            status = "deprecated" if status == "proposed" else status
        metric_id = normalize_metric_code(f"{canonical_code}__{source_sheet}__{raw_label}")[:160]
        normalized = normalize_text(raw_label)
        aliases = list(
            dict.fromkeys(
                [
                    raw_label,
                    normalized,
                    canonical_code.replace("_", " "),
                    *[alias.strip() for alias in aliases_extra.split("|") if alias.strip()],
                ]
            )
        )
        candidate = {
            "metric_id": metric_id,
            "canonical_code": canonical_code,
            "raw_label": raw_label,
            "label": raw_label,
            "aliases": "|".join(alias for alias in aliases if alias),
            "unit": unit,
            "department": department,
            "source_sheet": source_sheet,
            "section_path": section_path,
            "semantic_type": semantic_type,
            "aggregation": aggregation_for_metric(canonical_code, unit),
            "status": status,
            "confidence": round(confidence, 3),
            "evidence": evidence if not formula_signature_value else f"{evidence}; formula={formula_signature_value}",
        }
        self.candidates.setdefault(metric_id, candidate)
        return candidate

    def list_candidates(self) -> list[dict[str, Any]]:
        return sorted(self.candidates.values(), key=lambda item: (item["canonical_code"], item["source_sheet"], item["raw_label"]))

    def _aliases_from_memory(self, memory_priors: list[dict[str, Any]]) -> dict[str, str]:
        aliases: dict[str, str] = {}
        for relation in memory_priors:
            for key in ("source_metric_code", "target_metric_code"):
                code = str(relation.get(key) or "").strip()
                if code:
                    aliases[normalize_text(code)] = code
                    aliases[normalize_text(code.replace("_", " "))] = code
        return aliases

    def _aliases_from_metric_mappings(self, metric_mapping_priors: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        aliases: dict[str, dict[str, Any]] = {}
        for mapping in metric_mapping_priors:
            canonical_code = str(mapping.get("canonical_code") or mapping.get("metric_code") or "").strip()
            if not canonical_code:
                continue
            alias_values = [
                str(mapping.get("raw_label") or "").strip(),
                str(mapping.get("label") or "").strip(),
                canonical_code,
                canonical_code.replace("_", " "),
            ]
            raw_aliases = mapping.get("aliases") or []
            if isinstance(raw_aliases, str):
                alias_values.extend(alias.strip() for alias in raw_aliases.split("|") if alias.strip())
            else:
                alias_values.extend(str(alias).strip() for alias in raw_aliases if str(alias).strip())
            mapping_record = {
                "canonical_code": canonical_code,
                "raw_label": str(mapping.get("raw_label") or mapping.get("label") or "").strip(),
                "label": str(mapping.get("label") or mapping.get("raw_label") or canonical_code).strip(),
                "aliases": [alias for alias in alias_values if alias],
                "semantic_type": str(mapping.get("semantic_type") or "unknown"),
            }
            for alias in alias_values:
                normalized = normalize_text(alias)
                if normalized:
                    aliases.setdefault(normalized, mapping_record)
        return aliases

    def _baseline_mapping(self, raw_label: str) -> tuple[str, float, str, str]:
        normalized = normalize_text(raw_label)
        if normalized in self.mapping_aliases:
            return self.mapping_aliases[normalized]["canonical_code"], 1.0, "metric_mapping_memory", "memory_applied"
        if normalized in self.memory_aliases:
            return self.memory_aliases[normalized], 0.9, "memory_alias", "mapped"
        canonical_code, confidence, evidence = self.domain_pack.canonical_code_for_label(raw_label)
        status = "mapped" if evidence in {"domain_alias", "partial_domain_alias"} else "proposed"
        return canonical_code, confidence, evidence, status

    def _merge_semantic_suggestion(
        self,
        *,
        raw_label: str,
        baseline_code: str,
        baseline_confidence: float,
        baseline_evidence: str,
        baseline_status: str,
        baseline_semantic_type: str,
        suggestion: dict[str, Any],
    ) -> tuple[str, str, float, str, str, str]:
        suggested_code = str(suggestion.get("canonical_code") or "").strip()
        suggested_semantic = str(suggestion.get("semantic_type") or baseline_semantic_type).strip()
        suggested_confidence = float(suggestion.get("confidence") or 0.0)
        known_codes = (
            set(self.domain_pack.aliases.values())
            .union(self.memory_aliases.values())
            .union(item["canonical_code"] for item in self.mapping_aliases.values())
        )
        keep_baseline_code = baseline_evidence in {"domain_alias", "memory_alias", "metric_mapping_memory"}
        if baseline_evidence == "partial_domain_alias" and suggested_confidence < 0.78:
            keep_baseline_code = True
        canonical_code = baseline_code if keep_baseline_code or suggested_confidence < 0.45 else suggested_code
        semantic_type = suggested_semantic if suggested_semantic in SEMANTIC_TYPES and suggested_confidence >= 0.35 else baseline_semantic_type
        confidence = max(baseline_confidence, min(0.9, suggested_confidence))
        status = baseline_status
        if not keep_baseline_code and canonical_code in known_codes and suggested_confidence >= 0.7:
            status = "mapped"
        elif not keep_baseline_code:
            status = "proposed"
        reason = str(suggestion.get("reason") or "").strip()
        evidence = (
            f"{baseline_evidence}; ollama_semantic:{suggestion.get('model')}; "
            f"llm_code={suggested_code}; llm_semantic={semantic_type}; llm_reason={reason}"
        )
        aliases = str(suggestion.get("aliases") or "")
        if not suggested_code:
            canonical_code = baseline_code
            confidence = baseline_confidence
            evidence = f"{baseline_evidence}; ollama_semantic_empty"
        if self.domain_pack.is_technical_label(raw_label):
            semantic_type = "technical_check"
        return canonical_code, semantic_type, round(confidence, 3), status, evidence, aliases

    def _suggestion_key(self, *, raw_label: str, source_sheet: str, section_path: str) -> str:
        return normalize_metric_code(f"{source_sheet}__{section_path}__{raw_label}")[:180]


class FormulaRelationEngine:
    evidence_type = "formula"

    def discover(self, profile: dict[str, Any], row_metric_map: dict[tuple[str, int], str]) -> list[dict[str, Any]]:
        relations: dict[tuple[str, str], dict[str, Any]] = {}
        row_profiles = self._row_profiles(profile)
        value_columns_by_sheet = self._value_columns_by_sheet(profile)
        for sheet in profile.get("sheets", []):
            sheet_name = sheet["sheet_name"]
            value_columns = value_columns_by_sheet.get(sheet_name) or set()
            for formula in sheet.get("formula_cells", []):
                if value_columns and int(formula.get("col") or 0) not in value_columns:
                    continue
                target_row = int(formula["row"])
                target = row_metric_map.get((sheet_name, target_row))
                if not target:
                    continue
                references = formula.get("referenced_row_signs") or [
                    {"row": row_number, "sign": 0, "mode": "legacy"} for row_number in formula.get("referenced_rows", [])
                ]
                for reference in references:
                    source_row = int(reference.get("row") or 0)
                    source = row_metric_map.get((sheet_name, source_row))
                    if not source or source == target:
                        continue
                    reference_sign = int(reference.get("sign") or 0)
                    reference_mode = str(reference.get("mode") or "legacy")
                    rollup_support = self._has_rollup_support(
                        source_row=row_profiles.get((sheet_name, source_row), {}),
                        target_row=row_profiles.get((sheet_name, target_row), {}),
                        formula_text=str(formula.get("formula") or ""),
                        reference_mode=reference_mode,
                        reference_sign=reference_sign,
                    )
                    if rollup_support and reference_sign < 0:
                        relation_type = "inverse_driver"
                        edge_type = "inverse_driver"
                        confidence = 0.82
                    elif rollup_support:
                        relation_type = "formula_input"
                        edge_type = "component"
                        confidence = 0.86
                    else:
                        relation_type = "formula_reference"
                        edge_type = "driver"
                        confidence = 0.52
                    needs_approval = not rollup_support
                    needs_approval_reason = "" if rollup_support else "formula reference is not a clear child-to-parent rollup"
                    key = (source, target)
                    candidate = {
                        "source_metric_code": source,
                        "target_metric_code": target,
                        "relation_type": relation_type,
                        "edge_type": edge_type,
                        "confidence": confidence,
                        "evidence_type": self.evidence_type,
                        "evidence": (
                            f"{sheet_name}!{formula['coordinate']} uses row {source_row} "
                            f"({reference_mode}, sign={reference_sign})."
                        ),
                        "period_window": "",
                        "source_document_id": None,
                        "needs_approval": needs_approval,
                        "needs_approval_reason": needs_approval_reason,
                    }
                    current = relations.get(key)
                    if current is None or self._candidate_rank(candidate) > self._candidate_rank(current):
                        relations[key] = candidate
        return list(relations.values())[:300]

    def _value_columns_by_sheet(self, profile: dict[str, Any]) -> dict[str, set[int]]:
        columns: dict[str, set[int]] = {}
        for sheet in profile.get("sheets", []):
            header_candidates = sheet.get("header_candidates") or []
            if not header_candidates:
                continue
            columns[str(sheet["sheet_name"])] = {
                int(item["col"])
                for item in header_candidates[0].get("period_columns", [])
                if item.get("col") is not None
            }
        return columns

    def _row_profiles(self, profile: dict[str, Any]) -> dict[tuple[str, int], dict[str, Any]]:
        rows: dict[tuple[str, int], dict[str, Any]] = {}
        for sheet in profile.get("sheets", []):
            sheet_name = sheet["sheet_name"]
            for row in sheet.get("candidate_metric_rows", []):
                rows[(sheet_name, int(row["row_number"]))] = row
        return rows

    def _has_rollup_support(
        self,
        *,
        source_row: dict[str, Any],
        target_row: dict[str, Any],
        formula_text: str,
        reference_mode: str,
        reference_sign: int,
    ) -> bool:
        if not source_row or not target_row:
            return False
        if reference_sign == 0:
            return False
        source_number = int(source_row.get("row_number") or 0)
        target_number = int(target_row.get("row_number") or 0)
        source_level = int(source_row.get("layout_level") or 0)
        target_level = int(target_row.get("layout_level") or 0)
        if int(source_row.get("parent_row_number") or 0) == target_number:
            return True
        if not self._is_additive_rollup_formula(formula_text, reference_mode):
            return False
        if source_number > target_number and source_level >= target_level:
            return True
        return source_number > target_number and source_level > target_level

    def _is_additive_rollup_formula(self, formula_text: str, reference_mode: str) -> bool:
        if reference_mode in {"direct_additive", "direct_subtractive", "sum_argument", "sum_range"}:
            return True
        normalized_formula = formula_text.upper().replace(" ", "")
        return "SUM(" in normalized_formula or "СУММ(" in normalized_formula

    def _candidate_rank(self, relation: dict[str, Any]) -> tuple[int, float]:
        approval_rank = 1 if not relation.get("needs_approval") else 0
        return approval_rank, float(relation.get("confidence") or 0.0)


class StructureRelationEngine:
    evidence_type = "row_structure"

    def __init__(self, domain_pack: DomainPack | None = None):
        self.domain_pack = domain_pack or DomainPack.load("generic")

    def discover(self, profile: dict[str, Any], row_metric_map: dict[tuple[str, int], str]) -> list[dict[str, Any]]:
        relations: dict[tuple[str, str], dict[str, Any]] = {}
        rows_by_sheet_number = self._rows_by_sheet_number(profile)
        for sheet in profile.get("sheets", []):
            sheet_name = sheet["sheet_name"]
            for row in sheet.get("candidate_metric_rows", []):
                source = row_metric_map.get((sheet_name, row["row_number"]))
                parent_row_number = row.get("parent_row_number")
                target = row_metric_map.get((sheet_name, int(parent_row_number))) if parent_row_number else None
                if not source or not target or source == target:
                    continue
                parent_row = rows_by_sheet_number.get((sheet_name, int(parent_row_number))) if parent_row_number else {}
                source_semantic = self.domain_pack.semantic_type_for_label(str(row.get("label") or ""), source)
                target_semantic = self.domain_pack.semantic_type_for_label(str(parent_row.get("label") or ""), target)
                hierarchy_evidence = str(row.get("hierarchy_evidence") or "section_path")
                auto_approve = self._is_strong_component_hierarchy(
                    source_semantic=source_semantic,
                    target_semantic=target_semantic,
                    hierarchy_evidence=hierarchy_evidence,
                )
                confidence = 0.7 if hierarchy_evidence == "layout_indent_or_outline" else 0.66 if auto_approve else 0.58
                key = (source, target)
                relations.setdefault(
                    key,
                    {
                        "source_metric_code": source,
                        "target_metric_code": target,
                        "relation_type": "component",
                        "edge_type": "component",
                        "confidence": confidence,
                        "evidence_type": self.evidence_type,
                        "evidence": (
                            f"Row '{row['label']}' is inside section '{row.get('parent_label')}'. "
                            f"Evidence: {hierarchy_evidence}; source_semantic={source_semantic}; "
                            f"target_semantic={target_semantic}."
                        ),
                        "period_window": "",
                        "source_document_id": None,
                        "needs_approval": not auto_approve,
                        "needs_approval_reason": "" if auto_approve else "row hierarchy is structural evidence only",
                    },
                )
        return list(relations.values())[:300]

    def _rows_by_sheet_number(self, profile: dict[str, Any]) -> dict[tuple[str, int], dict[str, Any]]:
        rows: dict[tuple[str, int], dict[str, Any]] = {}
        for sheet in profile.get("sheets", []):
            sheet_name = sheet["sheet_name"]
            for row in sheet.get("candidate_metric_rows", []):
                rows[(sheet_name, int(row["row_number"]))] = row
        return rows

    def _is_strong_component_hierarchy(
        self,
        *,
        source_semantic: str,
        target_semantic: str,
        hierarchy_evidence: str,
    ) -> bool:
        if hierarchy_evidence not in {"section_marker", "layout_indent_or_outline"}:
            return False
        if source_semantic in {"technical_check", "denominator", "business_driver", "external_context", "unknown"}:
            return False
        return target_semantic in {"target_metric", "cost_component", "unknown"}


class DomainRuleEngine:
    evidence_type = "domain_rule"

    def __init__(self, domain_pack: DomainPack):
        self.domain_pack = domain_pack

    def discover(self, available_metric_codes: set[str]) -> list[dict[str, Any]]:
        relations: list[dict[str, Any]] = []
        for template in self.domain_pack.relation_templates:
            source = str(template.get("source_metric_code") or template.get("source") or "")
            target = str(template.get("target_metric_code") or template.get("target") or "")
            if source not in available_metric_codes or target not in available_metric_codes:
                continue
            relation_type = str(template.get("relation_type") or template.get("edge_type") or "driver")
            relations.append(
                {
                    "source_metric_code": source,
                    "target_metric_code": target,
                    "relation_type": relation_type,
                    "edge_type": edge_type_for_relation(relation_type, source),
                    "confidence": float(template.get("confidence") or template.get("strength") or 0.82),
                    "evidence_type": self.evidence_type,
                    "evidence": str(template.get("reason") or "Domain pack relation template."),
                    "period_window": "",
                    "source_document_id": None,
                    "needs_approval": False,
                    "needs_approval_reason": "",
                }
            )
        return relations


class MemoryPriorEngine:
    evidence_type = "memory_prior"

    def discover(self, memory_priors: list[dict[str, Any]], available_metric_codes: set[str]) -> list[dict[str, Any]]:
        relations: list[dict[str, Any]] = []
        for prior in memory_priors:
            source = str(prior.get("source_metric_code") or "")
            target = str(prior.get("target_metric_code") or "")
            if not source or not target:
                continue
            if source not in available_metric_codes and target not in available_metric_codes:
                continue
            relation_type = str(prior.get("relation_type") or prior.get("edge_type") or "driver")
            relations.append(
                {
                    "source_metric_code": source,
                    "target_metric_code": target,
                    "relation_type": relation_type,
                    "edge_type": edge_type_for_relation(relation_type, source),
                    "confidence": float(prior.get("strength") or 1.0),
                    "evidence_type": self.evidence_type,
                    "evidence": str(prior.get("reason") or prior.get("note") or "Confirmed memory prior."),
                    "period_window": "",
                    "source_document_id": prior.get("source_document_id"),
                    "needs_approval": False,
                    "needs_approval_reason": "",
                }
            )
        return relations


class TextRelationEngine:
    evidence_type = "text"

    def discover(self, candidate_relations: list[dict[str, Any]], available_metric_codes: set[str]) -> list[dict[str, Any]]:
        relations: list[dict[str, Any]] = []
        for candidate in candidate_relations:
            source = str(candidate.get("source_metric_code") or candidate.get("source") or "")
            target = str(candidate.get("target_metric_code") or candidate.get("target") or "")
            if not source or not target:
                continue
            if source not in available_metric_codes or target not in available_metric_codes:
                continue
            relation_type = str(candidate.get("relation_type") or candidate.get("edge_type") or "driver")
            relations.append(
                {
                    "source_metric_code": source,
                    "target_metric_code": target,
                    "relation_type": relation_type,
                    "edge_type": edge_type_for_relation(relation_type, source),
                    "confidence": float(candidate.get("confidence") or 0.55),
                    "evidence_type": self.evidence_type,
                    "evidence": str(candidate.get("evidence") or candidate.get("note") or ""),
                    "period_window": "",
                    "source_document_id": candidate.get("source_document_id"),
                    "needs_approval": True,
                    "needs_approval_reason": "text evidence needs explicit confirmation",
                }
            )
        return relations


class StatisticalRelationEngine:
    evidence_type = "statistical"

    def __init__(self, *, min_periods: int = 12, strong_min_periods: int = 24):
        self.min_periods = min_periods
        self.strong_min_periods = strong_min_periods

    def discover(self, metric_series: dict[str, dict[str, float]]) -> list[dict[str, Any]]:
        metric_codes = sorted(metric_series)
        relations: list[dict[str, Any]] = []
        for source in metric_codes:
            for target in metric_codes:
                if source == target:
                    continue
                common_periods = sorted(set(metric_series[source]).intersection(metric_series[target]))
                if len(common_periods) < self.min_periods:
                    continue
                values_x = [metric_series[source][period] for period in common_periods]
                values_y = [metric_series[target][period] for period in common_periods]
                correlation = self._pearson(values_x, values_y)
                if correlation is None or abs(correlation) < 0.86:
                    continue
                relation_type = "driver" if correlation > 0 else "inverse_driver"
                confidence = min(0.72, 0.35 + abs(correlation) * 0.25 + min(len(common_periods), 24) / 100)
                relations.append(
                    {
                        "source_metric_code": source,
                        "target_metric_code": target,
                        "relation_type": relation_type,
                        "edge_type": edge_type_for_relation(relation_type, source),
                        "confidence": round(confidence, 3),
                        "evidence_type": self.evidence_type,
                        "evidence": f"Pearson correlation {correlation:.3f} over {len(common_periods)} periods.",
                        "period_window": f"{common_periods[0]}..{common_periods[-1]}",
                        "source_document_id": None,
                        "needs_approval": True,
                        "needs_approval_reason": "statistical relation needs business confirmation",
                    }
                )
        relations.sort(key=lambda item: (-float(item["confidence"]), item["source_metric_code"], item["target_metric_code"]))
        return relations[:100]

    def _pearson(self, values_x: list[float], values_y: list[float]) -> float | None:
        if len(values_x) != len(values_y) or len(values_x) < self.min_periods:
            return None
        mean_x = sum(values_x) / len(values_x)
        mean_y = sum(values_y) / len(values_y)
        centered_x = [value - mean_x for value in values_x]
        centered_y = [value - mean_y for value in values_y]
        denominator = math.sqrt(sum(value * value for value in centered_x) * sum(value * value for value in centered_y))
        if denominator == 0:
            return None
        return sum(left * right for left, right in zip(centered_x, centered_y, strict=True)) / denominator


class RelationScorer:
    EVIDENCE_WEIGHTS = {
        "domain_rule": 0.26,
        "formula": 0.22,
        "memory_prior": 0.22,
        "statistical": 0.12,
        "row_structure": 0.06,
        "text": 0.08,
    }

    def score(self, relation: dict[str, Any]) -> dict[str, Any]:
        confidence = float(relation.get("confidence") or 0.0)
        evidence_weight = self.EVIDENCE_WEIGHTS.get(str(relation.get("evidence_type") or ""), 0.04)
        score = min(1.0, confidence * 0.74 + evidence_weight)
        relation["score"] = round(score, 3)
        if relation.get("evidence_type") == "row_structure":
            if relation.get("needs_approval"):
                relation["needs_approval_reason"] = relation.get("needs_approval_reason") or "structural evidence is not causal"
        return relation


class ObservationScorer:
    def rank(self, observations: list[dict[str, Any]], metric_candidates: list[dict[str, Any]], dependencies: list[dict[str, Any]]) -> list[dict[str, Any]]:
        semantic_by_code = {item["canonical_code"]: item.get("semantic_type") for item in metric_candidates}
        inbound = Counter(item["target_metric_code"] for item in dependencies)
        ranked = []
        for observation in observations:
            metric_code = observation.get("metric_code")
            semantic_type = semantic_by_code.get(metric_code, "unknown")
            if semantic_type == "technical_check":
                continue
            base_score = float(observation.get("score") or 0.0)
            semantic_boost = 0.5 if semantic_type == "target_metric" else 0.25 if semantic_type == "cost_component" else 0.0
            graph_boost = min(0.35, inbound[metric_code] * 0.03)
            previous_value = abs(float(observation.get("previous_value") or 0.0))
            zero_penalty = 0.45 if previous_value < 1e-6 and abs(float(observation.get("delta_abs") or 0.0)) < 1.0 else 0.0
            observation = dict(observation)
            observation["noise_control_score"] = round(max(0.0, base_score + semantic_boost + graph_boost - zero_penalty), 4)
            ranked.append(observation)
        ranked.sort(key=lambda item: (-float(item.get("noise_control_score") or 0.0), item.get("metric_code") or ""))
        return ranked


class NormalizationAdapter:
    def __init__(
        self,
        *,
        domain_pack: DomainPack | None = None,
        memory_priors: list[dict[str, Any]] | None = None,
        metric_mapping_priors: list[dict[str, Any]] | None = None,
        metric_semantic_resolver: OllamaMetricSemanticResolver | None = None,
        relation_semantic_judge: Any | None = None,
    ):
        self.domain_pack = domain_pack or DomainPack.load("generic")
        self.memory_priors = list(memory_priors or [])
        self.relation_semantic_judge = relation_semantic_judge
        self.registry = MetricRegistry(
            domain_pack=self.domain_pack,
            memory_priors=self.memory_priors,
            metric_mapping_priors=metric_mapping_priors,
            semantic_resolver=metric_semantic_resolver,
        )
        self.profiler = WorkbookProfiler(domain_pack=self.domain_pack)
        self.relation_scorer = RelationScorer()

    def normalize_workbook(
        self,
        workbook_path: str | Path,
        output_dir: str | Path,
        *,
        department: str = "generic",
        source_workbook: str | None = None,
        source_document_id: str | None = None,
    ) -> DynamicRelationMemoryArtifacts:
        output_root = Path(output_dir)
        output_root.mkdir(parents=True, exist_ok=True)
        workbook_path = Path(workbook_path)
        source_workbook = source_workbook or workbook_path.name
        profile = self.profiler.profile_workbook(workbook_path)
        long_rows, pivot_rows, row_metric_map, warnings = self._normalize_rows(
            workbook_path=workbook_path,
            profile=profile,
            department=department,
            source_workbook=source_workbook,
        )
        metric_candidates = self.registry.list_candidates()
        relation_candidates = self._discover_relations(profile, row_metric_map, pivot_rows, source_document_id)
        relation_candidates = [self.relation_scorer.score(item) for item in relation_candidates]
        relation_candidates = self._dedupe_relations(relation_candidates)
        relation_candidates = self._apply_relation_quality_gate(relation_candidates, metric_candidates)
        relation_candidates, relation_gate_warnings = self._apply_relation_semantic_judge(
            relation_candidates,
            metric_candidates,
        )
        warnings.extend(relation_gate_warnings)
        relation_candidates = [self.relation_scorer.score(item) for item in relation_candidates]
        relation_candidates = self._dedupe_relations(relation_candidates)

        workbook_profile_path = output_root / "workbook_profile.json"
        normalized_facts_path = output_root / "normalized_facts.xlsx"
        pivoted_facts_path = output_root / "pivoted_facts.xlsx"
        metric_candidates_path = output_root / "metric_candidates.xlsx"
        relation_candidates_path = output_root / "relation_candidates.xlsx"
        dependency_rules_path = output_root / "dependency_rules.xlsx"
        quality_report_path = output_root / "normalization_quality_report.json"
        generated_manifest_path = output_root / "generated_manifest.yaml"

        workbook_profile_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
        self._write_xlsx(normalized_facts_path, "normalized_facts", LONG_FACT_HEADERS, long_rows)
        pivot_headers = [*PIVOT_DIMENSIONS, *sorted({row["metric_code"] for row in long_rows})]
        self._write_xlsx(pivoted_facts_path, "pivoted_facts", pivot_headers, pivot_rows)
        self._write_xlsx(metric_candidates_path, "metric_candidates", METRIC_CANDIDATE_HEADERS, metric_candidates)
        self._write_xlsx(relation_candidates_path, "relation_candidates", RELATION_CANDIDATE_HEADERS, relation_candidates)
        dependency_rows = self._dependency_rule_rows(relation_candidates)
        self._write_xlsx(
            dependency_rules_path,
            "dependency_rules",
            ["source_metric_code", "target_metric_code", "edge_type", "reason", "strength"],
            dependency_rows,
        )
        quality_report = self._quality_report(
            profile=profile,
            long_rows=long_rows,
            pivot_rows=pivot_rows,
            metric_candidates=metric_candidates,
            relation_candidates=relation_candidates,
            warnings=warnings,
        )
        quality_report_path.write_text(json.dumps(quality_report, ensure_ascii=False, indent=2), encoding="utf-8")
        manifest = self._manifest_for_outputs(pivoted_facts_path, metric_candidates_path, dependency_rules_path, pivot_headers)
        generated_manifest_path.write_text(yaml.safe_dump(manifest, sort_keys=False, allow_unicode=True), encoding="utf-8")

        return DynamicRelationMemoryArtifacts(
            workbook_profile_path=workbook_profile_path,
            normalized_facts_path=normalized_facts_path,
            pivoted_facts_path=pivoted_facts_path,
            metric_candidates_path=metric_candidates_path,
            relation_candidates_path=relation_candidates_path,
            dependency_rules_path=dependency_rules_path,
            quality_report_path=quality_report_path,
            generated_manifest_path=generated_manifest_path,
            profile=profile,
            metric_candidates=metric_candidates,
            relation_candidates=relation_candidates,
            warnings=[*profile.get("warnings", []), *warnings],
        )

    def _normalize_rows(
        self,
        *,
        workbook_path: Path,
        profile: dict[str, Any],
        department: str,
        source_workbook: str,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[tuple[str, int], str], list[str]]:
        workbook = load_workbook(workbook_path, data_only=True, read_only=False)
        long_rows: list[dict[str, Any]] = []
        pivot: dict[tuple[str, str, str, str], dict[str, Any]] = {}
        row_metric_map: dict[tuple[str, int], str] = {}
        warnings: list[str] = []
        try:
            sheets_by_name = {sheet.title: sheet for sheet in workbook.worksheets}
            self.registry.prime_semantic_suggestions(
                self._metric_inputs_for_semantic_resolution(
                    profile=profile,
                    department=department,
                )
            )
            warnings.extend(self.registry.warnings)
            for sheet_profile in profile.get("sheets", []):
                if not sheet_profile.get("header_candidates"):
                    continue
                sheet_name = sheet_profile["sheet_name"]
                sheet = sheets_by_name[sheet_name]
                header = sheet_profile["header_candidates"][0]
                period_cols = self._period_value_columns(header["period_columns"], sheet_name, warnings)
                for metric_row in sheet_profile.get("candidate_metric_rows", []):
                    if metric_row.get("is_technical") or not metric_row.get("value_count"):
                        continue
                    candidate = self.registry.propose_metric(
                        raw_label=metric_row["label"],
                        unit=metric_row.get("unit") or "",
                        department=department,
                        source_sheet=sheet_name,
                        section_path=metric_row.get("section_path") or "",
                    )
                    if candidate["semantic_type"] == "technical_check":
                        continue
                    metric_code = candidate["canonical_code"]
                    row_metric_map[(sheet_name, int(metric_row["row_number"]))] = metric_code
                    for period_col in period_cols:
                        period = str(period_col["period"])
                        scenario = str(period_col["scenario"])
                        value = _safe_float(sheet.cell(int(metric_row["row_number"]), int(period_col["col"])).value)
                        if value is None:
                            continue
                        long_row = {
                            "period": period,
                            "scenario": scenario,
                            "department": department,
                            "source_workbook": source_workbook,
                            "sheet": sheet_name,
                            "section_path": metric_row.get("section_path") or "",
                            "metric_code": metric_code,
                            "raw_label": metric_row["label"],
                            "unit": metric_row.get("unit") or "",
                            "value": value,
                        }
                        long_rows.append(long_row)
                        key = (period, sheet_name, source_workbook, scenario)
                        pivot_row = pivot.setdefault(
                            key,
                            {
                                "month": period,
                                "report_sheet": sheet_name,
                                "source_workbook": source_workbook,
                                "scenario": scenario,
                            },
                        )
                        if metric_code in pivot_row and isinstance(pivot_row[metric_code], (int, float)):
                            pivot_row[metric_code] = float(pivot_row[metric_code]) + value
                        else:
                            pivot_row[metric_code] = value
        finally:
            workbook.close()
        pivot_rows = sorted(pivot.values(), key=lambda item: (item["month"], item["report_sheet"]))
        return long_rows, pivot_rows, row_metric_map, warnings

    def _metric_inputs_for_semantic_resolution(
        self,
        *,
        profile: dict[str, Any],
        department: str,
    ) -> list[dict[str, Any]]:
        inputs: list[dict[str, Any]] = []
        seen: set[tuple[str, str, str]] = set()
        for sheet_profile in profile.get("sheets", []):
            sheet_name = str(sheet_profile.get("sheet_name") or "")
            for metric_row in sheet_profile.get("candidate_metric_rows", []):
                if metric_row.get("is_technical") or not metric_row.get("value_count"):
                    continue
                raw_label = str(metric_row.get("label") or "")
                section_path = str(metric_row.get("section_path") or "")
                key = (sheet_name, section_path, raw_label)
                if key in seen:
                    continue
                seen.add(key)
                inputs.append(
                    {
                        "raw_label": raw_label,
                        "unit": str(metric_row.get("unit") or ""),
                        "department": department,
                        "source_sheet": sheet_name,
                        "section_path": section_path,
                    }
                )
        return inputs

    def _period_value_columns(
        self,
        period_columns: list[dict[str, Any]],
        sheet_name: str,
        warnings: list[str],
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        seen_periods: Counter[str] = Counter()
        seen_period_scenarios: Counter[tuple[str, str]] = Counter()
        for item in sorted(period_columns, key=lambda value: value["col"]):
            period = str(item["period"])
            scenario = str(item.get("scenario") or "fact")
            seen_periods[period] += 1
            seen_period_scenarios[(period, scenario)] += 1
            if seen_period_scenarios[(period, scenario)] > 1:
                scenario = f"{scenario}_{seen_period_scenarios[(period, scenario)]}"
            result.append({"period": period, "scenario": scenario, "col": int(item["col"])})
        for period, count in seen_periods.items():
            if count > 1:
                scenarios = ", ".join(
                    f"{item['period']}:{item['scenario']}@col{item['col']}" for item in result if item["period"] == period
                )
                warnings.append(f"Sheet {sheet_name}: duplicate period {period}; preserved scenario columns: {scenarios}.")
        return result

    def _discover_relations(
        self,
        profile: dict[str, Any],
        row_metric_map: dict[tuple[str, int], str],
        pivot_rows: list[dict[str, Any]],
        source_document_id: str | None,
    ) -> list[dict[str, Any]]:
        available_metric_codes = {metric_code for metric_code in row_metric_map.values()}
        relations = [
            *FormulaRelationEngine().discover(profile, row_metric_map),
            *StructureRelationEngine(self.domain_pack).discover(profile, row_metric_map),
            *DomainRuleEngine(self.domain_pack).discover(available_metric_codes),
            *MemoryPriorEngine().discover(self.memory_priors, available_metric_codes),
            *StatisticalRelationEngine().discover(self._metric_series_from_pivot(pivot_rows)),
        ]
        for relation in relations:
            relation["source_document_id"] = relation.get("source_document_id") or source_document_id
        return relations

    def _metric_series_from_pivot(self, pivot_rows: list[dict[str, Any]]) -> dict[str, dict[str, float]]:
        values_by_metric_month: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
        for row in pivot_rows:
            month = str(row.get("month") or "")
            if not month:
                continue
            for key, value in row.items():
                if key in PIVOT_DIMENSIONS:
                    continue
                numeric = _safe_float(value)
                if numeric is not None:
                    values_by_metric_month[key][month].append(numeric)
        series: dict[str, dict[str, float]] = {}
        for metric_code, month_map in values_by_metric_month.items():
            aggregation = aggregation_for_metric(metric_code)
            series[metric_code] = {}
            for month, values in month_map.items():
                series[metric_code][month] = sum(values) / len(values) if aggregation == "mean" else sum(values)
        return series

    def _dedupe_relations(self, relations: list[dict[str, Any]]) -> list[dict[str, Any]]:
        precedence = {"domain_rule": 4, "formula": 3, "memory_prior": 3, "statistical": 2, "row_structure": 1}
        by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
        for relation in relations:
            if relation.get("rejected_by_relation_gate"):
                continue
            key = (
                relation["source_metric_code"],
                relation["target_metric_code"],
                relation.get("relation_type") or relation.get("edge_type") or "driver",
            )
            current = by_key.get(key)
            if current is None:
                by_key[key] = relation
                continue
            current_rank = (precedence.get(str(current.get("evidence_type")), 0), float(current.get("score") or 0.0))
            new_rank = (precedence.get(str(relation.get("evidence_type")), 0), float(relation.get("score") or 0.0))
            if new_rank > current_rank:
                by_key[key] = relation
        return sorted(
            by_key.values(),
            key=lambda item: (-float(item.get("score") or 0.0), item["source_metric_code"], item["target_metric_code"]),
        )

    def _apply_relation_quality_gate(
        self,
        relations: list[dict[str, Any]],
        metric_candidates: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        semantic_by_code = self._metric_semantics(metric_candidates)
        gated = [self._gate_single_relation(dict(relation), semantic_by_code) for relation in relations]
        return self._guard_reciprocal_component_cycles(gated, semantic_by_code)

    def _metric_semantics(self, metric_candidates: list[dict[str, Any]]) -> dict[str, str]:
        priority = {
            "technical_check": 6,
            "target_metric": 5,
            "cost_component": 4,
            "business_driver": 3,
            "denominator": 3,
            "external_context": 2,
            "unknown": 1,
        }
        result: dict[str, str] = {}
        for item in metric_candidates:
            code = str(item.get("canonical_code") or item.get("metric_code") or "")
            semantic = str(item.get("semantic_type") or "unknown")
            if not code:
                continue
            current = result.get(code, "unknown")
            if priority.get(semantic, 0) >= priority.get(current, 0):
                result[code] = semantic
        return result

    def _gate_single_relation(self, relation: dict[str, Any], semantic_by_code: dict[str, str]) -> dict[str, Any]:
        evidence_type = str(relation.get("evidence_type") or "")
        if evidence_type in {"domain_rule", "memory_prior"}:
            return relation
        source = str(relation.get("source_metric_code") or "")
        target = str(relation.get("target_metric_code") or "")
        source_semantic = semantic_by_code.get(source, "unknown")
        target_semantic = semantic_by_code.get(target, "unknown")
        relation["source_semantic_type"] = source_semantic
        relation["target_semantic_type"] = target_semantic
        if relation.get("edge_type") != "component":
            return relation
        reason = ""
        if target_semantic in {"technical_check", "denominator", "business_driver", "external_context"}:
            reason = f"component target semantic_type={target_semantic} is not a rollup metric"
        elif source_semantic == "target_metric" and target_semantic in {"cost_component", "denominator", "business_driver"}:
            reason = "component direction contradicts metric semantics"
        elif source_semantic in {"denominator", "business_driver", "external_context"} and target_semantic == "cost_component":
            reason = "driver or denominator evidence should not be auto-approved as a cost component"
        if reason:
            relation["needs_approval"] = True
            relation["needs_approval_reason"] = _append_reason(relation.get("needs_approval_reason"), reason)
            relation["confidence"] = min(float(relation.get("confidence") or 0.0), 0.52)
        return relation

    def _guard_reciprocal_component_cycles(
        self,
        relations: list[dict[str, Any]],
        semantic_by_code: dict[str, str],
    ) -> list[dict[str, Any]]:
        approved_components: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
        for relation in relations:
            if relation.get("needs_approval") or relation.get("edge_type") != "component":
                continue
            source = str(relation.get("source_metric_code") or "")
            target = str(relation.get("target_metric_code") or "")
            if not source or not target or source == target:
                continue
            approved_components[tuple(sorted((source, target)))].append(relation)
        for reciprocal_group in approved_components.values():
            directions = {(item.get("source_metric_code"), item.get("target_metric_code")) for item in reciprocal_group}
            if len(directions) < 2:
                continue
            keep = max(reciprocal_group, key=lambda item: self._relation_direction_rank(item, semantic_by_code))
            for relation in reciprocal_group:
                if relation is keep:
                    continue
                relation["needs_approval"] = True
                relation["needs_approval_reason"] = _append_reason(
                    relation.get("needs_approval_reason"),
                    "reciprocal component cycle guard",
                )
                relation["confidence"] = min(float(relation.get("confidence") or 0.0), 0.52)
        return relations

    def _relation_direction_rank(self, relation: dict[str, Any], semantic_by_code: dict[str, str]) -> tuple[int, int, float]:
        evidence_priority = {"domain_rule": 5, "memory_prior": 4, "formula": 3, "row_structure": 2, "statistical": 1}
        target_priority = {"target_metric": 5, "cost_component": 4, "unknown": 3, "business_driver": 2, "denominator": 2}
        source = str(relation.get("source_metric_code") or "")
        target = str(relation.get("target_metric_code") or "")
        source_semantic = semantic_by_code.get(source, "unknown")
        target_semantic = semantic_by_code.get(target, "unknown")
        source_penalty = -2 if source_semantic == "target_metric" and target_semantic != "target_metric" else 0
        return (
            evidence_priority.get(str(relation.get("evidence_type") or ""), 0),
            target_priority.get(target_semantic, 0) + source_penalty,
            float(relation.get("score") or relation.get("confidence") or 0.0),
        )

    def _apply_relation_semantic_judge(
        self,
        relations: list[dict[str, Any]],
        metric_candidates: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], list[str]]:
        if not self.relation_semantic_judge:
            return relations, []
        contexts = self._relation_gate_contexts(relations, metric_candidates)
        if not contexts:
            return relations, []
        decisions, warnings = self.relation_semantic_judge.judge_batch(contexts)
        relation_by_key = {str(relation.get("_relation_gate_key")): relation for relation in relations}
        for key, decision in decisions.items():
            relation = relation_by_key.get(key)
            if not relation:
                continue
            self._apply_relation_gate_decision(relation, decision)
        return [item for item in relations if not item.get("rejected_by_relation_gate")], warnings

    def _relation_gate_contexts(
        self,
        relations: list[dict[str, Any]],
        metric_candidates: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        metric_info = self._metric_info_by_code(metric_candidates)
        contexts: list[dict[str, Any]] = []
        for index, relation in enumerate(relations):
            if relation.get("evidence_type") not in {"formula", "row_structure", "statistical", "text"}:
                continue
            key = f"rel_{index}"
            relation["_relation_gate_key"] = key
            source = str(relation.get("source_metric_code") or "")
            target = str(relation.get("target_metric_code") or "")
            contexts.append(
                {
                    "key": key,
                    "source_metric_code": source,
                    "source_label": metric_info.get(source, {}).get("label", source),
                    "source_semantic_type": metric_info.get(source, {}).get("semantic_type", "unknown"),
                    "target_metric_code": target,
                    "target_label": metric_info.get(target, {}).get("label", target),
                    "target_semantic_type": metric_info.get(target, {}).get("semantic_type", "unknown"),
                    "edge_type": relation.get("edge_type"),
                    "relation_type": relation.get("relation_type"),
                    "evidence_type": relation.get("evidence_type"),
                    "evidence": relation.get("evidence"),
                    "needs_approval": bool(relation.get("needs_approval")),
                    "confidence": relation.get("confidence"),
                }
            )
        contexts.sort(
            key=lambda item: (
                bool(item.get("needs_approval")),
                -float(item.get("confidence") or 0.0),
                str(item.get("evidence_type") or ""),
                str(item.get("source_metric_code") or ""),
                str(item.get("target_metric_code") or ""),
            )
        )
        return contexts[:RELATION_GATE_CONTEXT_LIMIT]

    def _metric_info_by_code(self, metric_candidates: list[dict[str, Any]]) -> dict[str, dict[str, str]]:
        result: dict[str, dict[str, str]] = {}
        for item in metric_candidates:
            code = str(item.get("canonical_code") or item.get("metric_code") or "")
            if not code or code in result and result[code].get("semantic_type") != "unknown":
                continue
            result[code] = {
                "label": str(item.get("label") or item.get("raw_label") or code),
                "semantic_type": str(item.get("semantic_type") or "unknown"),
            }
        return result

    def _apply_relation_gate_decision(self, relation: dict[str, Any], decision: dict[str, Any]) -> None:
        gate_decision = str(decision.get("decision") or "keep_pending")
        relation["relation_gate_decision"] = gate_decision
        relation["relation_gate_model"] = str(decision.get("model") or "")
        reason = str(decision.get("reason") or "")
        if reason:
            relation["evidence"] = f"{relation.get('evidence') or ''} LLM gate: {reason}".strip()
        edge_type = str(decision.get("edge_type") or "")
        if edge_type:
            relation["edge_type"] = edge_type
            relation["relation_type"] = "component" if edge_type == "component" else edge_type
        confidence = _safe_float(decision.get("confidence"))
        if confidence is not None and confidence > 0:
            relation["confidence"] = round(max(0.0, min(1.0, float(confidence))), 3)
        if gate_decision == "approve":
            relation["needs_approval"] = False
            relation["needs_approval_reason"] = ""
        elif gate_decision == "reject":
            relation["rejected_by_relation_gate"] = True
            relation["needs_approval"] = True
            relation["needs_approval_reason"] = _append_reason(
                relation.get("needs_approval_reason"),
                "rejected by local LLM relation gate",
            )
        else:
            relation["needs_approval"] = True
            relation["needs_approval_reason"] = _append_reason(
                relation.get("needs_approval_reason"),
                "local LLM relation gate kept this relation pending",
            )

    def _dependency_rule_rows(self, relations: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows = []
        for relation in relations:
            if relation.get("needs_approval"):
                continue
            rows.append(
                {
                    "source_metric_code": relation["source_metric_code"],
                    "target_metric_code": relation["target_metric_code"],
                    "edge_type": relation.get("edge_type") or edge_type_for_relation(str(relation.get("relation_type") or "driver")),
                    "reason": f"{relation.get('evidence_type')}: {relation.get('evidence')}",
                    "strength": relation.get("score") or relation.get("confidence") or 0.5,
                }
            )
        return rows

    def _quality_report(
        self,
        *,
        profile: dict[str, Any],
        long_rows: list[dict[str, Any]],
        pivot_rows: list[dict[str, Any]],
        metric_candidates: list[dict[str, Any]],
        relation_candidates: list[dict[str, Any]],
        warnings: list[str],
    ) -> dict[str, Any]:
        technical_candidates = [item for item in metric_candidates if item.get("semantic_type") == "technical_check"]
        duplicate_period_count = sum(len(sheet.get("duplicate_periods") or []) for sheet in profile.get("sheets", []))
        return {
            "workbook": profile.get("workbook_name"),
            "sheet_count": len(profile.get("sheets", [])),
            "long_fact_row_count": len(long_rows),
            "pivot_fact_row_count": len(pivot_rows),
            "metric_candidate_count": len(metric_candidates),
            "relation_candidate_count": len(relation_candidates),
            "technical_metric_candidate_count": len(technical_candidates),
            "duplicate_period_warning_count": duplicate_period_count,
            "relation_evidence_types": dict(Counter(item.get("evidence_type") for item in relation_candidates)),
            "metric_statuses": dict(Counter(item.get("status") for item in metric_candidates)),
            "warnings": [*profile.get("warnings", []), *warnings],
        }

    def _manifest_for_outputs(
        self,
        pivoted_facts_path: Path,
        metric_candidates_path: Path,
        dependency_rules_path: Path,
        pivot_headers: list[str],
    ) -> dict[str, Any]:
        metric_codes = [header for header in pivot_headers if header not in PIVOT_DIMENSIONS]
        return {
            "version": 1,
            "poc_id": "relation-memory-dynamic-normalized",
            "status": "generated",
            "goal": {"summary": "Generated normalized relation-memory manifest."},
            "incoming_csvs": [
                {
                    "key": "normalized_facts",
                    "path": self._relative_to_project(pivoted_facts_path),
                    "required": True,
                    "kind": "numeric_facts",
                    "grain": PIVOT_DIMENSIONS,
                    "required_columns": [*PIVOT_DIMENSIONS, *metric_codes],
                    "derives": {"metrics": metric_codes, "dimensions": PIVOT_DIMENSIONS},
                },
                {
                    "key": "metric_dictionary",
                    "path": self._relative_to_project(metric_candidates_path),
                    "required": False,
                    "kind": "metadata",
                    "required_columns": [
                        "metric_code",
                        "label",
                        "description",
                        "aliases",
                        "sensitivity_level",
                        "allow_roles",
                        "preferred_dataset",
                    ],
                },
                {
                    "key": "dependency_rules",
                    "path": self._relative_to_project(dependency_rules_path),
                    "required": False,
                    "kind": "metadata",
                    "required_columns": [
                        "source_metric_code",
                        "target_metric_code",
                        "edge_type",
                        "reason",
                        "strength",
                    ],
                },
            ],
            "golden_queries": [],
        }

    def _relative_to_project(self, path: Path) -> str:
        try:
            return str(path.resolve().relative_to(BASE_DIR.resolve()))
        except ValueError:
            return str(path)

    def _write_xlsx(self, path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name[:31]
        worksheet.append(headers)
        for row in rows:
            if sheet_name == "metric_candidates":
                row = self._metric_dictionary_export_row(row)
            worksheet.append([row.get(header) for header in headers])
        workbook.save(path)
        workbook.close()

    def _metric_dictionary_export_row(self, row: dict[str, Any]) -> dict[str, Any]:
        exported = dict(row)
        exported.setdefault("metric_code", row.get("canonical_code"))
        exported.setdefault("description", row.get("evidence") or "")
        exported.setdefault("sensitivity_level", "internal")
        exported.setdefault("allow_roles", "admin,analyst")
        exported.setdefault("preferred_dataset", "normalized_facts")
        return exported


def make_dynamic_output_dir(root: str | Path, prefix: str = "relation_memory_dynamic") -> Path:
    path = Path(root) / f"{prefix}_{uuid.uuid4().hex[:12]}"
    path.mkdir(parents=True, exist_ok=True)
    return path
