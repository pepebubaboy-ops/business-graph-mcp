from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.services.relation_memory_ingestion import normalize_metric_code
from app.services.relation_memory_poc import _canonical_header, _canonical_metric_code, _safe_float


DIMENSION_COLUMNS = {
    "month",
    "region",
    "business_unit",
    "warehouse",
    "product_category",
    "scenario",
    "report_sheet",
    "source_workbook",
}
ROW_METRIC_HEADER_CANDIDATES = {"показатель", "metric", "metrics", "indicator", "метрика"}
MAX_GENERATED_METRIC_CODE_LENGTH = 96
ROW_REPORT_RELATION_SOURCE = "row_report_structure"


def normalize_row_metric_xlsx(
    path: Path,
    index: int,
    temp_dir: Path,
    inferred_dependency_priors: list[dict[str, Any]],
) -> Path | None:
    temp_root = Path(temp_dir)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        month_rows: dict[str, dict[str, float]] = {}
        metric_codes: list[str] = []
        used_codes: set[str] = set()
        dependency_keys: set[tuple[str, str]] = set()
        for sheet in workbook.worksheets:
            header = _find_row_metric_header(sheet)
            if not header:
                continue
            header_row, metric_col, period_cols = header
            report_events: list[dict[str, Any]] = []
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                label = _cell_value(row, metric_col)
                marker = _report_marker(label)
                if marker:
                    report_events.append(
                        {"kind": marker, "row_number": row_number, "label": str(label or "")}
                    )
                    continue
                if not _looks_like_metric_label(label):
                    continue
                values_by_month: dict[str, float] = {}
                for col_idx, month in period_cols.items():
                    numeric_value = _safe_float(_cell_value(row, col_idx))
                    if numeric_value is not None:
                        values_by_month[month] = numeric_value
                if len(values_by_month) < 2:
                    continue

                metric_code = _generated_metric_code(
                    sheet.title, str(label), row_number, used_codes
                )
                used_codes.add(metric_code)
                metric_codes.append(metric_code)
                report_events.append(
                    {
                        "kind": "metric",
                        "row_number": row_number,
                        "label": str(label),
                        "metric_code": metric_code,
                    }
                )
                for month, value in values_by_month.items():
                    month_rows.setdefault(month, {})[metric_code] = value
            _add_row_report_dependency_priors(
                sheet_name=sheet.title,
                events=report_events,
                dependency_priors=inferred_dependency_priors,
                dependency_keys=dependency_keys,
            )
    finally:
        workbook.close()

    if len(month_rows) < 2 or not metric_codes:
        return None

    normalized_workbook = Workbook()
    normalized_sheet = normalized_workbook.active
    normalized_sheet.title = "normalized_facts"
    normalized_sheet.append(["month", *metric_codes])
    for month in sorted(month_rows):
        values = month_rows[month]
        normalized_sheet.append([month, *[values.get(metric_code) for metric_code in metric_codes]])

    normalized_path = temp_root / f"normalized_{index}_{normalize_metric_code(path.stem)}.xlsx"
    normalized_workbook.save(normalized_path)
    normalized_workbook.close()
    return normalized_path


def column_has_numeric_values(
    canonical_header: str, raw_headers: list[str], rows: list[dict[str, Any]]
) -> bool:
    for raw_header in raw_headers:
        raw_canonical = _canonical_header(raw_header)
        if raw_canonical == str(raw_header).strip():
            raw_canonical = _canonical_metric_code(raw_header)
        if raw_canonical != canonical_header:
            continue
        return any(_safe_float(row.get(raw_header)) is not None for row in rows[:50])
    return False


def _find_row_metric_header(sheet: Any) -> tuple[int, int, dict[int, str]] | None:
    max_scan_rows = min(sheet.max_row or 0, 10)
    for row_idx in range(1, max_scan_rows + 1):
        values = next(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), ())
        metric_col = None
        period_cols: dict[int, str] = {}
        for col_idx, value in enumerate(values, start=1):
            text = str(value or "").strip()
            normalized_text = normalize_metric_code(text) if text else ""
            if normalized_text == "metric" and text.lower() not in {"metric", "metrics"}:
                normalized_text = ""
            if normalized_text in ROW_METRIC_HEADER_CANDIDATES:
                metric_col = col_idx
            month = _month_from_period_header(value)
            if month:
                period_cols[col_idx] = month
        if metric_col and len(period_cols) >= 2:
            return row_idx, metric_col, period_cols
    return None


def _month_from_period_header(value: Any) -> str | None:
    if isinstance(value, datetime):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, date):
        return f"{value.year:04d}-{value.month:02d}"
    return None


def _looks_like_metric_label(value: Any) -> bool:
    text = str(value or "").strip()
    if len(text) < 3:
        return False
    lowered = text.lower()
    if lowered.endswith(":") and not any(char.isdigit() for char in lowered):
        return False
    return (
        lowered not in {"показатель", "в том числе", "в том числе:"}
        and "расшифровка" not in lowered
    )


def _report_marker(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    normalized = normalize_metric_code(text)
    lowered = text.lower()
    if normalized == "показатель":
        return "header"
    if "в том числе" in lowered:
        return "parts"
    if "расшифровка" in lowered:
        return "breakdown"
    return None


def _add_row_report_dependency_priors(
    *,
    sheet_name: str,
    events: list[dict[str, Any]],
    dependency_priors: list[dict[str, Any]],
    dependency_keys: set[tuple[str, str]],
) -> None:
    segment_parent: dict[str, Any] | None = None
    current_parent: dict[str, Any] | None = None
    previous_metric: dict[str, Any] | None = None
    allow_subsections = False

    for event in events:
        if event["kind"] == "header":
            segment_parent = None
            current_parent = None
            previous_metric = None
            allow_subsections = False
            continue
        if event["kind"] == "parts":
            current_parent = previous_metric or segment_parent
            allow_subsections = True
            continue
        if event["kind"] == "breakdown":
            current_parent = segment_parent or previous_metric
            allow_subsections = False
            continue

        if event["kind"] != "metric":
            continue
        if segment_parent is None:
            segment_parent = event
            current_parent = event
            previous_metric = event
            continue

        is_subsection = allow_subsections and _looks_like_subsection_parent(event["label"])
        targets = [segment_parent] if is_subsection else [current_parent or segment_parent]
        if (
            not is_subsection
            and current_parent
            and segment_parent
            and current_parent["metric_code"] != segment_parent["metric_code"]
        ):
            targets.append(segment_parent)

        for target in targets:
            _append_report_dependency_prior(
                sheet_name=sheet_name,
                source=event,
                target=target,
                dependency_priors=dependency_priors,
                dependency_keys=dependency_keys,
            )

        if is_subsection:
            current_parent = event
        previous_metric = event


def _append_report_dependency_prior(
    *,
    sheet_name: str,
    source: dict[str, Any],
    target: dict[str, Any],
    dependency_priors: list[dict[str, Any]],
    dependency_keys: set[tuple[str, str]],
) -> None:
    source_code = str(source.get("metric_code") or "")
    target_code = str(target.get("metric_code") or "")
    if not source_code or not target_code or source_code == target_code:
        return
    key = (source_code, target_code)
    if key in dependency_keys:
        return
    dependency_keys.add(key)
    source_label = str(source.get("label") or source_code)
    target_label = str(target.get("label") or target_code)
    edge_type = _report_relation_edge_type(source_label)
    dependency_priors.append(
        {
            "source_metric_code": source_code,
            "target_metric_code": target_code,
            "edge_type": edge_type,
            "strength": 0.7,
            "source": ROW_REPORT_RELATION_SOURCE,
            "note": (
                f"В листе «{sheet_name}» строка «{source_label}» находится в блоке показателя "
                f"«{target_label}»."
            ),
        }
    )


def _looks_like_subsection_parent(label: str) -> bool:
    lowered = label.lower()
    if ":" in label[:80]:
        return True
    return lowered.startswith(
        (
            "себестоимость накл. нтк",
            "себестоимость накл. своего",
            "себестоимость своего парка",
            "нтк ",
            "свой парк",
        )
    )


def _report_relation_edge_type(source_label: str) -> str:
    lowered = source_label.lower()
    driver_tokens = ("количество", "кол-во", "пробег", "накладные", "водители", "выручка", "доля")
    if any(token in lowered for token in driver_tokens):
        return "driver"
    return "component"


def _cell_value(row: tuple[Any, ...], one_based_col: int) -> Any:
    index = one_based_col - 1
    if index >= len(row):
        return None
    return row[index]


def _generated_metric_code(
    sheet_name: str, label: str, row_number: int, used_codes: set[str]
) -> str:
    sheet_code = normalize_metric_code(sheet_name)[:24].strip("_")
    label_code = normalize_metric_code(label)[:MAX_GENERATED_METRIC_CODE_LENGTH].strip("_")
    base = "_".join(part for part in [sheet_code, label_code] if part)
    if not base:
        base = f"metric_{row_number}"
    code = base
    suffix = 2
    while code in used_codes:
        code = f"{base}_{suffix}"
        suffix += 1
    return code
