from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import get_column_letter, range_boundaries

from app.services.relation_memory_agent import LlmJsonError, RelationMemoryLlmClient, normalize_candidate_payload
from app.services.relation_memory_business_literature import BusinessLiteratureRetriever
from app.services.relation_memory_ingestion import normalize_metric_code


SCAN_UNIT_ROW_SPAN = 24
SCAN_UNIT_COL_SPAN = 12
ROW_GAP_TOLERANCE = 2
COL_GAP_TOLERANCE = 2
MAX_GENERATED_METRIC_CODE_LENGTH = 96
RELATION_PAIR_BATCH_SIZE = 8
FULL_PAIR_SCAN_UNIT_LIMIT = 8


@dataclass
class AgenticWorkbookEnrichmentResult:
    metrics: list[dict[str, Any]] = field(default_factory=list)
    metric_dictionary_rows: list[dict[str, Any]] = field(default_factory=list)
    dependency_rule_rows: list[dict[str, Any]] = field(default_factory=list)
    candidate_relations: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class WorkbookBlock:
    id: str
    sheet_name: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    non_empty_cells: int
    formula_cells: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "sheet_name": self.sheet_name,
            "range": _range_string(self.min_row, self.max_row, self.min_col, self.max_col),
            "min_row": self.min_row,
            "max_row": self.max_row,
            "min_col": self.min_col,
            "max_col": self.max_col,
            "non_empty_cells": self.non_empty_cells,
            "formula_cells": self.formula_cells,
        }


@dataclass
class WorkbookScanUnit:
    id: str
    sheet_name: str
    block_id: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "sheet_name": self.sheet_name,
            "block_id": self.block_id,
            "range": _range_string(self.min_row, self.max_row, self.min_col, self.max_col),
            "min_row": self.min_row,
            "max_row": self.max_row,
            "min_col": self.min_col,
            "max_col": self.max_col,
        }


@dataclass
class WorkbookFormulaCell:
    sheet_name: str
    cell: str
    formula: str
    references: list[dict[str, Any]]


@dataclass
class WorkbookSheetIndex:
    sheet_name: str
    max_row: int
    max_col: int
    non_empty_cells: int
    formula_cells: int
    merged_ranges: list[str]
    blocks: list[WorkbookBlock] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "max_row": self.max_row,
            "max_col": self.max_col,
            "non_empty_cells": self.non_empty_cells,
            "formula_cells": self.formula_cells,
            "merged_ranges": self.merged_ranges,
            "blocks": [block.to_dict() for block in self.blocks],
        }


@dataclass
class WorkbookIndex:
    document_id: str
    filename: str
    sheets: list[WorkbookSheetIndex]
    scan_units: list[WorkbookScanUnit]
    formula_cells: list[WorkbookFormulaCell]


@dataclass
class _RowSegment:
    row_index: int
    start_col: int
    end_col: int
    non_empty_cells: int
    formula_cells: int


@dataclass
class _BlockBuilder:
    id: str
    sheet_name: str
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    non_empty_cells: int
    formula_cells: int

    @classmethod
    def from_segment(cls, *, block_id: str, sheet_name: str, segment: _RowSegment) -> _BlockBuilder:
        return cls(
            id=block_id,
            sheet_name=sheet_name,
            min_row=segment.row_index,
            max_row=segment.row_index,
            min_col=segment.start_col,
            max_col=segment.end_col,
            non_empty_cells=segment.non_empty_cells,
            formula_cells=segment.formula_cells,
        )

    def extend(self, segment: _RowSegment) -> None:
        self.max_row = max(self.max_row, segment.row_index)
        self.min_col = min(self.min_col, segment.start_col)
        self.max_col = max(self.max_col, segment.end_col)
        self.non_empty_cells += segment.non_empty_cells
        self.formula_cells += segment.formula_cells

    def absorb(self, other: _BlockBuilder) -> None:
        self.min_row = min(self.min_row, other.min_row)
        self.max_row = max(self.max_row, other.max_row)
        self.min_col = min(self.min_col, other.min_col)
        self.max_col = max(self.max_col, other.max_col)
        self.non_empty_cells += other.non_empty_cells
        self.formula_cells += other.formula_cells

    def build(self) -> WorkbookBlock:
        return WorkbookBlock(
            id=self.id,
            sheet_name=self.sheet_name,
            min_row=self.min_row,
            max_row=self.max_row,
            min_col=self.min_col,
            max_col=self.max_col,
            non_empty_cells=self.non_empty_cells,
            formula_cells=self.formula_cells,
        )


class AgenticWorkbookRelationEnricher:
    """Deterministic workbook enrichment with optional LLM relation judging."""

    def __init__(
        self,
        *,
        llm_client: RelationMemoryLlmClient | None = None,
        literature_retriever: BusinessLiteratureRetriever | None = None,
        enable_llm_relation_judging: bool = False,
    ):
        self.llm_client = llm_client or RelationMemoryLlmClient()
        self.literature_retriever = literature_retriever or BusinessLiteratureRetriever()
        self.enable_llm_relation_judging = enable_llm_relation_judging

    def enrich_workbook(self, path: str | Path, *, document_id: str, filename: str) -> AgenticWorkbookEnrichmentResult:
        result = AgenticWorkbookEnrichmentResult()
        workbook = load_workbook(path, data_only=False, read_only=False)
        try:
            index = _build_workbook_index(workbook=workbook, document_id=document_id, filename=filename)
            seen_metric_codes: set[str] = set()
            proposed_pairs: list[dict[str, Any]] = []
            formula_pairs: list[dict[str, Any]] = []
            seen_pair_keys: set[tuple[str, str]] = set()
            scan_unit_metric_groups: list[list[dict[str, Any]]] = []

            for scan_unit in index.scan_units:
                tool_outputs = _run_scan_tools(workbook=workbook, index=index, scan_unit=scan_unit)
                scan_unit_metrics: list[dict[str, Any]] = []
                for candidate_row in _build_metric_candidate_rows(tool_outputs):
                    metric = _metric_from_candidate_row(
                        candidate_row,
                        document_id=document_id,
                        scan_unit=scan_unit,
                        used_metric_codes=seen_metric_codes,
                    )
                    result.metrics.append(metric)
                    scan_unit_metrics.append(metric)
                if scan_unit_metrics:
                    scan_unit_metric_groups.append(scan_unit_metrics)

            formula_pairs = _propose_formula_relation_pairs(index=index, metrics=result.metrics)
            _append_unique_pairs(proposed_pairs, formula_pairs, seen_pair_keys=seen_pair_keys)
            for scan_unit_metrics in scan_unit_metric_groups:
                _append_unique_pairs(
                    proposed_pairs,
                    _propose_scan_unit_relation_pairs(scan_unit_metrics),
                    seen_pair_keys=seen_pair_keys,
                )

            result.metric_dictionary_rows = [_metric_dictionary_row(metric) for metric in result.metrics]
            relations = self._judge_relation_pairs(index=index, proposed_pairs=proposed_pairs, warnings=result.warnings)
            relations.extend(_deterministic_formula_relations(formula_pairs, document_id=document_id))
            result.candidate_relations = _dedupe_relations(relations)
            result.dependency_rule_rows = [_dependency_rule_row(relation) for relation in result.candidate_relations]
        finally:
            workbook.close()
        return result

    def _judge_relation_pairs(
        self,
        *,
        index: WorkbookIndex,
        proposed_pairs: list[dict[str, Any]],
        warnings: list[str],
    ) -> list[dict[str, Any]]:
        if not self.enable_llm_relation_judging or not proposed_pairs:
            return []
        relations: list[dict[str, Any]] = []
        for batch in _batched(proposed_pairs, RELATION_PAIR_BATCH_SIZE):
            try:
                relations.extend(self._classify_relation_batch(index=index, proposed_pairs=batch))
            except LlmJsonError as exc:
                warnings.append(f"agentic_relation_llm_fallback:{exc}")
        return relations

    def _classify_relation_batch(self, *, index: WorkbookIndex, proposed_pairs: list[dict[str, Any]]) -> list[dict[str, Any]]:
        business_literature_context = self.literature_retriever.retrieve_for_pairs(proposed_pairs)
        system_prompt = (
            "You judge relationships between existing spreadsheet metric candidates. "
            "Do not create, rename, merge, or extract metrics. "
            "Use only metric codes that appear in proposed_pairs. "
            "Use business_literature_context for grounding, but do not create a relation from literature alone. "
            "Return strict JSON with key candidate_relations only."
        )
        payload = {
            "task": "classify_metric_relations",
            "document_id": index.document_id,
            "proposed_pairs": proposed_pairs,
            "business_literature_context": business_literature_context,
            "instruction": (
                "For each proposed pair, either return one directed relation or omit it. "
                "Relations must include source_metric_code, target_metric_code, edge_type, "
                "lag_period, note, evidence, confidence, and source_document_id."
            ),
        }
        response = _normalize_task_payload(
            payload=self.llm_client.chat_json(system_prompt=system_prompt, user_payload=payload),
            list_fields=("candidate_relations",),
            alias_map={"candidate_relations": ("relations", "extracted_relations", "relation_candidates")},
        )
        if "candidate_relations" not in response or not isinstance(response["candidate_relations"], list):
            raise LlmJsonError("Relation batch response must include candidate_relations list.")

        allowed_codes = {
            str(metric["code"])
            for pair in proposed_pairs
            for metric in (pair.get("left_metric", {}), pair.get("right_metric", {}))
            if metric.get("code")
        }
        allowed_pairs = {
            tuple(
                sorted(
                    (
                        str(pair.get("left_metric", {}).get("code") or ""),
                        str(pair.get("right_metric", {}).get("code") or ""),
                    )
                )
            )
            for pair in proposed_pairs
            if pair.get("left_metric", {}).get("code") and pair.get("right_metric", {}).get("code")
        }

        valid_relations: list[dict[str, Any]] = []
        for raw_relation in response["candidate_relations"]:
            normalized = normalize_candidate_payload(raw_relation, default_source="llm_agentic_excel_relation")
            if normalized is None:
                continue
            source_code = normalized["source_metric_code"]
            target_code = normalized["target_metric_code"]
            if source_code not in allowed_codes or target_code not in allowed_codes:
                continue
            if tuple(sorted((source_code, target_code))) not in allowed_pairs:
                continue
            normalized["source_document_id"] = normalized.get("source_document_id") or index.document_id
            normalized["evidence_type"] = normalized.get("evidence_type") or "llm_agentic"
            valid_relations.append(normalized)
        return valid_relations


def _build_workbook_index(*, workbook: Workbook, document_id: str, filename: str) -> WorkbookIndex:
    sheet_indexes: list[WorkbookSheetIndex] = []
    scan_units: list[WorkbookScanUnit] = []
    formula_cells: list[WorkbookFormulaCell] = []

    for sheet in workbook.worksheets:
        max_row = int(sheet.max_row or 0)
        max_col = int(sheet.max_column or 0)
        active_blocks: dict[str, _BlockBuilder] = {}
        finalized_blocks: list[_BlockBuilder] = []
        block_counter = 1
        sheet_non_empty_cells = 0
        sheet_formula_cells = 0

        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            row_index = row[0].row if row else 0
            row_segments, row_formula_cells = _extract_row_segments(row)
            for formula_item in row_formula_cells:
                formula_cells.append(
                    WorkbookFormulaCell(
                        sheet_name=sheet.title,
                        cell=formula_item["cell"],
                        formula=formula_item["formula"],
                        references=_parse_formula_references(formula_item["formula"], current_sheet=sheet.title),
                    )
                )
            sheet_formula_cells += len(row_formula_cells)

            touched_block_ids: set[str] = set()
            for segment in row_segments:
                sheet_non_empty_cells += segment.non_empty_cells
                overlapping_ids = [
                    block_id
                    for block_id, block in active_blocks.items()
                    if _segment_overlaps_block(segment=segment, block=block)
                ]
                if not overlapping_ids:
                    block_id = f"{sheet.title}::block_{block_counter}"
                    block_counter += 1
                    active_blocks[block_id] = _BlockBuilder.from_segment(
                        block_id=block_id,
                        sheet_name=sheet.title,
                        segment=segment,
                    )
                    touched_block_ids.add(block_id)
                    continue

                primary_id = overlapping_ids[0]
                primary = active_blocks[primary_id]
                primary.extend(segment)
                touched_block_ids.add(primary_id)
                for merge_id in overlapping_ids[1:]:
                    if merge_id == primary_id:
                        continue
                    primary.absorb(active_blocks[merge_id])
                    active_blocks.pop(merge_id, None)

            stale_ids = [
                block_id
                for block_id, block in active_blocks.items()
                if block_id not in touched_block_ids and block.max_row < row_index - ROW_GAP_TOLERANCE
            ]
            for stale_id in stale_ids:
                finalized_blocks.append(active_blocks.pop(stale_id))

        finalized_blocks.extend(active_blocks.values())
        workbook_blocks = [builder.build() for builder in finalized_blocks]
        sheet_index = WorkbookSheetIndex(
            sheet_name=sheet.title,
            max_row=max_row,
            max_col=max_col,
            non_empty_cells=sheet_non_empty_cells,
            formula_cells=sheet_formula_cells,
            merged_ranges=[str(item) for item in sheet.merged_cells.ranges],
            blocks=workbook_blocks,
        )
        sheet_indexes.append(sheet_index)
        for block in workbook_blocks:
            scan_units.extend(_build_scan_units_for_block(block=block))

    return WorkbookIndex(document_id=document_id, filename=filename, sheets=sheet_indexes, scan_units=scan_units, formula_cells=formula_cells)


def _extract_row_segments(row: tuple[Any, ...]) -> tuple[list[_RowSegment], list[dict[str, Any]]]:
    segments: list[_RowSegment] = []
    formula_cells: list[dict[str, Any]] = []
    segment_start: int | None = None
    segment_non_empty = 0
    segment_formula = 0
    empty_gap = 0

    for cell in row:
        value = cell.value
        if isinstance(value, str) and value.startswith("="):
            formula_cells.append({"cell": cell.coordinate, "formula": value})
        if _is_non_empty(value):
            if segment_start is None:
                segment_start = cell.column
                segment_non_empty = 0
                segment_formula = 0
                empty_gap = 0
            segment_non_empty += 1
            if isinstance(value, str) and value.startswith("="):
                segment_formula += 1
            empty_gap = 0
            continue
        if segment_start is not None:
            empty_gap += 1
            if empty_gap <= COL_GAP_TOLERANCE:
                continue
            segments.append(
                _RowSegment(
                    row_index=cell.row,
                    start_col=segment_start,
                    end_col=cell.column - empty_gap,
                    non_empty_cells=segment_non_empty,
                    formula_cells=segment_formula,
                )
            )
            segment_start = None
            segment_non_empty = 0
            segment_formula = 0
            empty_gap = 0

    if segment_start is not None and row:
        last_cell = row[-1]
        segments.append(
            _RowSegment(
                row_index=last_cell.row,
                start_col=segment_start,
                end_col=last_cell.column - empty_gap,
                non_empty_cells=segment_non_empty,
                formula_cells=segment_formula,
            )
        )
    return segments, formula_cells


def _segment_overlaps_block(*, segment: _RowSegment, block: _BlockBuilder) -> bool:
    if segment.row_index > block.max_row + ROW_GAP_TOLERANCE + 1:
        return False
    return not (
        segment.end_col < block.min_col - COL_GAP_TOLERANCE - 1
        or segment.start_col > block.max_col + COL_GAP_TOLERANCE + 1
    )


def _build_scan_units_for_block(*, block: WorkbookBlock) -> list[WorkbookScanUnit]:
    scan_units: list[WorkbookScanUnit] = []
    row_start = block.min_row
    while row_start <= block.max_row:
        row_end = min(row_start + SCAN_UNIT_ROW_SPAN - 1, block.max_row)
        col_start = block.min_col
        while col_start <= block.max_col:
            col_end = min(col_start + SCAN_UNIT_COL_SPAN - 1, block.max_col)
            scan_units.append(
                WorkbookScanUnit(
                    id=f"{block.sheet_name}::{block.id}::R{row_start}-{row_end}::C{col_start}-{col_end}",
                    sheet_name=block.sheet_name,
                    block_id=block.id,
                    min_row=row_start,
                    max_row=row_end,
                    min_col=col_start,
                    max_col=col_end,
                )
            )
            col_start = col_end + 1
        row_start = row_end + 1
    return scan_units


def _run_scan_tools(*, workbook: Workbook, index: WorkbookIndex, scan_unit: WorkbookScanUnit) -> dict[str, Any]:
    return {
        "describe_scan_unit": _tool_describe_scan_unit(index=index, scan_unit=scan_unit),
        "read_scan_unit_cells": _tool_read_scan_unit_cells(workbook=workbook, scan_unit=scan_unit),
        "read_scan_unit_formula_cells": _tool_read_scan_unit_formula_cells(index=index, scan_unit=scan_unit),
    }


def _metric_from_candidate_row(
    raw: dict[str, Any],
    *,
    document_id: str,
    scan_unit: WorkbookScanUnit,
    used_metric_codes: set[str],
) -> dict[str, Any]:
    label = str(raw.get("label") or "").strip()
    row_index = raw.get("row_index")
    base_code = normalize_metric_code(label)[:MAX_GENERATED_METRIC_CODE_LENGTH].strip("_")
    if not base_code or base_code == "metric":
        base_code = f"metric_{row_index or len(used_metric_codes) + 1}"
    metric_code = base_code
    suffix = 2
    while metric_code in used_metric_codes:
        metric_code = f"{base_code}_{suffix}"
        suffix += 1
    used_metric_codes.add(metric_code)

    label_cell_ref = str(raw.get("label_cell") or "").strip()
    value_cell_refs = [
        str(item.get("cell") or "").strip()
        for item in raw.get("value_cells", []) or []
        if str(item.get("cell") or "").strip()
    ]
    aliases = [label, metric_code, metric_code.replace("_", " ")]
    return {
        "code": metric_code,
        "label": label,
        "aliases": [item for item in dict.fromkeys(aliases) if item],
        "source_document_id": document_id,
        "source_sheet": scan_unit.sheet_name,
        "source_block_id": scan_unit.block_id,
        "source_scan_unit_id": scan_unit.id,
        "row_index": row_index,
        "label_cell_ref": label_cell_ref,
        "value_cell_refs": value_cell_refs,
        "cell_refs": [cell_ref for cell_ref in [label_cell_ref, *value_cell_refs] if cell_ref],
        "source": "agentic_excel_metric_candidate",
        "confidence": 1.0,
    }


def _metric_dictionary_row(metric: dict[str, Any]) -> dict[str, Any]:
    code = str(metric.get("code") or "")
    label = str(metric.get("label") or code)
    return {
        "metric_code": code,
        "label": label,
        "description": f"Metric identified from workbook scan unit: {label}",
        "aliases": "|".join(str(item) for item in metric.get("aliases", []) if item),
        "sensitivity_level": "internal",
        "allow_roles": "admin,analyst",
        "preferred_dataset": "agentic_excel_enrichment",
        "semantic_type": "unknown",
    }


def _dependency_rule_row(relation: dict[str, Any]) -> dict[str, Any]:
    return {
        "source_metric_code": relation.get("source_metric_code"),
        "target_metric_code": relation.get("target_metric_code"),
        "edge_type": relation.get("edge_type") or "driver",
        "reason": relation.get("note") or relation.get("evidence") or "Agentic workbook enrichment.",
        "strength": relation.get("confidence") or relation.get("score") or 0.7,
    }


def _build_metric_candidate_rows(tool_outputs: dict[str, Any]) -> list[dict[str, Any]]:
    candidate_rows: list[dict[str, Any]] = []
    rows = tool_outputs.get("read_scan_unit_cells", {}).get("rows", [])
    for row in rows:
        cells = row.get("cells", [])
        text_cells = [
            cell
            for cell in cells
            if isinstance(cell.get("value"), str)
            and str(cell.get("value") or "").strip()
            and not str(cell.get("value") or "").lstrip().startswith("=")
        ]
        value_cells = [
            cell
            for cell in cells
            if isinstance(cell.get("value"), (int, float))
            or (isinstance(cell.get("value"), str) and str(cell.get("value") or "").startswith("="))
        ]
        if not text_cells or not value_cells:
            continue
        for text_cell in text_cells:
            text_cell_ref = str(text_cell.get("cell") or "")
            if not text_cell_ref:
                continue
            _, label_col = _cell_to_row_col(text_cell_ref)
            values_after_label = [
                cell
                for cell in value_cells
                if cell.get("cell") and _cell_to_row_col(str(cell.get("cell")))[1] > label_col
            ]
            candidate_label = str(text_cell.get("value") or "").strip()
            if values_after_label and _looks_like_metric_row_label(candidate_label):
                candidate_rows.append(
                    {
                        "row_index": row.get("row_index"),
                        "label": candidate_label,
                        "label_cell": text_cell_ref,
                        "value_cells": [
                            {"cell": cell.get("cell"), "value": cell.get("value")}
                            for cell in values_after_label[:8]
                        ],
                    }
                )
                break
    return candidate_rows


def _looks_like_metric_row_label(label: str) -> bool:
    normalized_label = " ".join(label.lower().split())
    if len(label) < 3 or normalized_label in {"показатель", "metric", "metrics"}:
        return False
    if not re.search(r"[a-zа-яё]", normalized_label, flags=re.IGNORECASE):
        return False
    if normalized_label.startswith("#") or normalized_label.startswith("показатель"):
        return False
    if normalized_label in {"абсолютные данные", "абсолютные значения", "комментарии", "прошлый мес"}:
        return False
    if "в том числе" in normalized_label or "расшифровка" in normalized_label:
        return False
    return True


def _propose_scan_unit_relation_pairs(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    pairs: list[dict[str, Any]] = []
    seen_pair_keys: set[tuple[str, str]] = set()

    def append_pair(left_metric: dict[str, Any], right_metric: dict[str, Any], reason: str) -> None:
        left_code = str(left_metric.get("code") or "")
        right_code = str(right_metric.get("code") or "")
        if not left_code or not right_code or left_code == right_code:
            return
        key = tuple(sorted((left_code, right_code)))
        if key in seen_pair_keys:
            return
        seen_pair_keys.add(key)
        pairs.append(
            _relation_pair_payload(
                left_metric,
                right_metric,
                proposal_reason=reason,
                evidence="Metrics appear near each other in the same bounded workbook scan unit.",
            )
        )

    if len(metrics) <= FULL_PAIR_SCAN_UNIT_LIMIT:
        for left_index, left_metric in enumerate(metrics):
            for right_metric in metrics[left_index + 1 :]:
                append_pair(left_metric, right_metric, "same_scan_unit")
        return pairs

    anchor_metric = metrics[0]
    for metric in metrics[1:]:
        append_pair(anchor_metric, metric, "scan_unit_anchor")
    for left_metric, right_metric in zip(metrics, metrics[1:], strict=False):
        append_pair(left_metric, right_metric, "adjacent_metric_rows")
    return pairs


def _propose_formula_relation_pairs(*, index: WorkbookIndex, metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    formula_by_cell = {
        (formula_cell.sheet_name, formula_cell.cell): formula_cell
        for formula_cell in index.formula_cells
    }
    pairs: list[dict[str, Any]] = []
    for target_metric in metrics:
        target_sheet = str(target_metric.get("source_sheet") or "")
        for target_cell in target_metric.get("value_cell_refs", []) or []:
            formula_cell = formula_by_cell.get((target_sheet, str(target_cell)))
            if formula_cell is None:
                continue
            for source_metric in metrics:
                if source_metric.get("code") == target_metric.get("code"):
                    continue
                if not _metric_intersects_formula_references(source_metric, formula_cell.references):
                    continue
                pairs.append(
                    _relation_pair_payload(
                        source_metric,
                        target_metric,
                        proposal_reason="formula_reference",
                        evidence=(
                            f"{target_sheet}!{formula_cell.cell} formula {formula_cell.formula[:180]!r} "
                            f"references cells attached to {source_metric.get('code')}."
                        ),
                    )
                )
    return pairs


def _deterministic_formula_relations(proposed_pairs: list[dict[str, Any]], *, document_id: str) -> list[dict[str, Any]]:
    relations = []
    for pair in proposed_pairs:
        left = pair.get("left_metric", {})
        right = pair.get("right_metric", {})
        relations.append(
            {
                "source_metric_code": left.get("code"),
                "target_metric_code": right.get("code"),
                "edge_type": "driver",
                "relation_type": "driver",
                "note": "Formula reference links source row values to target metric formula.",
                "evidence": pair.get("evidence") or "",
                "evidence_type": "formula",
                "confidence": 0.86,
                "score": 0.86,
                "needs_approval_reason": "Formula-derived relation from nonstandard workbook layout.",
                "source": "agentic_formula_reference",
                "source_document_id": document_id,
            }
        )
    return relations


def _relation_pair_payload(
    left_metric: dict[str, Any],
    right_metric: dict[str, Any],
    *,
    proposal_reason: str,
    evidence: str,
) -> dict[str, Any]:
    left_code = str(left_metric.get("code") or "")
    right_code = str(right_metric.get("code") or "")
    return {
        "pair_id": f"{left_code}__{right_code}",
        "left_metric": _metric_prompt_payload(left_metric),
        "right_metric": _metric_prompt_payload(right_metric),
        "proposal_reason": proposal_reason,
        "evidence": evidence,
    }


def _metric_prompt_payload(metric: dict[str, Any]) -> dict[str, Any]:
    return {
        "code": metric.get("code"),
        "label": metric.get("label"),
        "sheet_name": metric.get("source_sheet"),
        "row_index": metric.get("row_index"),
        "cell_refs": metric.get("cell_refs", []),
        "value_cell_refs": metric.get("value_cell_refs", []),
    }


def _append_unique_pairs(target: list[dict[str, Any]], candidates: list[dict[str, Any]], *, seen_pair_keys: set[tuple[str, str]]) -> None:
    for pair in candidates:
        left_code = str(pair.get("left_metric", {}).get("code") or "")
        right_code = str(pair.get("right_metric", {}).get("code") or "")
        if not left_code or not right_code or left_code == right_code:
            continue
        key = tuple(sorted((left_code, right_code)))
        if key in seen_pair_keys:
            continue
        seen_pair_keys.add(key)
        target.append(pair)


def _dedupe_relations(relations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for relation in relations:
        normalized = normalize_candidate_payload(relation, default_source=str(relation.get("source") or "agentic_excel_relation"))
        if normalized is None:
            continue
        key = (
            normalized["source_metric_code"],
            normalized["target_metric_code"],
            normalized["edge_type"],
            str(normalized.get("lag_period") or ""),
        )
        current = deduped.get(key)
        if current is None or float(normalized.get("confidence") or 0.0) > float(current.get("confidence") or 0.0):
            deduped[key] = normalized
    return list(deduped.values())


def _metric_intersects_formula_references(metric: dict[str, Any], references: list[dict[str, Any]]) -> bool:
    metric_sheet = str(metric.get("source_sheet") or "")
    value_cell_refs = [str(item) for item in metric.get("value_cell_refs", []) or [] if str(item)]
    if not metric_sheet or not value_cell_refs:
        return False
    for reference in references:
        if str(reference.get("sheet_name") or "") != metric_sheet:
            continue
        reference_range = str(reference.get("range") or "").strip()
        if reference_range and any(_cell_ref_in_range(cell_ref, reference_range) for cell_ref in value_cell_refs):
            return True
    return False


def _cell_ref_in_range(cell_ref: str, range_ref: str) -> bool:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        row, col = _cell_to_row_col(cell_ref)
    except (TypeError, ValueError):
        return False
    return min_row <= row <= max_row and min_col <= col <= max_col


def _tool_describe_scan_unit(*, index: WorkbookIndex, scan_unit: WorkbookScanUnit) -> dict[str, Any]:
    sheet_summary = next(sheet for sheet in index.sheets if sheet.sheet_name == scan_unit.sheet_name)
    block_summary = next(block for block in sheet_summary.blocks if block.id == scan_unit.block_id)
    return {"scan_unit": scan_unit.to_dict(), "block": block_summary.to_dict(), "sheet": sheet_summary.to_dict()}


def _tool_read_scan_unit_cells(*, workbook: Workbook, scan_unit: WorkbookScanUnit) -> dict[str, Any]:
    sheet = workbook[scan_unit.sheet_name]
    rows: list[dict[str, Any]] = []
    for row_index in range(scan_unit.min_row, scan_unit.max_row + 1):
        sparse_cells: list[dict[str, Any]] = []
        for col_index in range(scan_unit.min_col, scan_unit.max_col + 1):
            cell = sheet.cell(row=row_index, column=col_index)
            if _is_non_empty(cell.value):
                sparse_cells.append({"cell": cell.coordinate, "value": _json_scalar(cell.value)})
        if sparse_cells:
            rows.append({"row_index": row_index, "cells": sparse_cells})
    return {
        "sheet_name": scan_unit.sheet_name,
        "range": _range_string(scan_unit.min_row, scan_unit.max_row, scan_unit.min_col, scan_unit.max_col),
        "rows": rows,
    }


def _tool_read_scan_unit_formula_cells(*, index: WorkbookIndex, scan_unit: WorkbookScanUnit) -> dict[str, Any]:
    formula_cells = []
    for formula_cell in index.formula_cells:
        if formula_cell.sheet_name != scan_unit.sheet_name:
            continue
        row, col = _cell_to_row_col(formula_cell.cell)
        if scan_unit.min_row <= row <= scan_unit.max_row and scan_unit.min_col <= col <= scan_unit.max_col:
            formula_cells.append(
                {
                    "sheet_name": formula_cell.sheet_name,
                    "cell": formula_cell.cell,
                    "references": [
                        f"{reference.get('sheet_name')}!{reference.get('range')}"
                        for reference in formula_cell.references
                        if reference.get("range")
                    ],
                }
            )
    return {"sheet_name": scan_unit.sheet_name, "scan_unit_id": scan_unit.id, "formula_cells": formula_cells}


def _parse_formula_references(formula_text: str, *, current_sheet: str) -> list[dict[str, Any]]:
    references: list[dict[str, Any]] = []
    try:
        tokens = Tokenizer(formula_text).items
    except Exception:
        return references
    for token in tokens:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        reference = token.value.strip()
        if not reference:
            continue
        ref_sheet, ref_range = _split_formula_reference(reference, current_sheet=current_sheet)
        if ref_range:
            references.append({"sheet_name": ref_sheet, "range": ref_range})
    return references


def _split_formula_reference(reference: str, *, current_sheet: str) -> tuple[str, str]:
    if "!" not in reference:
        return current_sheet, reference.replace("$", "")
    sheet_part, range_part = reference.split("!", 1)
    return sheet_part.strip("'").strip() or current_sheet, range_part.replace("$", "")


def _cell_to_row_col(cell_ref: str) -> tuple[int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_ref)
    if min_col != max_col or min_row != max_row:
        raise ValueError(f"Expected single cell reference, got: {cell_ref}")
    return min_row, min_col


def _range_string(min_row: int, max_row: int, min_col: int, max_col: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _is_non_empty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _json_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    return text


def _batched(items: list[dict[str, Any]], batch_size: int) -> list[list[dict[str, Any]]]:
    return [items[index : index + batch_size] for index in range(0, len(items), batch_size)]


def _normalize_task_payload(
    *,
    payload: dict[str, Any],
    list_fields: tuple[str, ...],
    alias_map: dict[str, tuple[str, ...]] | None = None,
) -> dict[str, Any]:
    alias_map = alias_map or {}
    normalized = dict(payload)
    if isinstance(normalized.get("result"), dict):
        merged = dict(normalized["result"])
        for key, value in normalized.items():
            if key != "result":
                merged.setdefault(key, value)
        normalized = merged
    field_name = normalized.get("name")
    field_value = normalized.get("value")
    if isinstance(field_name, str) and field_name.strip() in list_fields and isinstance(field_value, list):
        normalized[field_name.strip()] = field_value
    for field in list_fields:
        if isinstance(normalized.get(field), list):
            continue
        for alias in alias_map.get(field, ()):
            if isinstance(normalized.get(alias), list):
                normalized[field] = normalized[alias]
                break
    return normalized
