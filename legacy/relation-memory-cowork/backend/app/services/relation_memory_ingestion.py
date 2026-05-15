from __future__ import annotations

import re
import tempfile
import uuid
from dataclasses import asdict, dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".txt", ".md", ".pdf"}
DIMENSION_HEADERS = {"month", "region", "business_unit", "warehouse", "product_category", "scenario"}


class IngestionError(ValueError):
    pass


@dataclass
class UploadedFilePayload:
    filename: str
    content: bytes


@dataclass
class DocumentRecord:
    id: str
    filename: str
    file_type: str
    size_bytes: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class TextChunk:
    id: str
    document_id: str
    text: str
    index: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class TableSummary:
    document_id: str
    sheet_name: str
    headers: list[str]
    row_count: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class DetectedMetric:
    code: str
    label: str
    source_document_id: str
    source: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class DocumentBundle:
    documents: list[DocumentRecord] = field(default_factory=list)
    text_chunks: list[TextChunk] = field(default_factory=list)
    tables: list[TableSummary] = field(default_factory=list)
    detected_metrics: list[DetectedMetric] = field(default_factory=list)
    xlsx_files: list[UploadedFilePayload] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "documents": [item.to_dict() for item in self.documents],
            "text_chunks": [item.to_dict() for item in self.text_chunks],
            "tables": [item.to_dict() for item in self.tables],
            "detected_metrics": [item.to_dict() for item in self.detected_metrics],
            "warnings": list(self.warnings),
        }


def normalize_metric_code(value: str) -> str:
    normalized = re.sub(r"[^a-zа-яё0-9]+", "_", value.lower(), flags=re.IGNORECASE).strip("_")
    if not normalized:
        return "metric"
    return normalized


def ingest_file_payloads(files: list[UploadedFilePayload]) -> DocumentBundle:
    bundle = DocumentBundle()
    for payload in files:
        extension = Path(payload.filename).suffix.lower()
        file_type = extension.lstrip(".") if extension else "unknown"

        document = DocumentRecord(
            id=str(uuid.uuid4()),
            filename=payload.filename,
            file_type=file_type,
            size_bytes=len(payload.content),
        )
        bundle.documents.append(document)

        if extension not in SUPPORTED_EXTENSIONS:
            bundle.warnings.append(
                f"Unsupported file type for {payload.filename}: {extension or '<none>'}"
            )
            continue

        if extension in {".xlsx", ".xlsm"}:
            bundle.xlsx_files.append(payload)
            _ingest_xlsx(payload, document, bundle)
        elif extension in {".txt", ".md"}:
            text = payload.content.decode("utf-8", errors="replace")
            bundle.text_chunks.extend(_chunk_text(document.id, text))
        elif extension == ".pdf":
            text = _extract_pdf_text(payload.content)
            if text.strip():
                bundle.text_chunks.extend(_chunk_text(document.id, text))
            else:
                bundle.warnings.append(f"No extractable text found in PDF: {payload.filename}")
    return bundle


def _ingest_xlsx(payload: UploadedFilePayload, document: DocumentRecord, bundle: DocumentBundle) -> None:
    workbook = load_workbook(BytesIO(payload.content), data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            row_iter = sheet.iter_rows(values_only=True)
            try:
                raw_headers = next(row_iter)
            except StopIteration:
                continue
            headers = [str(item).strip() for item in raw_headers if item not in (None, "")]
            row_count = sum(1 for _ in row_iter)
            bundle.tables.append(
                TableSummary(
                    document_id=document.id,
                    sheet_name=sheet.title,
                    headers=headers,
                    row_count=row_count,
                )
            )
            for header in headers:
                code = normalize_metric_code(header)
                if code in DIMENSION_HEADERS:
                    continue
                bundle.detected_metrics.append(
                    DetectedMetric(
                        code=code,
                        label=header,
                        source_document_id=document.id,
                        source="xlsx_header",
                    )
                )
    finally:
        workbook.close()


def _extract_pdf_text(content: bytes) -> str:
    reader = PdfReader(BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _chunk_text(document_id: str, text: str, *, chunk_size: int = 4000) -> list[TextChunk]:
    clean_text = re.sub(r"\s+", " ", text).strip()
    if not clean_text:
        return []
    chunks: list[TextChunk] = []
    for index, start in enumerate(range(0, len(clean_text), chunk_size)):
        chunks.append(
            TextChunk(
                id=f"{document_id}:{index}",
                document_id=document_id,
                text=clean_text[start : start + chunk_size],
                index=index,
            )
        )
    return chunks


def write_xlsx_payloads_to_tempdir(payloads: list[UploadedFilePayload], temp_dir: str | Path) -> list[Path]:
    paths: list[Path] = []
    root = Path(temp_dir)
    root.mkdir(parents=True, exist_ok=True)
    for index, payload in enumerate(payloads):
        safe_name = re.sub(r"[^a-zA-Zа-яА-ЯёЁ0-9_.-]+", "_", payload.filename).strip("_") or f"upload_{index}.xlsx"
        output_path = root / f"{index}_{safe_name}"
        output_path.write_bytes(payload.content)
        paths.append(output_path)
    return paths


def make_tempdir() -> tempfile.TemporaryDirectory[str]:
    return tempfile.TemporaryDirectory(prefix="relation_memory_session_")
