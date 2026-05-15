from __future__ import annotations

import json
import math
import re
from dataclasses import asdict, dataclass, field
from itertools import combinations
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from app.services.relation_memory_poc_validator import RelationMemoryPocValidator

HEADER_ALIASES = {
    "месяц": "month",
    "регион": "region",
    "бизнес-направление": "business_unit",
    "склад": "warehouse",
    "товарная_категория": "product_category",
    "заказы": "orders",
    "объем": "volume",
    "выручка": "revenue",
    "средняя_цена": "avg_price",
    "доля_возвратов": "return_rate",
    "валовая_маржа": "gross_margin",
    "транспортные_затраты": "freight_cost",
    "затраты_на_хранение": "storage_cost",
    "совокупные_затраты": "total_cost",
    "доля_своевременной_доставки": "on_time_delivery_rate",
    "доля_дефицита": "stockout_rate",
    "дни_запаса": "inventory_days",
    "код_метрики": "metric_code",
    "название": "label",
    "описание": "description",
    "синонимы": "aliases",
    "уровень_чувствительности": "sensitivity_level",
    "разрешенные_роли": "allow_roles",
    "предпочтительный_датасет": "preferred_dataset",
    "код_исходной_метрики": "source_metric_code",
    "код_целевой_метрики": "target_metric_code",
    "тип_связи": "edge_type",
    "обоснование": "reason",
    "сила_связи": "strength",
    "период_сравнения": "lag_period",
    "важность": "severity",
    "текст_правила": "rule_text",
    "активно": "is_active",
    "код_правила": "code",
    "коды_метрик": "metric_codes",
    "текст_вердикта": "verdict_text",
    "приоритет": "priority",
}

METRIC_CODE_ALIASES = {
    "revenue": "revenue",
    "выручка": "revenue",
    "orders": "orders",
    "заказы": "orders",
    "volume": "volume",
    "объем": "volume",
    "avg price": "avg_price",
    "avg_price": "avg_price",
    "average price": "avg_price",
    "средняя цена": "avg_price",
    "средняя_цена": "avg_price",
    "return rate": "return_rate",
    "return_rate": "return_rate",
    "доля возвратов": "return_rate",
    "доля_возвратов": "return_rate",
    "gross margin": "gross_margin",
    "gross_margin": "gross_margin",
    "валовая маржа": "gross_margin",
    "валовая_маржа": "gross_margin",
    "freight cost": "freight_cost",
    "freight_cost": "freight_cost",
    "транспортные затраты": "freight_cost",
    "транспортные_затраты": "freight_cost",
    "storage cost": "storage_cost",
    "storage_cost": "storage_cost",
    "затраты на хранение": "storage_cost",
    "затраты_на_хранение": "storage_cost",
    "total cost": "total_cost",
    "total_cost": "total_cost",
    "совокупные затраты": "total_cost",
    "совокупные_затраты": "total_cost",
    "on time delivery rate": "on_time_delivery_rate",
    "on_time_delivery_rate": "on_time_delivery_rate",
    "доля своевременной доставки": "on_time_delivery_rate",
    "доля_своевременной_доставки": "on_time_delivery_rate",
    "stockout rate": "stockout_rate",
    "stockout_rate": "stockout_rate",
    "доля дефицита": "stockout_rate",
    "доля_дефицита": "stockout_rate",
    "inventory days": "inventory_days",
    "inventory_days": "inventory_days",
    "дни запаса": "inventory_days",
    "дни_запаса": "inventory_days",
}

DATASET_KEY_ALIASES = {
    "sales_monthly": "sales_monthly",
    "продажи по месяцам": "sales_monthly",
    "продажи_по_месяцам": "sales_monthly",
    "logistics_costs_monthly": "logistics_costs_monthly",
    "логистические затраты по месяцам": "logistics_costs_monthly",
    "логистические_затраты_по_месяцам": "logistics_costs_monthly",
    "warehouse_ops_monthly": "warehouse_ops_monthly",
    "складские операции по месяцам": "warehouse_ops_monthly",
    "складские_операции_по_месяцам": "warehouse_ops_monthly",
}

SENSITIVITY_ALIASES = {
    "public": "public",
    "публичный": "public",
    "internal": "internal",
    "внутренний": "internal",
    "confidential": "confidential",
    "конфиденциальный": "confidential",
}

ROLE_ALIASES = {
    "admin": "admin",
    "администратор": "admin",
    "analyst": "analyst",
    "аналитик": "analyst",
    "finance_manager": "finance_manager",
    "финансовый_менеджер": "finance_manager",
    "sales_manager": "sales_manager",
    "менеджер_по_продажам": "sales_manager",
    "ops_manager": "ops_manager",
    "операционный_менеджер": "ops_manager",
    "pricing_manager": "pricing_manager",
    "менеджер_по_ценообразованию": "pricing_manager",
}

EDGE_TYPE_ALIASES = {
    "driver": "driver",
    "драйвер": "driver",
    "inverse_driver": "inverse_driver",
    "обратный_драйвер": "inverse_driver",
    "component": "component",
    "компонент": "component",
}

LAG_PERIOD_ALIASES = {
    "previous_period": "previous_period",
    "предыдущий_период": "previous_period",
}

SEVERITY_ALIASES = {
    "info": "info",
    "информация": "info",
    "warning": "warning",
    "предупреждение": "warning",
}

VERDICT_CODE_ALIASES = {
    "ground_numeric_claims": "ground_numeric_claims",
    "обосновывать_числовые_утверждения": "ground_numeric_claims",
    "margin_drop_requires_dual_cause_check": "margin_drop_requires_dual_cause_check",
    "падение_маржи_требует_проверки_двух_причин": "margin_drop_requires_dual_cause_check",
    "service_issues_can_raise_returns": "service_issues_can_raise_returns",
    "сервисные_сбои_могут_увеличить_возвраты": "service_issues_can_raise_returns",
    "cost_rollup_must_be_consistent": "cost_rollup_must_be_consistent",
    "роллап_затрат_должен_быть_согласован": "cost_rollup_must_be_consistent",
}

DIMENSION_VALUE_ALIASES = {
    "region": {
        "ru": "ru",
        "россия": "ru",
        "eu": "eu",
        "европа": "eu",
        "apac": "apac",
        "атр": "apac",
    },
    "business_unit": {
        "retail": "retail",
        "розница": "retail",
        "b2b": "b2b",
        "ecommerce": "ecommerce",
        "электронная коммерция": "ecommerce",
    },
    "warehouse": {
        "moscow_dc": "moscow_dc",
        "рц москва": "moscow_dc",
        "berlin_dc": "berlin_dc",
        "рц берлин": "berlin_dc",
        "singapore_dc": "singapore_dc",
        "рц сингапур": "singapore_dc",
    },
    "product_category": {
        "electronics": "electronics",
        "электроника": "electronics",
        "furniture": "furniture",
        "мебель": "furniture",
        "apparel": "apparel",
        "одежда": "apparel",
    },
}

ENTITY_DISPLAY_LABELS = {
    "region": {
        "ru": "Россия",
        "eu": "Европа",
        "apac": "АТР",
    },
    "business_unit": {
        "retail": "Розница",
        "b2b": "B2B",
        "ecommerce": "Электронная коммерция",
    },
    "warehouse": {
        "moscow_dc": "РЦ Москва",
        "berlin_dc": "РЦ Берлин",
        "singapore_dc": "РЦ Сингапур",
    },
    "product_category": {
        "electronics": "Электроника",
        "furniture": "Мебель",
        "apparel": "Одежда",
    },
}

DATASET_DISPLAY_LABELS = {
    "sales_monthly": "Продажи по месяцам",
    "logistics_costs_monthly": "Логистические затраты по месяцам",
    "warehouse_ops_monthly": "Складские операции по месяцам",
    "government_context_monthly": "Государственный контекст по месяцам",
    "weather_context_monthly": "Погодный контекст по месяцам",
    "ftl_market_context_monthly": "Рыночный контекст FTL по месяцам",
    "ftl_cost_external_context": "Внешний контекст себестоимости FTL",
    "ftl_cost_report": "Себестоимость FTL февраль 2026",
    "atp_analytics_report": "Аналитическая записка АТП февраль 2026",
}

DIMENSION_DISPLAY_LABELS = {
    "month": "Месяц",
    "region": "Регион",
    "business_unit": "Бизнес-направление",
    "warehouse": "Склад",
    "product_category": "Товарная категория",
}


def _project_root() -> Path:
    return Path(__file__).resolve().parents[3]


def _normalize_text(value: str) -> str:
    lowered = value.lower().replace("_", " ")
    lowered = re.sub(r"[^a-zа-яё0-9 ]+", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


HEADER_ALIASES = {_normalize_text(key): value for key, value in HEADER_ALIASES.items()}
METRIC_CODE_ALIASES = {_normalize_text(key): value for key, value in METRIC_CODE_ALIASES.items()}
DATASET_KEY_ALIASES = {_normalize_text(key): value for key, value in DATASET_KEY_ALIASES.items()}
SENSITIVITY_ALIASES = {_normalize_text(key): value for key, value in SENSITIVITY_ALIASES.items()}
ROLE_ALIASES = {_normalize_text(key): value for key, value in ROLE_ALIASES.items()}
EDGE_TYPE_ALIASES = {_normalize_text(key): value for key, value in EDGE_TYPE_ALIASES.items()}
LAG_PERIOD_ALIASES = {_normalize_text(key): value for key, value in LAG_PERIOD_ALIASES.items()}
SEVERITY_ALIASES = {_normalize_text(key): value for key, value in SEVERITY_ALIASES.items()}
VERDICT_CODE_ALIASES = {_normalize_text(key): value for key, value in VERDICT_CODE_ALIASES.items()}
DIMENSION_VALUE_ALIASES = {
    dimension_key: {_normalize_text(key): value for key, value in alias_map.items()}
    for dimension_key, alias_map in DIMENSION_VALUE_ALIASES.items()
}


def _normalize_code(value: str) -> str:
    return re.sub(r"[^a-zа-яё0-9]+", "_", value.lower(), flags=re.IGNORECASE).strip("_")


def _lookup_alias(value: Any, mapping: dict[str, str]) -> str:
    normalized = _normalize_text(str(value or ""))
    return mapping.get(normalized, str(value or "").strip())


def _canonical_header(value: str) -> str:
    return HEADER_ALIASES.get(_normalize_text(value), str(value).strip())


def _canonical_metric_code(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    alias = METRIC_CODE_ALIASES.get(_normalize_text(text))
    return alias or _normalize_code(text)


def _canonical_dataset_key(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    alias = DATASET_KEY_ALIASES.get(_normalize_text(text))
    return alias or _normalize_code(text)


def _canonical_dimension_value(dimension_key: str, value: Any) -> str:
    if value in (None, ""):
        return ""
    mapping = DIMENSION_VALUE_ALIASES.get(dimension_key)
    if not mapping:
        return str(value).strip()
    return _lookup_alias(value, mapping)


def _canonical_sensitivity(value: Any) -> str:
    return _lookup_alias(value, SENSITIVITY_ALIASES)


def _canonical_role(value: Any) -> str:
    return _lookup_alias(value, ROLE_ALIASES)


def _canonical_edge_type(value: Any) -> str:
    return _lookup_alias(value, EDGE_TYPE_ALIASES)


def _canonical_lag_period(value: Any) -> str:
    return _lookup_alias(value, LAG_PERIOD_ALIASES)


def _canonical_severity(value: Any) -> str:
    return _lookup_alias(value, SEVERITY_ALIASES)


def _canonical_verdict_code(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    alias = VERDICT_CODE_ALIASES.get(_normalize_text(text))
    return alias or _normalize_code(text)


def _titleize(value: str) -> str:
    return value.replace("_", " ").strip().title()


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _split_multi(value: Any, separator: str = "|") -> list[str]:
    return [part.strip() for part in str(value or "").split(separator) if part.strip()]


def _sheet_rows(
    path: Path, sheet_name: str | None = None
) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"{path.name} missing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]
    row_iter = sheet.iter_rows(values_only=True)
    headers = [str(item) for item in next(row_iter)]
    rows = [dict(zip(headers, values)) for values in row_iter]
    workbook.close()
    return headers, rows


def _safe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _pearson_correlation(values_x: list[float], values_y: list[float]) -> float | None:
    if len(values_x) != len(values_y) or len(values_x) < 3:
        return None
    mean_x = sum(values_x) / len(values_x)
    mean_y = sum(values_y) / len(values_y)
    centered_x = [value - mean_x for value in values_x]
    centered_y = [value - mean_y for value in values_y]
    numerator = sum(left * right for left, right in zip(centered_x, centered_y, strict=True))
    denominator = math.sqrt(
        sum(value * value for value in centered_x) * sum(value * value for value in centered_y)
    )
    if denominator == 0:
        return None
    return numerator / denominator


def _normalized_mae(predicted: list[float], actual: list[float]) -> float | None:
    if len(predicted) != len(actual) or not actual:
        return None
    scale = max(1.0, sum(abs(value) for value in actual) / len(actual))
    return (
        sum(abs(left - right) for left, right in zip(predicted, actual, strict=True))
        / len(actual)
        / scale
    )


def _aggregation_mode_for_metric(metric_code: str) -> str:
    average_tokens = ("rate", "avg", "margin", "index", "days", "temperature")
    if any(token in metric_code for token in average_tokens):
        return "mean"
    return "sum"


def _aggregate_values(values: list[float], mode: str) -> float:
    if not values:
        return 0.0
    if mode == "mean":
        return sum(values) / len(values)
    return sum(values)


def _edge_sign(edge_type: str) -> int:
    return -1 if edge_type == "inverse_driver" else 1


def _round_number(value: float | None, digits: int = 4) -> float | None:
    if value is None:
        return None
    return round(value, digits)


def _default_metric_record(code: str, dataset_key: str) -> dict[str, Any]:
    label = _titleize(code)
    return {
        "code": code,
        "label": label,
        "display_name": label,
        "name": label,
        "description": f"Inferred metric '{code}' from dataset '{dataset_key}'.",
        "aliases": [code, code.replace("_", " "), _titleize(code).lower()],
        "sensitivity_level": "internal",
        "allow_roles": ["admin", "analyst"],
        "preferred_dataset": dataset_key,
        "semantic_type": "unknown",
        "source": "inferred",
    }


@dataclass
class RelationMemorySnapshot:
    datasets: list[dict[str, Any]] = field(default_factory=list)
    metrics: list[dict[str, Any]] = field(default_factory=list)
    formulas: list[dict[str, Any]] = field(default_factory=list)
    dimensions: list[dict[str, Any]] = field(default_factory=list)
    entity_values: list[dict[str, Any]] = field(default_factory=list)
    dependencies: list[dict[str, Any]] = field(default_factory=list)
    lag_rules: list[dict[str, Any]] = field(default_factory=list)
    verdict_rules: list[dict[str, Any]] = field(default_factory=list)
    observations: list[dict[str, Any]] = field(default_factory=list)
    hypotheses: list[dict[str, Any]] = field(default_factory=list)
    inquiry_questions: list[dict[str, Any]] = field(default_factory=list)
    golden_queries: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class RelationMemoryPocBuilder:
    def __init__(
        self, manifest_path: str | Path, dependency_priors: list[dict[str, Any]] | None = None
    ):
        self.manifest_path = Path(manifest_path)
        self.project_root = _project_root()
        self.dependency_priors = list(dependency_priors or [])
        with self.manifest_path.open("r", encoding="utf-8") as handle:
            self.manifest = yaml.safe_load(handle)

    def build_snapshot(self) -> RelationMemorySnapshot:
        snapshot = RelationMemorySnapshot()
        metric_records: dict[str, dict[str, Any]] = {}
        dataset_records: dict[str, dict[str, Any]] = {}
        formula_records: dict[str, dict[str, Any]] = {}
        dimension_records: dict[str, dict[str, Any]] = {}
        entity_records: dict[str, dict[str, Any]] = {}
        dependency_records: dict[str, dict[str, Any]] = {}
        lag_records: dict[str, dict[str, Any]] = {}
        verdict_records: dict[str, dict[str, Any]] = {}
        fact_contexts: list[dict[str, Any]] = []

        for contract in self.manifest.get("incoming_csvs", []):
            path = self.project_root / str(contract["path"])
            required = bool(contract.get("required"))
            if not path.exists():
                if required:
                    raise FileNotFoundError(f"Required PoC input missing: {path}")
                snapshot.warnings.append(f"Optional input missing: {path}")
                continue

            headers, rows = _sheet_rows(path, contract.get("sheet_name") or contract.get("sheet"))
            headers, rows = self._canonicalize_rows(contract=contract, headers=headers, rows=rows)
            self._validate_contract(path=path, headers=headers, contract=contract)
            if contract.get("kind") == "numeric_facts":
                dataset_key = str(contract["key"])
                fact_contexts.append(
                    {
                        "dataset_key": dataset_key,
                        "dimensions": list(contract.get("grain", [])),
                        "metrics": list(contract.get("derives", {}).get("metrics", [])),
                        "rows": rows,
                    }
                )
                self._ingest_fact_dataset(
                    contract=contract,
                    path=path,
                    rows=rows,
                    dataset_records=dataset_records,
                    metric_records=metric_records,
                    formula_records=formula_records,
                    dimension_records=dimension_records,
                    entity_records=entity_records,
                )
            else:
                self._ingest_metadata(
                    contract=contract,
                    rows=rows,
                    metric_records=metric_records,
                    dependency_records=dependency_records,
                    lag_records=lag_records,
                    verdict_records=verdict_records,
                )

        self._infer_dependencies_from_facts(
            fact_contexts=fact_contexts,
            metric_records=metric_records,
            dependency_records=dependency_records,
        )
        self._apply_dependency_heuristics(metric_records, dependency_records)
        self._apply_dependency_priors(metric_records, dependency_records)
        self._apply_rule_fallbacks(metric_records, lag_records, verdict_records)
        self._enrich_display_metadata(
            dataset_records=dataset_records,
            metric_records=metric_records,
            formula_records=formula_records,
            dimension_records=dimension_records,
            entity_records=entity_records,
            lag_records=lag_records,
            verdict_records=verdict_records,
        )
        inquiry_bundle = self._build_inquiry_bundle(
            fact_contexts=fact_contexts,
            metric_records=metric_records,
            dependency_records=dependency_records,
            lag_records=lag_records,
        )

        snapshot.datasets = sorted(dataset_records.values(), key=lambda item: item["key"])
        snapshot.metrics = sorted(metric_records.values(), key=lambda item: item["code"])
        snapshot.formulas = sorted(formula_records.values(), key=lambda item: item["id"])
        snapshot.dimensions = sorted(dimension_records.values(), key=lambda item: item["key"])
        snapshot.entity_values = sorted(entity_records.values(), key=lambda item: item["id"])
        snapshot.dependencies = sorted(
            dependency_records.values(),
            key=lambda item: (item["source_metric_code"], item["target_metric_code"]),
        )
        snapshot.lag_rules = sorted(lag_records.values(), key=lambda item: item["id"])
        snapshot.verdict_rules = sorted(verdict_records.values(), key=lambda item: item["code"])
        snapshot.observations = inquiry_bundle["observations"]
        snapshot.hypotheses = inquiry_bundle["hypotheses"]
        snapshot.inquiry_questions = inquiry_bundle["inquiry_questions"]
        snapshot.golden_queries = list(self.manifest.get("golden_queries", []))
        return snapshot

    def _build_inquiry_bundle(
        self,
        *,
        fact_contexts: list[dict[str, Any]],
        metric_records: dict[str, dict[str, Any]],
        dependency_records: dict[str, dict[str, Any]],
        lag_records: dict[str, dict[str, Any]],
    ) -> dict[str, list[dict[str, Any]]]:
        metric_series = self._build_metric_time_series(fact_contexts)
        observations = self._infer_observations(metric_series, metric_records, dependency_records)
        hypotheses = self._infer_hypotheses(
            observations=observations,
            metric_series=metric_series,
            metric_records=metric_records,
            dependency_records=dependency_records,
            lag_records=lag_records,
        )
        inquiry_questions = self._build_inquiry_questions(
            observations=observations,
            hypotheses=hypotheses,
            metric_records=metric_records,
        )
        return {
            "observations": observations,
            "hypotheses": hypotheses,
            "inquiry_questions": inquiry_questions,
        }

    def _build_metric_time_series(
        self, fact_contexts: list[dict[str, Any]]
    ) -> dict[str, dict[str, Any]]:
        metric_series: dict[str, dict[str, Any]] = {}
        for context in fact_contexts:
            if "month" not in context.get("dimensions", []):
                continue
            monthly_values: dict[str, dict[str, list[float]]] = {}
            for row in context.get("rows", []):
                month = str(row.get("month") or "").strip()
                if not month:
                    continue
                for metric_code in context.get("metrics", []):
                    numeric_value = _safe_float(row.get(metric_code))
                    if numeric_value is None:
                        continue
                    monthly_values.setdefault(metric_code, {}).setdefault(month, []).append(
                        numeric_value
                    )
            for metric_code, month_map in monthly_values.items():
                aggregation = _aggregation_mode_for_metric(metric_code)
                points = [
                    {
                        "month": month,
                        "value": _round_number(_aggregate_values(values, aggregation), 6),
                        "sample_size": len(values),
                    }
                    for month, values in sorted(month_map.items())
                ]
                if len(points) < 2:
                    continue
                metric_series[metric_code] = {
                    "metric_code": metric_code,
                    "dataset_key": context["dataset_key"],
                    "aggregation": aggregation,
                    "points": points,
                }
        return metric_series

    def _infer_observations(
        self,
        metric_series: dict[str, dict[str, Any]],
        metric_records: dict[str, dict[str, Any]],
        dependency_records: dict[str, dict[str, Any]],
    ) -> list[dict[str, Any]]:
        observations: list[dict[str, Any]] = []
        inbound_counts: dict[str, int] = {}
        outbound_counts: dict[str, int] = {}
        for dependency in dependency_records.values():
            inbound_counts[dependency["target_metric_code"]] = (
                inbound_counts.get(dependency["target_metric_code"], 0) + 1
            )
            outbound_counts[dependency["source_metric_code"]] = (
                outbound_counts.get(dependency["source_metric_code"], 0) + 1
            )
        for metric_code, series in metric_series.items():
            if self._is_context_metric(metric_code, metric_records):
                continue
            semantic_type = self._metric_semantic_type(metric_code, metric_records)
            if semantic_type == "technical_check":
                continue
            points = series["points"]
            current_point = points[-1]
            previous_point = points[-2]
            delta_abs = float(current_point["value"]) - float(previous_point["value"])
            midpoint_scale = max(
                (abs(float(current_point["value"])) + abs(float(previous_point["value"]))) / 2.0,
                1e-9,
            )
            delta_pct = None
            if abs(float(previous_point["value"])) > 1e-9:
                delta_pct = delta_abs / abs(float(previous_point["value"]))
            raw_score = min(10.0, abs(delta_abs) / midpoint_scale)
            direction_sign = 1 if delta_abs > 0 else -1 if delta_abs < 0 else 0
            observations.append(
                {
                    "id": f"{metric_code}:{previous_point['month']}->{current_point['month']}",
                    "metric_code": metric_code,
                    "dataset_key": series["dataset_key"],
                    "kind": "level_shift",
                    "reference_mode": "up"
                    if direction_sign > 0
                    else "down"
                    if direction_sign < 0
                    else "flat",
                    "direction_sign": direction_sign,
                    "aggregation": series["aggregation"],
                    "current_period": current_point["month"],
                    "previous_period": previous_point["month"],
                    "current_value": _round_number(float(current_point["value"]), 4),
                    "previous_value": _round_number(float(previous_point["value"]), 4),
                    "delta_abs": _round_number(delta_abs, 4),
                    "delta_pct": _round_number(delta_pct, 4),
                    "score": _round_number(raw_score, 4),
                    "raw_score": _round_number(raw_score, 4),
                    "semantic_type": semantic_type,
                    "inbound_dependency_count": inbound_counts.get(metric_code, 0),
                    "outbound_dependency_count": outbound_counts.get(metric_code, 0),
                    "recent_points": points[-3:],
                }
            )
        total_abs_delta = (
            sum(abs(float(item.get("delta_abs") or 0.0)) for item in observations) or 1.0
        )
        for observation in observations:
            semantic_type = str(observation.get("semantic_type") or "unknown")
            semantic_boost = {
                "target_metric": 0.6,
                "cost_component": 0.35,
                "business_driver": 0.2,
                "denominator": 0.1,
            }.get(semantic_type, 0.0)
            graph_boost = min(
                0.5,
                0.05
                * (
                    int(observation.get("inbound_dependency_count") or 0)
                    + int(observation.get("outbound_dependency_count") or 0)
                ),
            )
            materiality_score = min(
                1.0, abs(float(observation.get("delta_abs") or 0.0)) / total_abs_delta * 10
            )
            near_zero_penalty = (
                0.75
                if abs(float(observation.get("previous_value") or 0.0)) < 1e-6
                and abs(float(observation.get("delta_abs") or 0.0)) < 1.0
                else 0.0
            )
            observation["score"] = _round_number(
                max(
                    0.0,
                    float(observation.get("raw_score") or 0.0)
                    + materiality_score
                    + semantic_boost
                    + graph_boost
                    - near_zero_penalty,
                ),
                4,
            )
        observations.sort(key=lambda item: (-float(item["score"]), item["metric_code"]))
        return observations[:20]

    def _infer_hypotheses(
        self,
        *,
        observations: list[dict[str, Any]],
        metric_series: dict[str, dict[str, Any]],
        metric_records: dict[str, dict[str, Any]],
        dependency_records: dict[str, dict[str, Any]],
        lag_records: dict[str, dict[str, Any]],
    ) -> list[dict[str, Any]]:
        reverse_edges: dict[str, list[dict[str, Any]]] = {}
        forward_edges: dict[str, list[dict[str, Any]]] = {}
        for dependency in dependency_records.values():
            reverse_edges.setdefault(dependency["target_metric_code"], []).append(dependency)
            forward_edges.setdefault(dependency["source_metric_code"], []).append(dependency)

        active_lag_metrics = {
            record["metric_code"] for record in lag_records.values() if record.get("is_active")
        }

        hypotheses: list[dict[str, Any]] = []
        for observation in observations:
            target_metric = observation["metric_code"]
            target_direction_sign = int(observation["direction_sign"])
            if target_direction_sign == 0:
                continue
            current_period = observation["current_period"]
            previous_period = observation["previous_period"]
            recent_points = observation.get("recent_points", [])
            two_periods_back = str(recent_points[0]["month"]) if len(recent_points) >= 3 else None
            ranked_candidates: list[dict[str, Any]] = []

            for path in self._find_upstream_paths(target_metric, reverse_edges):
                source_metric = path["metric_codes"][0]
                source_series = metric_series.get(source_metric)
                if not source_series:
                    continue
                source_points = {
                    item["month"]: float(item["value"]) for item in source_series["points"]
                }
                current_source_value = source_points.get(current_period)
                previous_source_value = source_points.get(previous_period)
                older_source_value = (
                    source_points.get(two_periods_back) if two_periods_back else None
                )

                current_source_sign = 0
                lagged_source_sign = 0
                if current_source_value is not None and previous_source_value is not None:
                    current_source_delta = current_source_value - previous_source_value
                    current_source_sign = (
                        1 if current_source_delta > 0 else -1 if current_source_delta < 0 else 0
                    )
                if previous_source_value is not None and older_source_value is not None:
                    lagged_source_delta = previous_source_value - older_source_value
                    lagged_source_sign = (
                        1 if lagged_source_delta > 0 else -1 if lagged_source_delta < 0 else 0
                    )

                expected_source_sign = target_direction_sign * int(path["path_sign"])
                current_support = (
                    current_source_sign == expected_source_sign and current_source_sign != 0
                )
                lag_support = (
                    lagged_source_sign == expected_source_sign
                    and lagged_source_sign != 0
                    and not current_support
                )
                lag_is_modeled = (
                    source_metric in active_lag_metrics or target_metric in active_lag_metrics
                )

                raw_score = float(path["total_strength"]) / max(int(path["hops"]), 1)
                if current_support:
                    raw_score += 1.1
                elif lag_support:
                    raw_score += 0.9
                    if lag_is_modeled:
                        raw_score += 0.15
                if int(path["hops"]) > 1:
                    raw_score += 0.35
                if self._is_context_metric(source_metric, metric_records):
                    raw_score += 0.15
                confidence = round(min(0.99, raw_score / 3.0), 3)

                if lag_support:
                    mechanism_type = "lagged_driver"
                    support_window = "previous_period"
                elif int(path["hops"]) > 1 and current_support:
                    mechanism_type = "multi_hop_driver"
                    support_window = "current"
                elif current_support and self._is_context_metric(source_metric, metric_records):
                    mechanism_type = "context_driver"
                    support_window = "current"
                elif current_support:
                    mechanism_type = "direct_driver"
                    support_window = "current"
                else:
                    mechanism_type = "unconfirmed_driver"
                    support_window = "unconfirmed"

                target_label = self._metric_label(target_metric, metric_records)
                source_label = self._metric_label(source_metric, metric_records)
                path_text = self._path_to_text(path["metric_codes"], metric_records)
                if mechanism_type == "lagged_driver":
                    explanation = (
                        f"Изменение «{target_label}» в {current_period} больше похоже на эффект «{source_label}» "
                        f"с лагом в один период по цепочке {path_text}."
                    )
                elif mechanism_type in {"multi_hop_driver", "context_driver", "direct_driver"}:
                    explanation = (
                        f"Изменение «{target_label}» в {current_period} согласуется с драйвером «{source_label}» "
                        f"по цепочке {path_text}."
                    )
                else:
                    explanation = (
                        f"Граф допускает влияние «{source_label}» на «{target_label}» по цепочке {path_text}, "
                        "но временной ряд это пока явно не подтверждает."
                    )

                ranked_candidates.append(
                    {
                        "id": f"{observation['id']}:{'->'.join(path['metric_codes'])}",
                        "observation_id": observation["id"],
                        "target_metric_code": target_metric,
                        "source_metric_code": source_metric,
                        "metric_codes": path["metric_codes"],
                        "path_edge_types": path["path_edge_types"],
                        "hops": path["hops"],
                        "mechanism_type": mechanism_type,
                        "support_window": support_window,
                        "confidence": confidence,
                        "explanation": explanation,
                    }
                )

            for path in self._find_downstream_paths(target_metric, forward_edges):
                impact_metric = path["metric_codes"][-1]
                impact_series = metric_series.get(impact_metric)
                if not impact_series:
                    continue
                impact_points = {
                    item["month"]: float(item["value"]) for item in impact_series["points"]
                }
                current_impact_value = impact_points.get(current_period)
                previous_impact_value = impact_points.get(previous_period)
                if current_impact_value is None or previous_impact_value is None:
                    continue
                impact_delta = current_impact_value - previous_impact_value
                impact_sign = 1 if impact_delta > 0 else -1 if impact_delta < 0 else 0
                expected_impact_sign = target_direction_sign * int(path["path_sign"])
                current_support = impact_sign == expected_impact_sign and impact_sign != 0

                raw_score = float(path["total_strength"]) / max(int(path["hops"]), 1)
                if current_support:
                    raw_score += 0.75
                if int(path["hops"]) > 1:
                    raw_score += 0.2
                confidence = round(min(0.88, raw_score / 2.5), 3)
                if not current_support and confidence < 0.35:
                    continue

                observed_label = self._metric_label(target_metric, metric_records)
                impact_label = self._metric_label(impact_metric, metric_records)
                path_text = self._path_to_text(path["metric_codes"], metric_records)
                mechanism_type = (
                    "downstream_impact" if current_support else "possible_downstream_impact"
                )
                if current_support:
                    explanation = (
                        f"Изменение «{observed_label}» в {current_period} совпадает по знаку с downstream-метрикой "
                        f"«{impact_label}» по цепочке {path_text}."
                    )
                else:
                    explanation = (
                        f"Граф показывает, что «{observed_label}» входит в downstream-цепочку до «{impact_label}» "
                        f"по пути {path_text}, но период пока не дает сильного численного подтверждения."
                    )

                ranked_candidates.append(
                    {
                        "id": f"{observation['id']}:downstream:{'->'.join(path['metric_codes'])}",
                        "observation_id": observation["id"],
                        "target_metric_code": target_metric,
                        "source_metric_code": target_metric,
                        "impact_metric_code": impact_metric,
                        "metric_codes": path["metric_codes"],
                        "path_edge_types": path["path_edge_types"],
                        "hops": path["hops"],
                        "mechanism_type": mechanism_type,
                        "support_window": "current" if current_support else "graph_only",
                        "confidence": confidence,
                        "explanation": explanation,
                    }
                )

            ranked_candidates.sort(
                key=lambda item: (
                    -float(item["confidence"]),
                    item["hops"],
                    str(item.get("source_metric_code") or ""),
                ),
            )
            selected = ranked_candidates[:3]
            if not selected or float(selected[0]["confidence"]) < 0.25:
                target_label = self._metric_label(target_metric, metric_records)
                selected.append(
                    {
                        "id": f"{observation['id']}:boundary_gap",
                        "observation_id": observation["id"],
                        "target_metric_code": target_metric,
                        "source_metric_code": None,
                        "metric_codes": [target_metric],
                        "path_edge_types": [],
                        "hops": 0,
                        "mechanism_type": "boundary_gap",
                        "support_window": "outside_graph",
                        "confidence": 0.2,
                        "explanation": (
                            f"Для изменения «{target_label}» в {current_period} пока нет подтвержденной связи "
                            "в памяти или правилах. Нужно уточнить у пользователя, какой бизнес-драйвер мог "
                            "повлиять на эту метрику."
                        ),
                    }
                )
            hypotheses.extend(selected)

        hypotheses.sort(
            key=lambda item: (
                item["observation_id"],
                -float(item["confidence"]),
                item["hops"],
                str(item.get("source_metric_code") or ""),
            ),
        )
        return hypotheses

    def _find_upstream_paths(
        self,
        target_metric: str,
        reverse_edges: dict[str, list[dict[str, Any]]],
        *,
        max_hops: int = 3,
    ) -> list[dict[str, Any]]:
        queue: list[tuple[str, list[str], list[str], float, int]] = [
            (target_metric, [target_metric], [], 0.0, 1)
        ]
        paths: list[dict[str, Any]] = []
        while queue:
            current_metric, metric_codes, edge_types, total_strength, path_sign = queue.pop(0)
            if len(metric_codes) > max_hops + 1:
                continue
            for dependency in reverse_edges.get(current_metric, []):
                source_metric = dependency["source_metric_code"]
                if source_metric in metric_codes:
                    continue
                next_metric_codes = [source_metric, *metric_codes]
                next_edge_types = [dependency["edge_type"], *edge_types]
                next_total_strength = total_strength + float(dependency.get("strength") or 0.0)
                next_path_sign = path_sign * _edge_sign(
                    str(dependency.get("edge_type") or "driver")
                )
                paths.append(
                    {
                        "metric_codes": next_metric_codes,
                        "path_edge_types": next_edge_types,
                        "hops": len(next_metric_codes) - 1,
                        "total_strength": next_total_strength,
                        "path_sign": next_path_sign,
                    }
                )
                if len(next_metric_codes) - 1 < max_hops:
                    queue.append(
                        (
                            source_metric,
                            next_metric_codes,
                            next_edge_types,
                            next_total_strength,
                            next_path_sign,
                        )
                    )
        paths.sort(
            key=lambda item: (item["hops"], -float(item["total_strength"]), item["metric_codes"][0])
        )
        return paths[:20]

    def _find_downstream_paths(
        self,
        source_metric: str,
        forward_edges: dict[str, list[dict[str, Any]]],
        *,
        max_hops: int = 3,
    ) -> list[dict[str, Any]]:
        queue: list[tuple[str, list[str], list[str], float, int]] = [
            (source_metric, [source_metric], [], 0.0, 1)
        ]
        paths: list[dict[str, Any]] = []
        while queue:
            current_metric, metric_codes, edge_types, total_strength, path_sign = queue.pop(0)
            if len(metric_codes) > max_hops + 1:
                continue
            for dependency in forward_edges.get(current_metric, []):
                next_metric = dependency["target_metric_code"]
                if next_metric in metric_codes:
                    continue
                next_metric_codes = [*metric_codes, next_metric]
                next_edge_types = [*edge_types, dependency["edge_type"]]
                next_total_strength = total_strength + float(dependency.get("strength") or 0.0)
                next_path_sign = path_sign * _edge_sign(
                    str(dependency.get("edge_type") or "driver")
                )
                paths.append(
                    {
                        "metric_codes": next_metric_codes,
                        "path_edge_types": next_edge_types,
                        "hops": len(next_metric_codes) - 1,
                        "total_strength": next_total_strength,
                        "path_sign": next_path_sign,
                    }
                )
                if len(next_metric_codes) - 1 < max_hops:
                    queue.append(
                        (
                            next_metric,
                            next_metric_codes,
                            next_edge_types,
                            next_total_strength,
                            next_path_sign,
                        )
                    )
        paths.sort(
            key=lambda item: (
                item["hops"],
                -float(item["total_strength"]),
                item["metric_codes"][-1],
            )
        )
        return paths[:20]

    def _build_inquiry_questions(
        self,
        *,
        observations: list[dict[str, Any]],
        hypotheses: list[dict[str, Any]],
        metric_records: dict[str, dict[str, Any]],
    ) -> list[dict[str, Any]]:
        hypotheses_by_observation: dict[str, list[dict[str, Any]]] = {}
        for hypothesis in hypotheses:
            hypotheses_by_observation.setdefault(hypothesis["observation_id"], []).append(
                hypothesis
            )

        inquiry_questions: list[dict[str, Any]] = []
        for observation in observations[:10]:
            ranked = hypotheses_by_observation.get(observation["id"], [])
            target_label = self._metric_label(observation["metric_code"], metric_records)
            current_period = observation["current_period"]
            ranked = sorted(ranked, key=lambda item: (-float(item["confidence"]), item["hops"]))
            valid_ranked = [
                hypothesis
                for hypothesis in ranked
                if self._is_valid_inquiry_hypothesis(
                    hypothesis, observation["metric_code"], metric_records
                )
            ]
            if not valid_ranked:
                inquiry_questions.append(self._boundary_probe_question(observation, target_label))
                continue

            top = valid_ranked[0]
            runner_up = next(
                (
                    hypothesis
                    for hypothesis in valid_ranked[1:]
                    if str(hypothesis.get("source_metric_code") or "")
                    != str(top.get("source_metric_code") or "")
                ),
                None,
            )
            hypothesis_ids = [top["id"]]

            if runner_up and float(runner_up["confidence"]) >= float(top["confidence"]) * 0.75:
                question_type = "disambiguate_drivers"
                top_source = self._metric_label(str(top["source_metric_code"]), metric_records)
                runner_source = self._metric_label(
                    str(runner_up["source_metric_code"]), metric_records
                )
                top_path = self._path_to_text(top["metric_codes"], metric_records)
                runner_path = self._path_to_text(runner_up["metric_codes"], metric_records)
                hypothesis_ids.append(runner_up["id"])
                prompt = (
                    f"Изменение «{target_label}» в {current_period} больше связано с «{top_source}» "
                    f"по цепочке {top_path} или с «{runner_source}» по цепочке {runner_path}? "
                    "Что из этого подтверждается по бизнес-контексту лучше?"
                )
            elif top["mechanism_type"] == "lagged_driver":
                question_type = "lag_probe"
                source_label = self._metric_label(str(top["source_metric_code"]), metric_records)
                prompt = (
                    f"Эффект «{source_label}» на «{target_label}» у вас обычно проявляется в том же месяце "
                    "или с лагом в один период?"
                )
            elif len(top.get("metric_codes", [])) > 2:
                question_type = "chain_probe"
                source_label = self._metric_label(str(top["source_metric_code"]), metric_records)
                path_text = self._path_to_text(top["metric_codes"], metric_records)
                prompt = (
                    f"Корректно ли считать, что «{source_label}» влияет на «{target_label}» через цепочку {path_text}, "
                    "а не через прямой локальный фактор?"
                )
            else:
                question_type = "driver_probe"
                source_label = self._metric_label(str(top["source_metric_code"]), metric_records)
                prompt = (
                    f"Подтверждаете ли вы, что изменение «{target_label}» в {current_period} связано прежде всего "
                    f"с драйвером «{source_label}»?"
                )

            priority = round(float(observation["score"]) + float(top["confidence"]), 3)
            inquiry_questions.append(
                {
                    "id": f"{observation['id']}:question",
                    "observation_id": observation["id"],
                    "target_metric_code": observation["metric_code"],
                    "question_type": question_type,
                    "prompt": prompt,
                    "hypothesis_ids": hypothesis_ids,
                    "priority": priority,
                }
            )

        inquiry_questions.sort(key=lambda item: (-float(item["priority"]), item["id"]))
        return inquiry_questions[:10]

    def _metric_label(self, metric_code: str, metric_records: dict[str, dict[str, Any]]) -> str:
        record = metric_records.get(metric_code)
        if not record:
            return _titleize(metric_code)
        return str(record.get("label") or record.get("display_name") or metric_code)

    def _path_to_text(
        self, metric_codes: list[str], metric_records: dict[str, dict[str, Any]]
    ) -> str:
        return " -> ".join(self._metric_label(code, metric_records) for code in metric_codes)

    def _is_context_metric(
        self, metric_code: str, metric_records: dict[str, dict[str, Any]]
    ) -> bool:
        preferred_dataset = str(metric_records.get(metric_code, {}).get("preferred_dataset") or "")
        return "context" in preferred_dataset

    def _metric_semantic_type(
        self, metric_code: str, metric_records: dict[str, dict[str, Any]]
    ) -> str:
        return str(metric_records.get(metric_code, {}).get("semantic_type") or "unknown")

    def _is_valid_inquiry_hypothesis(
        self,
        hypothesis: dict[str, Any],
        target_metric_code: str,
        metric_records: dict[str, dict[str, Any]],
    ) -> bool:
        if str(hypothesis.get("mechanism_type") or "") == "boundary_gap":
            return False
        source_metric_code = str(hypothesis.get("source_metric_code") or "")
        if not source_metric_code or source_metric_code == target_metric_code:
            return False
        if self._metric_semantic_type(source_metric_code, metric_records) == "technical_check":
            return False
        return self._has_distinct_upstream_driver(hypothesis, target_metric_code)

    def _has_distinct_upstream_driver(
        self, hypothesis: dict[str, Any], target_metric_code: str
    ) -> bool:
        metric_codes = [
            str(code or "") for code in hypothesis.get("metric_codes") or [] if str(code or "")
        ]
        if len(metric_codes) < 2:
            return False
        if metric_codes[-1] != target_metric_code:
            return False
        return any(code != target_metric_code for code in metric_codes[:-1])

    def _boundary_probe_question(
        self, observation: dict[str, Any], target_label: str
    ) -> dict[str, Any]:
        return {
            "id": f"{observation['id']}:question",
            "observation_id": observation["id"],
            "target_metric_code": observation["metric_code"],
            "question_type": "boundary_probe",
            "prompt": (
                f"Что в бизнес-процессе могло повлиять на изменение «{target_label}» в {observation['current_period']}? "
                "Если это связано с другой метрикой из отчета, укажите с какой."
            ),
            "hypothesis_ids": [],
            "priority": round(float(observation["score"]) + 0.2, 3),
        }

    def _enrich_display_metadata(
        self,
        *,
        dataset_records: dict[str, dict[str, Any]],
        metric_records: dict[str, dict[str, Any]],
        formula_records: dict[str, dict[str, Any]],
        dimension_records: dict[str, dict[str, Any]],
        entity_records: dict[str, dict[str, Any]],
        lag_records: dict[str, dict[str, Any]],
        verdict_records: dict[str, dict[str, Any]],
    ) -> None:
        for dataset_key, record in dataset_records.items():
            display_name = DATASET_DISPLAY_LABELS.get(
                dataset_key,
                str(Path(record["filename"]).stem).replace("_", " ").strip().title(),
            )
            record["label"] = display_name
            record["display_name"] = display_name
            record["name"] = display_name

        for dimension_key, record in dimension_records.items():
            display_name = DIMENSION_DISPLAY_LABELS.get(
                dimension_key, record.get("label") or _titleize(dimension_key)
            )
            record["label"] = display_name
            record["display_name"] = display_name
            record["name"] = display_name

        for record in metric_records.values():
            display_name = str(record.get("label") or _titleize(record["code"])).strip()
            record["display_name"] = display_name
            record["name"] = display_name

        for record in entity_records.values():
            display_name = str(record.get("label") or record.get("value") or "").strip()
            record["display_name"] = display_name
            record["name"] = display_name

        for record in formula_records.values():
            metric_label = metric_records.get(record["metric_code"], {}).get(
                "label", _titleize(record["metric_code"])
            )
            dataset_label = dataset_records.get(record["dataset_key"], {}).get(
                "label", record["dataset_key"]
            )
            display_name = f"{metric_label} из {dataset_label}"
            record["label"] = display_name
            record["display_name"] = display_name
            record["name"] = display_name

        for record in lag_records.values():
            metric_label = metric_records.get(record["metric_code"], {}).get(
                "label", _titleize(record["metric_code"])
            )
            display_name = f"Лаг для {metric_label}"
            record["label"] = display_name
            record["display_name"] = display_name
            record["name"] = display_name

        for record in verdict_records.values():
            display_name = str(record.get("label") or _titleize(record["code"])).strip()
            record["display_name"] = display_name
            record["name"] = display_name

    def _canonicalize_rows(
        self,
        *,
        contract: dict[str, Any],
        headers: list[str],
        rows: list[dict[str, Any]],
    ) -> tuple[list[str], list[dict[str, Any]]]:
        required_columns = set(str(item) for item in contract.get("required_columns", []))
        canonical_headers = []
        for header in headers:
            canonical_header = _canonical_header(header)
            if contract.get("kind") == "numeric_facts" and canonical_header not in required_columns:
                metric_header = _canonical_metric_code(header)
                if metric_header in required_columns:
                    canonical_header = metric_header
            canonical_headers.append(canonical_header)
        header_map = dict(zip(headers, canonical_headers, strict=True))
        canonical_rows: list[dict[str, Any]] = []

        for row in rows:
            canonical_row: dict[str, Any] = {}
            for raw_header, value in row.items():
                header = header_map[raw_header]
                canonical_row[header] = value
                if contract.get("kind") == "numeric_facts" and header in {
                    "region",
                    "business_unit",
                    "warehouse",
                    "product_category",
                }:
                    canonical_value = _canonical_dimension_value(header, value)
                    canonical_row[header] = canonical_value
                    raw_value = str(value or "").strip()
                    if raw_value and raw_value != canonical_value:
                        canonical_row[f"__display__{header}"] = raw_value
            canonical_rows.append(canonical_row)

        return canonical_headers, canonical_rows

    def _validate_contract(
        self, *, path: Path, headers: list[str], contract: dict[str, Any]
    ) -> None:
        missing = [
            column for column in contract.get("required_columns", []) if column not in headers
        ]
        if missing:
            raise ValueError(f"{path.name} missing required columns: {', '.join(missing)}")
        expected_sheet = contract.get("sheet_name") or contract.get("sheet") or path.stem
        workbook = load_workbook(path, read_only=True)
        actual_sheet = (
            str(expected_sheet)
            if expected_sheet in workbook.sheetnames
            else workbook.sheetnames[0]
            if workbook.sheetnames
            else None
        )
        workbook.close()
        if not actual_sheet:
            raise ValueError(f"{path.name} must contain at least one sheet")
        if contract.get("sheet_name") and actual_sheet != str(contract["sheet_name"]):
            raise ValueError(f"{path.name} missing sheet: {contract['sheet_name']}")

    def _ingest_fact_dataset(
        self,
        *,
        contract: dict[str, Any],
        path: Path,
        rows: list[dict[str, Any]],
        dataset_records: dict[str, dict[str, Any]],
        metric_records: dict[str, dict[str, Any]],
        formula_records: dict[str, dict[str, Any]],
        dimension_records: dict[str, dict[str, Any]],
        entity_records: dict[str, dict[str, Any]],
    ) -> None:
        dataset_key = str(contract["key"])
        dimensions = list(contract.get("grain", []))
        metrics = list(contract.get("derives", {}).get("metrics", []))
        dataset_records[dataset_key] = {
            "key": dataset_key,
            "filename": path.name,
            "path": self._display_path(path),
            "row_count": len(rows),
            "grain": dimensions,
            "dimensions": dimensions,
            "metrics": metrics,
            "kind": contract.get("kind", "numeric_facts"),
        }

        for code in metrics:
            metric_records.setdefault(code, _default_metric_record(code, dataset_key))
            formula_id = f"{dataset_key}:{code}:v1"
            formula_records[formula_id] = {
                "id": formula_id,
                "metric_code": code,
                "dataset_key": dataset_key,
                "version": 1,
                "expression": f"xlsx:{dataset_key}::{code}",
                "notes": f"Inferred direct column mapping from {path.name}.",
            }

        for dimension_key in dimensions:
            dimension_records.setdefault(
                dimension_key,
                {
                    "key": dimension_key,
                    "label": _titleize(dimension_key),
                    "aliases": [dimension_key, dimension_key.replace("_", " ")],
                },
            )
            values = sorted(
                {
                    str(row[dimension_key]).strip()
                    for row in rows
                    if row.get(dimension_key) not in (None, "")
                }
            )
            for value in values:
                entity_id = f"{dimension_key}:{_normalize_code(value)}"
                display_value = str(
                    next(
                        (
                            row.get(f"__display__{dimension_key}")
                            for row in rows
                            if str(row.get(dimension_key)).strip() == value
                            and row.get(f"__display__{dimension_key}")
                        ),
                        ENTITY_DISPLAY_LABELS.get(dimension_key, {}).get(value, value),
                    )
                ).strip()
                aliases = [value, value.replace("_", " ")]
                if display_value and display_value not in aliases:
                    aliases.append(display_value)
                record = entity_records.setdefault(
                    entity_id,
                    {
                        "id": entity_id,
                        "dimension_key": dimension_key,
                        "value": value,
                        "label": display_value or value,
                        "aliases": aliases,
                        "dataset_keys": [],
                    },
                )
                if dataset_key not in record["dataset_keys"]:
                    record["dataset_keys"].append(dataset_key)

    def _display_path(self, path: Path) -> str:
        try:
            return str(path.relative_to(self.project_root))
        except ValueError:
            return str(path)

    def _ingest_metadata(
        self,
        *,
        contract: dict[str, Any],
        rows: list[dict[str, Any]],
        metric_records: dict[str, dict[str, Any]],
        dependency_records: dict[str, dict[str, Any]],
        lag_records: dict[str, dict[str, Any]],
        verdict_records: dict[str, dict[str, Any]],
    ) -> None:
        key = str(contract["key"])
        if key == "metric_dictionary":
            for row in rows:
                code = _canonical_metric_code(row["metric_code"])
                preferred_dataset = _canonical_dataset_key(row.get("preferred_dataset"))
                record = metric_records.setdefault(
                    code,
                    _default_metric_record(code, preferred_dataset),
                )
                aliases = _split_multi(row.get("aliases"))
                aliases.extend([code, code.replace("_", " "), _titleize(code).lower()])
                record.update(
                    {
                        "code": code,
                        "label": str(row.get("label") or record["label"]).strip(),
                        "description": str(row.get("description") or record["description"]).strip(),
                        "aliases": list(dict.fromkeys(alias for alias in aliases if alias)),
                        "sensitivity_level": _canonical_sensitivity(
                            row.get("sensitivity_level") or record["sensitivity_level"]
                        ),
                        "allow_roles": [
                            _canonical_role(role)
                            for role in str(row.get("allow_roles") or "").split(",")
                            if role.strip()
                        ]
                        or record["allow_roles"],
                        "preferred_dataset": preferred_dataset or record["preferred_dataset"],
                        "semantic_type": str(
                            row.get("semantic_type") or record.get("semantic_type") or "unknown"
                        ).strip(),
                        "source": "metadata",
                    }
                )
            return

        if key == "dependency_rules":
            for row in rows:
                source = _canonical_metric_code(row["source_metric_code"])
                target = _canonical_metric_code(row["target_metric_code"])
                dep_key = f"{source}->{target}"
                dependency_records[dep_key] = {
                    "id": dep_key,
                    "source_metric_code": source,
                    "target_metric_code": target,
                    "edge_type": _canonical_edge_type(row.get("edge_type") or "driver"),
                    "reason": str(row.get("reason") or "").strip(),
                    "strength": float(row.get("strength") or 0.0),
                    "source": "metadata",
                }
            return

        if key == "lag_rules":
            for row in rows:
                metric_code = _canonical_metric_code(row["metric_code"])
                lag_period = _canonical_lag_period(row.get("lag_period") or "previous_period")
                lag_id = f"{metric_code}:{lag_period}"
                lag_records[lag_id] = {
                    "id": lag_id,
                    "metric_code": metric_code,
                    "lag_period": lag_period,
                    "severity": _canonical_severity(row.get("severity") or "info"),
                    "rule_text": str(row.get("rule_text") or "").strip(),
                    "is_active": _parse_bool(row.get("is_active")),
                    "source": "metadata",
                }
            return

        if key == "verdict_rules":
            for row in rows:
                code = _canonical_verdict_code(row["code"])
                verdict_records[code] = {
                    "code": code,
                    "label": str(row.get("label") or _titleize(code)),
                    "metric_codes": [
                        _canonical_metric_code(item)
                        for item in _split_multi(row.get("metric_codes"))
                    ],
                    "verdict_text": str(row.get("verdict_text") or "").strip(),
                    "priority": int(row.get("priority") or 100),
                    "is_active": _parse_bool(row.get("is_active")),
                    "source": "metadata",
                }

    def _apply_dependency_heuristics(
        self,
        metric_records: dict[str, dict[str, Any]],
        dependency_records: dict[str, dict[str, Any]],
    ) -> None:
        heuristic_edges = [
            (
                "revenue",
                "gross_margin",
                "driver",
                "Gross margin depends on the revenue side of the same dimensional slice.",
                0.9,
            ),
            (
                "inventory_days",
                "storage_cost",
                "driver",
                "Storage cost typically rises with inventory days for the same slice.",
                0.8,
            ),
        ]
        available_metrics = set(metric_records)
        for source, target, edge_type, reason, strength in heuristic_edges:
            if source not in available_metrics or target not in available_metrics:
                continue
            dep_key = f"{source}->{target}"
            dependency_records.setdefault(
                dep_key,
                {
                    "id": dep_key,
                    "source_metric_code": source,
                    "target_metric_code": target,
                    "edge_type": edge_type,
                    "reason": reason,
                    "strength": strength,
                    "source": "heuristic",
                },
            )

    def _apply_dependency_priors(
        self,
        metric_records: dict[str, dict[str, Any]],
        dependency_records: dict[str, dict[str, Any]],
    ) -> None:
        available_metrics = set(metric_records)
        for prior in self.dependency_priors:
            source = _canonical_metric_code(prior.get("source_metric_code"))
            target = _canonical_metric_code(prior.get("target_metric_code"))
            if not source or not target:
                continue
            if source not in available_metrics and target not in available_metrics:
                continue
            dep_key = f"{source}->{target}"
            dependency_records[dep_key] = {
                "id": dep_key,
                "source_metric_code": source,
                "target_metric_code": target,
                "edge_type": _canonical_edge_type(prior.get("edge_type") or "driver"),
                "reason": str(
                    prior.get("note") or prior.get("reason") or "Confirmed user memory prior."
                ).strip(),
                "strength": float(prior.get("strength") or 1.0),
                "source": str(prior.get("source") or "memory_prior"),
            }

    def _infer_dependencies_from_facts(
        self,
        *,
        fact_contexts: list[dict[str, Any]],
        metric_records: dict[str, dict[str, Any]],
        dependency_records: dict[str, dict[str, Any]],
    ) -> None:
        if not fact_contexts:
            return

        merged_rows = self._build_merged_metric_rows(fact_contexts)
        available_metrics = set(metric_records)

        self._infer_component_dependencies(merged_rows, available_metrics, dependency_records)
        self._infer_margin_dependencies(merged_rows, available_metrics, dependency_records)
        self._infer_revenue_dependencies(merged_rows, available_metrics, dependency_records)
        self._infer_statistical_dependencies(merged_rows, available_metrics, dependency_records)

    def _build_merged_metric_rows(
        self, fact_contexts: list[dict[str, Any]]
    ) -> list[dict[str, float]]:
        all_dimensions = sorted(
            {
                dimension_key
                for context in fact_contexts
                for dimension_key in context.get("dimensions", [])
            }
        )
        merged_rows: dict[tuple[tuple[str, str], ...], dict[str, float]] = {}
        for context in fact_contexts:
            for row in context.get("rows", []):
                row_key = tuple(
                    (dimension_key, str(row.get(dimension_key) or ""))
                    for dimension_key in all_dimensions
                )
                record = merged_rows.setdefault(row_key, {})
                for metric_code in context.get("metrics", []):
                    numeric_value = _safe_float(row.get(metric_code))
                    if numeric_value is not None:
                        record[metric_code] = numeric_value
        return list(merged_rows.values())

    def _infer_component_dependencies(
        self,
        merged_rows: list[dict[str, float]],
        available_metrics: set[str],
        dependency_records: dict[str, dict[str, Any]],
    ) -> None:
        targets = [metric_code for metric_code in available_metrics if "total" in metric_code]
        for target_metric in targets:
            candidate_sources = sorted(
                metric_code for metric_code in available_metrics if metric_code != target_metric
            )
            for left_metric, right_metric in combinations(candidate_sources, 2):
                triples = [
                    (row[left_metric], row[right_metric], row[target_metric])
                    for row in merged_rows
                    if left_metric in row and right_metric in row and target_metric in row
                ]
                if len(triples) < 24:
                    continue
                error = _normalized_mae(
                    [left + right for left, right, _ in triples],
                    [target for _, _, target in triples],
                )
                if error is None or error > 0.01:
                    continue
                for source_metric in (left_metric, right_metric):
                    dep_key = f"{source_metric}->{target_metric}"
                    dependency_records.setdefault(
                        dep_key,
                        {
                            "id": dep_key,
                            "source_metric_code": source_metric,
                            "target_metric_code": target_metric,
                            "edge_type": "component",
                            "reason": (
                                f"Inferred from tabular data: {target_metric} is approximately "
                                f"the sum of {left_metric} and {right_metric} across aligned rows."
                            ),
                            "strength": 0.99,
                            "source": "inferred",
                        },
                    )

    def _infer_margin_dependencies(
        self,
        merged_rows: list[dict[str, float]],
        available_metrics: set[str],
        dependency_records: dict[str, dict[str, Any]],
    ) -> None:
        targets = [
            metric_code
            for metric_code in available_metrics
            if any(token in metric_code for token in ("margin", "profit"))
        ]
        for target_metric in targets:
            candidate_sources = sorted(
                metric_code for metric_code in available_metrics if metric_code != target_metric
            )
            for positive_metric, negative_metric in combinations(candidate_sources, 2):
                ordered_pairs = [
                    (positive_metric, negative_metric),
                    (negative_metric, positive_metric),
                ]
                for positive_candidate, negative_candidate in ordered_pairs:
                    triples = [
                        (row[positive_candidate], row[negative_candidate], row[target_metric])
                        for row in merged_rows
                        if positive_candidate in row
                        and negative_candidate in row
                        and target_metric in row
                    ]
                    if len(triples) < 24:
                        continue
                    error = _normalized_mae(
                        [positive - negative for positive, negative, _ in triples],
                        [target for _, _, target in triples],
                    )
                    if error is None or error > 0.01:
                        continue
                    positive_key = f"{positive_candidate}->{target_metric}"
                    negative_key = f"{negative_candidate}->{target_metric}"
                    dependency_records.setdefault(
                        positive_key,
                        {
                            "id": positive_key,
                            "source_metric_code": positive_candidate,
                            "target_metric_code": target_metric,
                            "edge_type": "driver",
                            "reason": (
                                f"Inferred from tabular data: {target_metric} behaves like "
                                f"{positive_candidate} minus {negative_candidate} across aligned rows."
                            ),
                            "strength": 0.98,
                            "source": "inferred",
                        },
                    )
                    dependency_records.setdefault(
                        negative_key,
                        {
                            "id": negative_key,
                            "source_metric_code": negative_candidate,
                            "target_metric_code": target_metric,
                            "edge_type": "inverse_driver",
                            "reason": (
                                f"Inferred from tabular data: {target_metric} behaves like "
                                f"{positive_candidate} minus {negative_candidate} across aligned rows."
                            ),
                            "strength": 0.98,
                            "source": "inferred",
                        },
                    )
                    break

    def _infer_statistical_dependencies(
        self,
        merged_rows: list[dict[str, float]],
        available_metrics: set[str],
        dependency_records: dict[str, dict[str, Any]],
    ) -> None:
        inference_templates = {
            "revenue": {
                "orders": ("driver", 0.55),
                "avg_price": ("driver", 0.45),
                "volume": ("driver", 0.55),
                "stockout_rate": ("inverse_driver", 0.30),
                "return_rate": ("inverse_driver", 0.30),
            },
            "return_rate": {
                "on_time_delivery_rate": ("inverse_driver", 0.35),
                "stockout_rate": ("driver", 0.25),
            },
            "freight_cost": {
                "volume": ("driver", 0.55),
            },
            "storage_cost": {
                "inventory_days": ("driver", 0.35),
                "volume": ("driver", 0.35),
            },
            "gross_margin": {
                "revenue": ("driver", 0.50),
                "total_cost": ("inverse_driver", 0.35),
            },
        }

        for target_metric, candidates in inference_templates.items():
            if target_metric not in available_metrics:
                continue
            for source_metric, (edge_type, min_strength) in candidates.items():
                if source_metric not in available_metrics:
                    continue
                dep_key = f"{source_metric}->{target_metric}"
                if dep_key in dependency_records:
                    continue
                pairs = [
                    (row[source_metric], row[target_metric])
                    for row in merged_rows
                    if source_metric in row and target_metric in row
                ]
                if len(pairs) < 24:
                    continue
                correlation = _pearson_correlation(
                    [left for left, _ in pairs],
                    [right for _, right in pairs],
                )
                if correlation is None:
                    continue
                if edge_type == "driver" and correlation <= 0:
                    continue
                if edge_type == "inverse_driver" and correlation >= 0:
                    continue
                strength = abs(correlation)
                if strength < min_strength:
                    continue
                dependency_records[dep_key] = {
                    "id": dep_key,
                    "source_metric_code": source_metric,
                    "target_metric_code": target_metric,
                    "edge_type": edge_type,
                    "reason": (
                        f"Inferred from aligned fact rows: correlation between {source_metric} and "
                        f"{target_metric} is {correlation:.3f}."
                    ),
                    "strength": round(min(0.95, strength), 3),
                    "source": "inferred",
                }

    def _infer_revenue_dependencies(
        self,
        merged_rows: list[dict[str, float]],
        available_metrics: set[str],
        dependency_records: dict[str, dict[str, Any]],
    ) -> None:
        if "revenue" not in available_metrics:
            return

        def set_revenue_dependency(
            source_metric: str,
            edge_type: str,
            correlation: float,
            reason: str,
        ) -> None:
            dep_key = f"{source_metric}->revenue"
            if dep_key in dependency_records:
                return
            dependency_records[dep_key] = {
                "id": dep_key,
                "source_metric_code": source_metric,
                "target_metric_code": "revenue",
                "edge_type": edge_type,
                "reason": reason,
                "strength": round(min(0.95, abs(correlation)), 3),
                "source": "inferred",
            }

        if {"orders", "avg_price", "revenue"}.issubset(available_metrics):
            pairs = [
                (row["orders"], row["revenue"] / max(row["avg_price"], 1e-9))
                for row in merged_rows
                if all(metric in row for metric in ("orders", "avg_price", "revenue"))
            ]
            correlation = (
                _pearson_correlation(
                    [left for left, _ in pairs],
                    [right for _, right in pairs],
                )
                if len(pairs) >= 24
                else None
            )
            if correlation is not None and correlation >= 0.6:
                set_revenue_dependency(
                    "orders",
                    "driver",
                    correlation,
                    (
                        "Inferred from normalized fact rows: revenue divided by avg_price "
                        f"moves with orders, correlation {correlation:.3f}."
                    ),
                )

            pairs = [
                (row["avg_price"], row["revenue"] / max(row["orders"], 1e-9))
                for row in merged_rows
                if all(metric in row for metric in ("orders", "avg_price", "revenue"))
            ]
            correlation = (
                _pearson_correlation(
                    [left for left, _ in pairs],
                    [right for _, right in pairs],
                )
                if len(pairs) >= 24
                else None
            )
            if correlation is not None and correlation >= 0.6:
                set_revenue_dependency(
                    "avg_price",
                    "driver",
                    correlation,
                    (
                        "Inferred from normalized fact rows: revenue divided by orders "
                        f"moves with avg_price, correlation {correlation:.3f}."
                    ),
                )

        if {"stockout_rate", "orders", "avg_price", "revenue"}.issubset(available_metrics):
            pairs = [
                (
                    row["stockout_rate"],
                    row["revenue"] / max(row["orders"] * row["avg_price"], 1e-9),
                )
                for row in merged_rows
                if all(
                    metric in row for metric in ("stockout_rate", "orders", "avg_price", "revenue")
                )
            ]
            correlation = (
                _pearson_correlation(
                    [left for left, _ in pairs],
                    [right for _, right in pairs],
                )
                if len(pairs) >= 24
                else None
            )
            if correlation is not None and correlation <= -0.3:
                set_revenue_dependency(
                    "stockout_rate",
                    "inverse_driver",
                    correlation,
                    (
                        "Inferred from normalized fact rows: revenue scaled by orders and avg_price "
                        f"falls as stockout_rate rises, correlation {correlation:.3f}."
                    ),
                )

        if {"return_rate", "orders", "avg_price", "revenue"}.issubset(available_metrics):
            pairs = [
                (
                    row["return_rate"],
                    row["revenue"] / max(row["orders"] * row["avg_price"], 1e-9),
                )
                for row in merged_rows
                if all(
                    metric in row for metric in ("return_rate", "orders", "avg_price", "revenue")
                )
            ]
            correlation = (
                _pearson_correlation(
                    [left for left, _ in pairs],
                    [right for _, right in pairs],
                )
                if len(pairs) >= 24
                else None
            )
            if correlation is not None and correlation <= -0.3:
                set_revenue_dependency(
                    "return_rate",
                    "inverse_driver",
                    correlation,
                    (
                        "Inferred from normalized fact rows: revenue scaled by orders and avg_price "
                        f"falls as return_rate rises, correlation {correlation:.3f}."
                    ),
                )

    def _apply_rule_fallbacks(
        self,
        metric_records: dict[str, dict[str, Any]],
        lag_records: dict[str, dict[str, Any]],
        verdict_records: dict[str, dict[str, Any]],
    ) -> None:
        if not lag_records:
            for metric_code in sorted(metric_records):
                lag_id = f"{metric_code}:previous_period"
                lag_records[lag_id] = {
                    "id": lag_id,
                    "metric_code": metric_code,
                    "lag_period": "previous_period",
                    "severity": "info",
                    "rule_text": f"Compare {metric_code} against previous_period before finalizing a trend statement.",
                    "is_active": True,
                    "source": "fallback",
                }

        if "ground_numeric_claims" not in verdict_records:
            verdict_records["ground_numeric_claims"] = {
                "code": "ground_numeric_claims",
                "label": "Ground numeric claims",
                "metric_codes": sorted(metric_records),
                "verdict_text": "Numeric claims must be grounded in selected metrics and the exact dimensional slice.",
                "priority": 1,
                "is_active": True,
                "source": "fallback",
            }


def write_json_report(path: str | Path, payload: dict[str, Any]) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def summarize_snapshot(
    manifest_path: str | Path, snapshot: RelationMemorySnapshot
) -> dict[str, Any]:
    payload = snapshot.to_dict()
    return {
        "manifest": str(manifest_path),
        "warnings": snapshot.warnings,
        "dataset_count": len(payload["datasets"]),
        "metric_count": len(payload["metrics"]),
        "formula_count": len(payload["formulas"]),
        "dimension_count": len(payload["dimensions"]),
        "entity_value_count": len(payload["entity_values"]),
        "dependency_count": len(payload["dependencies"]),
        "lag_rule_count": len(payload["lag_rules"]),
        "verdict_rule_count": len(payload["verdict_rules"]),
        "observation_count": len(payload["observations"]),
        "hypothesis_count": len(payload["hypotheses"]),
        "inquiry_question_count": len(payload["inquiry_questions"]),
    }


def write_graph_reports(
    reports_dir: str | Path,
    manifest_path: str | Path,
    snapshot: RelationMemorySnapshot,
) -> dict[str, Path]:
    reports_path = Path(reports_dir)
    snapshot_path = write_json_report(reports_path / "graph_snapshot.json", snapshot.to_dict())
    summary_path = write_json_report(
        reports_path / "graph_summary.json",
        summarize_snapshot(manifest_path, snapshot),
    )
    return {
        "graph_snapshot": snapshot_path,
        "graph_summary": summary_path,
    }
