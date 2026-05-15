from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from business_graph_core.parsers.base import ParsedCell, ParsedFile, ParsedSheet

DEPENDENCY_RULE_COLUMNS = {
    "source_metric_code",
    "target_metric_code",
    "edge_type",
    "reason",
    "strength",
}


class ExcelParser:
    """Parse .xlsx workbooks without executing macros."""

    def parse(self, path: Path, *, file_id: str | None = None) -> ParsedFile:
        if not path.exists():
            raise FileNotFoundError(path)
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            raise ValueError(f"Unsupported Excel extension: {path.suffix}")

        workbook = load_workbook(path, data_only=False, read_only=True)
        parsed_sheets: list[ParsedSheet] = []
        for sheet in workbook.worksheets:
            parsed_cells: list[ParsedCell] = []
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    formula = value if isinstance(value, str) and value.startswith("=") else None
                    parsed_cells.append(
                        ParsedCell(
                            sheet=sheet.title,
                            address=cell.coordinate,
                            value=value,
                            formula=formula,
                        )
                    )
            parsed_sheets.append(
                ParsedSheet(
                    name=sheet.title,
                    cells=parsed_cells,
                    dependency_rule_rows=self._extract_dependency_rules(sheet),
                )
            )
        workbook.close()

        return ParsedFile(
            file_id=file_id or f"file:{uuid4().hex}",
            source_name=path.name,
            source_type="xlsx",
            sheets=parsed_sheets,
        )

    def _extract_dependency_rules(self, sheet) -> list[dict[str, object]]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        header_index = None
        normalized_header: list[str] = []
        for idx, row in enumerate(rows):
            normalized = [str(value).strip() if value is not None else "" for value in row]
            lowered = [value.lower() for value in normalized]
            if DEPENDENCY_RULE_COLUMNS.issubset(set(lowered)):
                header_index = idx
                normalized_header = lowered
                break

        if header_index is None:
            return []

        result: list[dict[str, object]] = []
        for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = dict(zip(normalized_header, row, strict=False))
            if not values.get("source_metric_code") or not values.get("target_metric_code"):
                continue
            values["_sheet"] = sheet.title
            values["_row"] = offset
            result.append(values)
        return result
